#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
处理已下载的录音文件，进行转写和质检
"""

import os
from pathlib import Path
from datetime import datetime
import pandas as pd
from call_quality_check import CallQualityChecker, DASHSCOPE_API_KEY, RECORDINGS_DIR

def process_downloaded_recordings():
    """处理已下载的录音文件"""
    print("=" * 60)
    print("处理已下载的录音文件")
    print("=" * 60)
    
    # 查找所有录音文件
    recording_files = list(RECORDINGS_DIR.glob("*_2025-12-11.mp3"))
    
    if not recording_files:
        print("未找到录音文件")
        return
    
    print(f"找到 {len(recording_files)} 个录音文件")
    
    checker = CallQualityChecker()
    results = []
    
    for i, audio_path in enumerate(recording_files, 1):
        # 从文件名提取顾问姓名
        advisor_name = audio_path.stem.replace("_2025-12-11", "")
        
        print(f"\n--- 处理第 {i}/{len(recording_files)} 个文件: {advisor_name} ---")
        print(f"   文件: {audio_path.name}")
        
        # 转写音频
        print(f"   正在转写音频...")
        transcript = None
        ai_result = None
        
        if DASHSCOPE_API_KEY:
            transcript = checker.audio_to_text(audio_path)
            
            # AI质检
            if transcript:
                print(f"   正在AI质检...")
                ai_result = checker.ai_quality_check(advisor_name, transcript)
        else:
            transcript = "（未配置API密钥，跳过转写）"
            ai_result = "（未配置API密钥，跳过质检）"
        
        # 记录结果
        results.append({
            'advisor': advisor_name,
            'date': '2025-12-11',
            'filename': audio_path.name,
            'transcript': transcript,
            'ai_result': ai_result,
            'status': '完成' if transcript else '转写失败'
        })
        
        print(f"   ✓ 处理完成")
    
    # 生成报告
    print(f"\n正在生成Excel报告...")
    data = []
    for result in results:
        data.append({
            '顾问姓名': result['advisor'],
            '录音日期': result['date'],
            '音频文件名': result['filename'],
            '通话文字摘要': result['transcript'][:200] + '...' if result['transcript'] and len(result['transcript']) > 200 else (result['transcript'] or '无'),
            'AI评分与点评': result['ai_result'] or '质检失败',
            '状态': result['status']
        })
    
    df = pd.DataFrame(data)
    output_file = f"质检日报_2025-12-11_已处理.xlsx"
    df.to_excel(output_file, index=False, engine='openpyxl')
    
    # 美化Excel格式，确保内容完整显示
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        
        wb = load_workbook(output_file)
        ws = wb.active
        
        # 设置表头样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # 设置列宽
        column_widths = {
            'A': 12,  # 顾问姓名
            'B': 12,  # 录音日期
            'C': 30,  # 音频文件名
            'D': 50,  # 通话文字摘要
            'E': 80,  # AI评分与点评（最重要的列，设置较宽）
            'F': 10,  # 状态
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 设置所有数据行的格式
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                # 启用自动换行
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                # 设置垂直对齐为顶部，这样多行文本会从顶部开始显示
        
        # 设置行高（自动调整以适应内容）
        ws.row_dimensions[1].height = 25  # 表头行高
        
        # 保存格式化的Excel
        wb.save(output_file)
        print(f"   Excel格式已优化，AI质检结果将完整显示")
    except Exception as e:
        print(f"   优化Excel格式时出错: {e}，但文件已保存")
    
    print(f"   报告已生成: {output_file}")
    print("\n" + "=" * 60)
    print("处理完成！")
    print(f"成功处理: {sum(1 for r in results if r['status'] == '完成')} 个文件")
    print("=" * 60)

if __name__ == "__main__":
    process_downloaded_recordings()

