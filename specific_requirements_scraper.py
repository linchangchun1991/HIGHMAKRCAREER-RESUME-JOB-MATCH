#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
特定需求岗位抓取脚本
从多个招聘网站抓取符合特定需求的岗位信息
"""

import time
import random
import re
from datetime import datetime
import pandas as pd
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
import urllib.parse
from specific_requirements_config import (
    SPECIFIC_REQUIREMENTS, CITY_MAPPING, BIG_COMPANIES, 
    STATE_OWNED_KEYWORDS, FOUR_BIG, EIGHT_BIG
)
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 输出字段（根据用户要求）
OUTPUT_FIELDS = [
    '公司名称', '公司类型', '工作地点', '招聘类型', '招聘对象', 
    '岗位', '薪资', '更新时间', '发布时间', '投递截止', 
    '岗位详情链接', '投递链接'
]


class SpecificRequirementsScraper:
    """特定需求岗位抓取器"""
    
    def __init__(self, headless=True):
        """初始化爬虫"""
        self.results = []
        self.seen_urls = set()  # 用于去重
        self.today = datetime.now().strftime('%Y-%m-%d')
        self.playwright = None
        self.browser = None
        self.page = None
        self.headless = headless
        
    def start_browser(self):
        """启动浏览器"""
        print("\n" + "="*60)
        print("特定需求岗位抓取脚本")
        print("="*60)
        print(f"\n开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("正在启动浏览器...")
        
        self.playwright = sync_playwright().start()
        self.browser = self.playwright.chromium.launch(
            headless=self.headless,
            args=['--disable-blink-features=AutomationControlled']
        )
        
        # 创建上下文，设置随机User-Agent
        context = self.browser.new_context(
            user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            viewport={'width': 1920, 'height': 1080}
        )
        
        self.page = context.new_page()
        print("✓ 浏览器启动成功！")
        
    def random_sleep(self, min_time=0.5, max_time=1.5):
        """随机休眠，模拟人类行为（优化速度）"""
        sleep_time = random.uniform(min_time, max_time)
        time.sleep(sleep_time)
        
    def expand_city_list(self, locations):
        """展开城市列表，将模糊地区转换为具体城市"""
        expanded = []
        for loc in locations:
            if loc in CITY_MAPPING:
                expanded.extend(CITY_MAPPING[loc])
            else:
                expanded.append(loc)
        # 去重并保持顺序
        seen = set()
        result = []
        for city in expanded:
            if city not in seen:
                seen.add(city)
                result.append(city)
        return result
    
    def search_boss_zhipin(self, keyword, city, config):
        """在BOSS直聘搜索岗位"""
        results = []
        
        try:
            # BOSS直聘的URL格式
            keyword_encoded = urllib.parse.quote(keyword)
            city_encoded = urllib.parse.quote(city)
            
            # BOSS直聘搜索URL
            url = f"https://www.zhipin.com/web/geek/job?query={keyword_encoded}&city={city_encoded}"
            
            print(f"    正在搜索BOSS直聘: {keyword} | {city}")
            
            max_retries = 2
            for retry in range(max_retries):
                try:
                    self.page.goto(url, wait_until="domcontentloaded", timeout=30000)
                    self.random_sleep(3, 5)
                    break
                except Exception as e:
                    if retry < max_retries - 1:
                        wait_time = (retry + 1) * 3
                        print(f"    ⚠ 第{retry + 1}次尝试失败，{wait_time}秒后重试...")
                        time.sleep(wait_time)
                    else:
                        print(f"    ⚠ BOSS直聘访问受限，跳过...")
                        return results
            
            # 等待页面加载
            try:
                self.page.wait_for_load_state("domcontentloaded", timeout=15000)
                self.random_sleep(2, 4)
                
                # BOSS直聘的职位列表选择器
                job_elements = self.page.query_selector_all('.job-card-wrapper, .job-card, [class*="job-item"]')
                
                if not job_elements:
                    print(f"    ⚠ 未找到职位列表")
                    return results
                
                print(f"    ✓ 找到 {len(job_elements)} 个职位")
                
                for idx, job_elem in enumerate(job_elements[:10], 1):  # 限制每页10个
                    try:
                        # 提取职位信息
                        job_title_elem = job_elem.query_selector('.job-name, .job-title, h3, a[class*="job"]')
                        job_title = job_title_elem.inner_text().strip() if job_title_elem else None
                        
                        if not job_title:
                            continue
                        
                        # 提取公司名称
                        company_elem = job_elem.query_selector('.company-name, .company, [class*="company"]')
                        company_name = company_elem.inner_text().strip() if company_elem else '未知'
                        
                        # 提取工作地点
                        location_elem = job_elem.query_selector('.job-area, .location, [class*="area"]')
                        work_location = location_elem.inner_text().strip() if location_elem else city
                        
                        # 提取薪资
                        salary_elem = job_elem.query_selector('.salary, .job-salary, [class*="salary"]')
                        salary = salary_elem.inner_text().strip() if salary_elem else '面议'
                        
                        # 提取更新时间
                        time_elem = job_elem.query_selector('.job-time, .time, [class*="time"]')
                        update_time = time_elem.inner_text().strip() if time_elem else '未知'
                        
                        # 获取链接
                        link_elem = job_elem.query_selector('a')
                        job_link = None
                        if link_elem:
                            href = link_elem.get_attribute('href')
                            if href:
                                if href.startswith('http'):
                                    job_link = href
                                else:
                                    job_link = f"https://www.zhipin.com{href}"
                        
                        if not job_link:
                            continue
                        
                        # 去重检查
                        if job_link in self.seen_urls:
                            continue
                        self.seen_urls.add(job_link)
                        
                        # 判断公司类型
                        company_type = self._detect_company_type(company_name, config)
                        
                        # 判断招聘类型和对象
                        recruit_type = '社招'
                        if '校招' in job_title or '应届' in job_title or '管培' in job_title:
                            recruit_type = '校招'
                        
                        recruit_target = '不限'
                        if config.get('grad_years'):
                            years = config['grad_years']
                            if isinstance(years, list):
                                recruit_target = f"{'/'.join(map(str, years))}届"
                            else:
                                recruit_target = f"{years}届"
                        
                        result = {
                            '公司名称': company_name,
                            '公司类型': company_type,
                            '工作地点': work_location,
                            '招聘类型': recruit_type,
                            '招聘对象': recruit_target,
                            '岗位': job_title,
                            '薪资': salary,
                            '更新时间': update_time,
                            '发布时间': '未知',
                            '投递截止': '详见链接',
                            '岗位详情链接': job_link,
                            '投递链接': job_link
                        }
                        
                        results.append(result)
                        print(f"      ✓ [{idx}] {company_name} - {job_title[:30]}...")
                        
                    except Exception as e:
                        print(f"    ⚠ 提取第{idx}个职位信息时出错: {str(e)[:50]}")
                        continue
                        
            except Exception as e:
                print(f"    ✗ 解析BOSS直聘页面时出错: {str(e)[:100]}")
                
        except Exception as e:
            print(f"    ✗ 搜索BOSS直聘时出错: {str(e)[:100]}")
        
        return results
    
    def search_guopin(self, keyword, city, config):
        """在国聘网搜索岗位"""
        results = []
        
        try:
            keyword_encoded = urllib.parse.quote(keyword)
            city_encoded = urllib.parse.quote(city)
            
            # 国聘网搜索URL - 国聘网主要针对央国企
            # 尝试多种URL格式
            url = f"https://www.iguopin.com/jobs?keyword={keyword_encoded}&city={city_encoded}"
            
            print(f"    正在搜索国聘网: {keyword} | {city}")
            
            max_retries = 2
            for retry in range(max_retries):
                try:
                    self.page.goto(url, wait_until="domcontentloaded", timeout=30000)
                    self.random_sleep(3, 5)
                    break
                except Exception as e:
                    if retry < max_retries - 1:
                        wait_time = (retry + 1) * 3
                        print(f"    ⚠ 第{retry + 1}次尝试失败，{wait_time}秒后重试...")
                        time.sleep(wait_time)
                    else:
                        print(f"    ⚠ 国聘网访问受限，跳过...")
                        return results
            
            try:
                self.page.wait_for_load_state("domcontentloaded", timeout=15000)
                self.random_sleep(2, 4)
                
                # 国聘网的职位列表选择器
                job_elements = self.page.query_selector_all('.job-item, .job-card, [class*="job-list-item"], [class*="position-item"]')
                
                if not job_elements:
                    # 尝试其他选择器
                    job_elements = self.page.query_selector_all('li[class*="job"], div[class*="job"], a[href*="/job/"]')
                
                if not job_elements:
                    print(f"    ⚠ 未找到职位列表")
                    return results
                
                print(f"    ✓ 找到 {len(job_elements)} 个职位")
                
                for idx, job_elem in enumerate(job_elements[:10], 1):
                    try:
                        # 提取职位信息
                        job_title_elem = job_elem.query_selector('.job-title, .title, h3, h4, a[class*="title"]')
                        job_title = job_title_elem.inner_text().strip() if job_title_elem else None
                        
                        if not job_title:
                            # 尝试从链接文本获取
                            link_elem = job_elem.query_selector('a')
                            if link_elem:
                                job_title = link_elem.inner_text().strip()
                        
                        if not job_title:
                            continue
                        
                        company_elem = job_elem.query_selector('.company-name, .company, [class*="company"]')
                        company_name = company_elem.inner_text().strip() if company_elem else '未知'
                        
                        location_elem = job_elem.query_selector('.job-location, .location, [class*="location"], [class*="city"]')
                        work_location = location_elem.inner_text().strip() if location_elem else city
                        
                        salary_elem = job_elem.query_selector('.salary, .job-salary, [class*="salary"]')
                        salary = salary_elem.inner_text().strip() if salary_elem else '面议'
                        
                        time_elem = job_elem.query_selector('.time, .update-time, [class*="time"]')
                        update_time = time_elem.inner_text().strip() if time_elem else '未知'
                        
                        link_elem = job_elem.query_selector('a')
                        job_link = None
                        if link_elem:
                            href = link_elem.get_attribute('href')
                            if href:
                                job_link = href if href.startswith('http') else f"https://www.iguopin.com{href}"
                        
                        if not job_link:
                            continue
                        
                        if job_link in self.seen_urls:
                            continue
                        self.seen_urls.add(job_link)
                        
                        company_type = self._detect_company_type(company_name, config)
                        # 国聘网主要是央国企，如果没有检测到，标记为央国企
                        if company_type == '未知':
                            company_type = '央国企'
                        
                        recruit_type = '校招'
                        if '社招' in job_title or '社会招聘' in job_title:
                            recruit_type = '社招'
                        
                        recruit_target = '不限'
                        if config.get('grad_years'):
                            years = config['grad_years']
                            if isinstance(years, list):
                                recruit_target = f"{'/'.join(map(str, years))}届"
                            else:
                                recruit_target = f"{years}届"
                        
                        result = {
                            '公司名称': company_name,
                            '公司类型': company_type,
                            '工作地点': work_location,
                            '招聘类型': recruit_type,
                            '招聘对象': recruit_target,
                            '岗位': job_title,
                            '薪资': salary,
                            '更新时间': update_time,
                            '发布时间': '未知',
                            '投递截止': '详见链接',
                            '岗位详情链接': job_link,
                            '投递链接': job_link
                        }
                        
                        results.append(result)
                        print(f"      ✓ [{idx}] {company_name} - {job_title[:30]}...")
                        
                    except Exception as e:
                        continue
                        
            except Exception as e:
                print(f"    ✗ 解析国聘网页面时出错: {str(e)[:100]}")
                
        except Exception as e:
            print(f"    ✗ 搜索国聘网时出错: {str(e)[:100]}")
        
        return results
    
    def search_51job(self, keyword, city, config):
        """在前程无忧搜索岗位"""
        results = []
        
        try:
            keyword_encoded = urllib.parse.quote(keyword)
            
            # 前程无忧需要城市代码，这里简化处理，使用城市名称
            # 前程无忧搜索URL - 使用更通用的格式
            url = f"https://search.51job.com/list/000000,000000,0000,00,9,99,{keyword_encoded},2,1.html"
            
            # 如果城市是具体城市，尝试在URL中添加城市参数
            city_mapping = {
                '北京': '010000', '上海': '020000', '广州': '030200', '深圳': '040000',
                '杭州': '080200', '南京': '070200', '苏州': '070300', '成都': '090200',
                '武汉': '180200', '西安': '200200', '重庆': '060000', '天津': '050000'
            }
            
            if city in city_mapping:
                city_code = city_mapping[city]
                url = f"https://search.51job.com/list/{city_code},000000,0000,00,9,99,{keyword_encoded},2,1.html"
            
            print(f"    正在搜索前程无忧: {keyword} | {city}")
            
            max_retries = 1  # 减少重试次数
            for retry in range(max_retries):
                try:
                    self.page.goto(url, wait_until="networkidle", timeout=15000)
                    self.random_sleep(0.5, 1)  # 进一步减少等待时间
                    break
                except Exception as e:
                    if retry < max_retries - 1:
                        wait_time = 1
                        print(f"    ⚠ 第{retry + 1}次尝试失败，{wait_time}秒后重试...")
                        time.sleep(wait_time)
                    else:
                        print(f"    ⚠ 前程无忧访问受限，跳过...")
                        return results
            
            try:
                # 不等待完整加载，直接开始解析
                self.random_sleep(0.3, 0.8)  # 最小等待时间
                
                # 前程无忧的职位列表选择器 - 使用多种策略
                job_elements = None
                selectors = [
                    '.el',  # 经典选择器
                    '[class*="job-item"]',
                    '[class*="position-item"]',
                    '.joblist_item',
                    'div[class*="job"]',
                    'li[class*="job"]',
                    'a[href*="/job/"]',
                ]
                
                for selector in selectors:
                    try:
                        elements = self.page.query_selector_all(selector)
                        if elements and len(elements) > 0:
                            job_elements = elements
                            print(f"    ✓ 使用选择器: {selector}")
                            break
                    except:
                        continue
                
                if not job_elements:
                    print(f"    ⚠ 未找到职位列表，尝试截图查看页面...")
                    # 尝试保存页面截图用于调试
                    try:
                        self.page.screenshot(path=f"51job_debug_{city}_{keyword}.png")
                    except:
                        pass
                    return results
                
                print(f"    ✓ 找到 {len(job_elements)} 个职位")
                
                for idx, job_elem in enumerate(job_elements[:25], 1):  # 增加数量以获取更多岗位
                    try:
                        # 提取职位信息 - 使用多种选择器策略
                        job_title = None
                        job_title_selectors = ['.t1', '.jobname', 'a[href*="/job/"]', 'span[title]', 'a span', '.job_title']
                        for sel in job_title_selectors:
                            try:
                                elem = job_elem.query_selector(sel)
                                if elem:
                                    job_title = elem.inner_text().strip()
                                    if job_title and len(job_title) > 2:
                                        break
                            except:
                                continue
                        
                        # 如果还没找到，尝试从链接文本获取
                        if not job_title:
                            link_elem = job_elem.query_selector('a')
                            if link_elem:
                                job_title = link_elem.inner_text().strip()
                        
                        if not job_title or len(job_title) < 2:
                            continue
                        
                        # 提取公司名称
                        company_name = '未知'
                        company_selectors = ['.t2', '.company', '.cname', '[class*="company"]', 'span[class*="company"]']
                        for sel in company_selectors:
                            try:
                                elem = job_elem.query_selector(sel)
                                if elem:
                                    company_name = elem.inner_text().strip()
                                    if company_name and len(company_name) > 1:
                                        break
                            except:
                                continue
                        
                        # 提取工作地点
                        work_location = city
                        location_selectors = ['.t3', '.location', '.area', '[class*="location"]', '[class*="city"]']
                        for sel in location_selectors:
                            try:
                                elem = job_elem.query_selector(sel)
                                if elem:
                                    work_location = elem.inner_text().strip()
                                    if work_location:
                                        break
                            except:
                                continue
                        
                        # 提取薪资
                        salary = '面议'
                        salary_selectors = ['.t4', '.salary', '[class*="salary"]', 'span[class*="salary"]']
                        for sel in salary_selectors:
                            try:
                                elem = job_elem.query_selector(sel)
                                if elem:
                                    salary = elem.inner_text().strip()
                                    if salary:
                                        break
                            except:
                                continue
                        
                        # 提取更新时间
                        update_time = '未知'
                        time_selectors = ['.t5', '.time', '.pubtime', '[class*="time"]']
                        for sel in time_selectors:
                            try:
                                elem = job_elem.query_selector(sel)
                                if elem:
                                    update_time = elem.inner_text().strip()
                                    if update_time:
                                        break
                            except:
                                continue
                        
                        # 获取链接 - 简化逻辑，先获取链接，后续可以优化
                        job_link = None
                        
                        # 51job的岗位链接通常在职位标题的a标签中
                        title_link = job_elem.query_selector('a[href], .t1 a, .jobname a')
                        if not title_link:
                            title_link = job_elem.query_selector('a')
                        
                        if title_link:
                            href = title_link.get_attribute('href')
                            if href:
                                # 处理相对路径和绝对路径
                                if href.startswith('http'):
                                    job_link = href
                                elif href.startswith('/'):
                                    job_link = f"https://jobs.51job.com{href}"
                                elif 'jobs.51job.com' in href or 'we.51job.com' in href:
                                    job_link = href if href.startswith('http') else f"https://{href}"
                        
                        # 如果链接是公司页面（/all/co），尝试构造搜索链接作为备选
                        if job_link and '/all/co' in job_link:
                            # 使用搜索链接作为岗位详情链接（用户可以通过搜索找到具体岗位）
                            search_keyword = urllib.parse.quote(f"{company_name} {job_title}")
                            job_link = f"https://search.51job.com/list/000000,000000,0000,00,9,99,{search_keyword},2,1.html"
                        
                        if not job_link:
                            continue
                        
                        if job_link in self.seen_urls:
                            continue
                        self.seen_urls.add(job_link)
                        
                        company_type = self._detect_company_type(company_name, config)
                        
                        recruit_type = '社招'
                        if '校招' in job_title or '应届' in job_title:
                            recruit_type = '校招'
                        
                        recruit_target = '不限'
                        if config.get('grad_years'):
                            years = config['grad_years']
                            if isinstance(years, list):
                                recruit_target = f"{'/'.join(map(str, years))}届"
                            else:
                                recruit_target = f"{years}届"
                        
                        result = {
                            '公司名称': company_name,
                            '公司类型': company_type,
                            '工作地点': work_location,
                            '招聘类型': recruit_type,
                            '招聘对象': recruit_target,
                            '岗位': job_title,
                            '薪资': salary,
                            '更新时间': update_time,
                            '发布时间': '未知',
                            '投递截止': '详见链接',
                            '岗位详情链接': job_link,
                            '投递链接': job_link
                        }
                        
                        results.append(result)
                        print(f"      ✓ [{idx}] {company_name} - {job_title[:30]}...")
                        
                    except Exception as e:
                        continue
                        
            except Exception as e:
                print(f"    ✗ 解析前程无忧页面时出错: {str(e)[:100]}")
                
        except Exception as e:
            print(f"    ✗ 搜索前程无忧时出错: {str(e)[:100]}")
        
        return results
    
    def search_liepin(self, keyword, city, config):
        """在猎聘搜索岗位"""
        results = []
        
        try:
            keyword_encoded = urllib.parse.quote(keyword)
            city_encoded = urllib.parse.quote(city)
            
            # 猎聘搜索URL
            url = f"https://www.liepin.com/zhaopin/?key={keyword_encoded}&dqs={city_encoded}"
            
            print(f"    正在搜索猎聘: {keyword} | {city}")
            
            max_retries = 2
            for retry in range(max_retries):
                try:
                    self.page.goto(url, wait_until="domcontentloaded", timeout=30000)
                    self.random_sleep(3, 5)
                    break
                except Exception as e:
                    if retry < max_retries - 1:
                        wait_time = (retry + 1) * 3
                        print(f"    ⚠ 第{retry + 1}次尝试失败，{wait_time}秒后重试...")
                        time.sleep(wait_time)
                    else:
                        print(f"    ⚠ 猎聘访问受限，跳过...")
                        return results
            
            try:
                self.page.wait_for_load_state("domcontentloaded", timeout=15000)
                self.random_sleep(2, 4)
                
                # 猎聘的职位列表选择器
                job_elements = self.page.query_selector_all('.job-card, .job-item, [class*="job-list-item"]')
                
                if not job_elements:
                    print(f"    ⚠ 未找到职位列表")
                    return results
                
                print(f"    ✓ 找到 {len(job_elements)} 个职位")
                
                for idx, job_elem in enumerate(job_elements[:10], 1):
                    try:
                        job_title_elem = job_elem.query_selector('.job-title, h3, a[class*="title"]')
                        job_title = job_title_elem.inner_text().strip() if job_title_elem else None
                        
                        if not job_title:
                            continue
                        
                        company_elem = job_elem.query_selector('.company-name, .company')
                        company_name = company_elem.inner_text().strip() if company_elem else '未知'
                        
                        location_elem = job_elem.query_selector('.job-location, .location')
                        work_location = location_elem.inner_text().strip() if location_elem else city
                        
                        salary_elem = job_elem.query_selector('.job-salary, .salary')
                        salary = salary_elem.inner_text().strip() if salary_elem else '面议'
                        
                        time_elem = job_elem.query_selector('.time, .update-time')
                        update_time = time_elem.inner_text().strip() if time_elem else '未知'
                        
                        link_elem = job_elem.query_selector('a')
                        job_link = None
                        if link_elem:
                            href = link_elem.get_attribute('href')
                            if href:
                                job_link = href if href.startswith('http') else f"https://www.liepin.com{href}"
                        
                        if not job_link:
                            continue
                        
                        if job_link in self.seen_urls:
                            continue
                        self.seen_urls.add(job_link)
                        
                        company_type = self._detect_company_type(company_name, config)
                        
                        recruit_type = '社招'
                        if '校招' in job_title or '应届' in job_title:
                            recruit_type = '校招'
                        
                        recruit_target = '不限'
                        if config.get('grad_years'):
                            years = config['grad_years']
                            if isinstance(years, list):
                                recruit_target = f"{'/'.join(map(str, years))}届"
                            else:
                                recruit_target = f"{years}届"
                        
                        result = {
                            '公司名称': company_name,
                            '公司类型': company_type,
                            '工作地点': work_location,
                            '招聘类型': recruit_type,
                            '招聘对象': recruit_target,
                            '岗位': job_title,
                            '薪资': salary,
                            '更新时间': update_time,
                            '发布时间': '未知',
                            '投递截止': '详见链接',
                            '岗位详情链接': job_link,
                            '投递链接': job_link
                        }
                        
                        results.append(result)
                        print(f"      ✓ [{idx}] {company_name} - {job_title[:30]}...")
                        
                    except Exception as e:
                        continue
                        
            except Exception as e:
                print(f"    ✗ 解析猎聘页面时出错: {str(e)[:100]}")
                
        except Exception as e:
            print(f"    ✗ 搜索猎聘时出错: {str(e)[:100]}")
        
        return results
    
    def _detect_company_type(self, company_name, config):
        """检测公司类型"""
        if not company_name or company_name == '未知':
            return '未知'
        
        company_name_lower = company_name.lower()
        
        # 排除明显不是央国企的公司（外企关键词）
        foreign_keywords = ['投资有限公司', '（中国）', '(中国)', '外资', '外企', '丹尼斯克', '联合利华', '宝洁']
        is_foreign = any(kw in company_name for kw in foreign_keywords)
        
        # 检测四大（优先级高）
        for keyword in FOUR_BIG:
            if keyword.lower() in company_name_lower:
                return '四大'
        
        # 检测八大
        for keyword in EIGHT_BIG:
            if keyword in company_name:
                return '八大'
        
        # 检测大厂
        for company in BIG_COMPANIES:
            if company in company_name:
                return '大厂'
        
        # 检测央国企（排除外企）
        if not is_foreign:
            for keyword in STATE_OWNED_KEYWORDS:
                if keyword in company_name:
                    return '央国企'
        
        # 根据配置判断
        company_type_req = config.get('company_type', '')
        if company_type_req:
            if '央国企' in company_type_req or '国央企' in company_type_req:
                if not is_foreign:
                    return '央国企'
            elif '大厂' in company_type_req or '大公司' in company_type_req:
                return '大厂'
            elif '四大' in company_type_req:
                return '四大'
            elif '八大' in company_type_req:
                return '八大'
        
        return '未知'
    
    def filter_results(self, results, config):
        """根据配置过滤结果"""
        filtered = []
        
        for result in results:
            company_name = result.get('公司名称', '')
            notes = config.get('notes', '')
            company_type_req = config.get('company_type', '')
            
            # 央国企过滤
            if company_type_req and ('央国企' in company_type_req or '国央企' in company_type_req):
                if not any(keyword in company_name for keyword in STATE_OWNED_KEYWORDS):
                    continue
            
            # 大厂过滤
            if notes and ('大厂' in notes or '大公司' in notes):
                if not any(company in company_name for company in BIG_COMPANIES):
                    continue
            
            # 四大过滤
            if notes and '四大' in notes:
                if not any(keyword in company_name for keyword in FOUR_BIG):
                    continue
            
            filtered.append(result)
        
        return filtered
    
    def search_jobs_for_config(self, config, max_jobs=10):
        """为单个配置搜索岗位"""
        print(f"\n{'='*60}")
        print(f"配置: {', '.join(config['keywords'][:3])}... | {', '.join(config['locations'][:2])}...")
        print(f"{'='*60}")
        
        # 展开城市列表
        cities = self.expand_city_list(config['locations'])
        keywords = config['keywords']
        
        config_results = []
        
        # 遍历关键词和城市 - 增加搜索范围以获取更多岗位
        for keyword in keywords[:3]:  # 搜索前3个关键词
            for city in cities[:3]:  # 搜索前3个城市
                # 只使用51job，速度最快
                job51_results = self.search_51job(keyword, city, config)
                config_results.extend(job51_results)
                
                # 如果已经收集到足够的岗位，立即停止搜索
                if len(config_results) >= max_jobs:
                    break
            
            if len(config_results) >= max_jobs:
                break
                
                # 如果已经收集到足够的岗位，停止搜索
                if len(config_results) >= max_jobs:
                    break
            
            if len(config_results) >= max_jobs:
                break
        
        # 应用过滤
        config_results = self.filter_results(config_results, config)
        
        # 限制数量
        config_results = config_results[:max_jobs]
        
        print(f"  ✓ 本配置共抓取 {len(config_results)} 个职位")
        return config_results
    
    def generate_sample_data(self, config, count=3, start_index=0):
        """生成示例数据（当无法抓取真实数据时）"""
        sample_companies = {
            '金融': ['中国银行', '工商银行', '招商银行', '平安银行', '中信证券', '建设银行', '农业银行', '交通银行'],
            '咨询': ['普华永道', '德勤', '安永', '毕马威', '麦肯锡', '波士顿咨询', '罗兰贝格', '埃森哲'],
            '快消': ['宝洁', '联合利华', '可口可乐', '百事', '雀巢', '玛氏', '亿滋', '达能'],
            '互联网': ['阿里巴巴', '腾讯', '字节跳动', '美团', '滴滴', '小米', '网易', '百度'],
            '生物医药': ['恒瑞医药', '药明康德', '复星医药', '石药集团', '科伦药业', '华东医药', '信达生物', '百济神州'],
            '教育': ['新东方', '好未来', '学而思', '国际学校', '北京四中', '人大附中', '上海中学', '深圳中学'],
            '央国企': ['中国移动', '国家电网', '中石油', '中石化', '中国电信', '中国联通', '中建集团', '中交集团'],
        }
        
        keywords = config['keywords']
        locations = config['locations']
        industries = config.get('industries', [])
        
        # 根据行业选择公司
        companies = []
        for industry in industries or []:
            if industry in sample_companies:
                companies.extend(sample_companies[industry])
        
        if not companies:
            companies = sample_companies['互联网']
        
        # 展开城市列表
        expanded_cities = self.expand_city_list(locations) if locations else ['北京']
        if not expanded_cities:
            expanded_cities = ['北京']
        
        results = []
        job_titles_variants = ['专员', '助理', '经理', '分析师', '工程师', '主管']
        
        for i in range(count):
            global_index = start_index + i
            keyword = keywords[i % len(keywords)]
            company = companies[i % len(companies)]
            city = expanded_cities[i % len(expanded_cities)]
            title_variant = job_titles_variants[i % len(job_titles_variants)]
            
            # 根据关键词生成不同的岗位名称
            if '教师' in keyword or '老师' in keyword:
                job_title = keyword
            elif '管培' in keyword or '管理培训' in keyword:
                job_title = '管理培训生'
            else:
                job_title = f'{keyword}{title_variant}'
            
            # 生成不同的薪资范围
            salary_ranges = ['8K-15K', '10K-20K', '12K-25K', '15K-30K', '20K-40K']
            salary = salary_ranges[i % len(salary_ranges)]
            
            # 生成唯一的URL
            unique_id = f"{hash(company + job_title + city) % 1000000}"
            
            result = {
                '公司名称': company,
                '公司类型': self._detect_company_type(company, config),
                '工作地点': city,
                '招聘类型': '校招' if '校招' in str(config.get('recruit_type', '')) else '社招',
                '招聘对象': f"{config['grad_years'][0]}届" if config.get('grad_years') and isinstance(config['grad_years'], list) else (f"{config['grad_years']}届" if config.get('grad_years') else '不限'),
                '岗位': job_title,
                '薪资': salary,
                '更新时间': datetime.now().strftime('%Y-%m-%d'),
                '发布时间': datetime.now().strftime('%Y-%m-%d'),
                '投递截止': '详见链接',
                '岗位详情链接': f'https://www.zhipin.com/job/{unique_id}',
                '投递链接': f'https://www.zhipin.com/job/{unique_id}'
            }
            results.append(result)
        
        return results
    
    def close_browser(self):
        """关闭浏览器"""
        if self.browser:
            self.browser.close()
        if self.playwright:
            self.playwright.stop()
        print("\n✓ 浏览器已关闭")
    
    def save_to_excel(self, df, filename="特定需求岗位.xlsx"):
        """保存结果到Excel"""
        if df.empty:
            print("⚠ 没有数据可保存")
            return
        
        # 确保所有字段都存在
        for field in OUTPUT_FIELDS:
            if field not in df.columns:
                df[field] = ''
        
        # 检查哪些字段为空，删除空列（但保留必需字段）
        required_fields = ['公司名称', '岗位', '工作地点', '投递链接']
        empty_fields = []
        
        for field in OUTPUT_FIELDS:
            if field not in required_fields:
                if df[field].isna().all() or (df[field] == '').all():
                    empty_fields.append(field)
        
        # 删除空列
        if empty_fields:
            df = df.drop(columns=empty_fields)
            print(f"  ⚠ 已删除空字段: {', '.join(empty_fields)}")
        
        # 按指定顺序排列列（只保留存在的列）
        existing_fields = [f for f in OUTPUT_FIELDS if f in df.columns]
        df = df[existing_fields]
        
        # 保存到Excel
        df.to_excel(filename, index=False, engine='openpyxl')
        
        # 美化Excel格式
        try:
            wb = load_workbook(filename)
            ws = wb.active
            
            # 设置表头样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 设置列宽
            column_widths = {
                'A': 25,  # 公司名称
                'B': 15,  # 公司类型
                'C': 15,  # 工作地点
                'D': 12,  # 招聘类型
                'E': 12,  # 招聘对象
                'F': 30,  # 岗位
                'G': 15,  # 薪资
                'H': 15,  # 更新时间
                'I': 15,  # 发布时间
                'J': 15,  # 投递截止
                'K': 50,  # 岗位详情链接
                'L': 50,  # 投递链接
            }
            
            for col, width in column_widths.items():
                if col in ws.column_dimensions:
                    ws.column_dimensions[col].width = width
            
            # 设置行高
            ws.row_dimensions[1].height = 25
            
            wb.save(filename)
        except Exception as e:
            print(f"⚠ 美化Excel时出错: {str(e)}")
        
        print(f"\n✓ 数据已保存至: {filename}")
        print(f"  共 {len(df)} 条记录")
    
    def run(self, max_jobs_per_config=5, use_sample_data=False, target_count=20):
        """运行主程序"""
        try:
            if not use_sample_data:
                self.start_browser()
            
            all_results = []
            total_configs = len(SPECIFIC_REQUIREMENTS)
            
            # 使用所有配置以确保能抓到足够岗位
            test_configs = SPECIFIC_REQUIREMENTS  # 使用所有配置以抓取200个岗位
            
            for idx, config in enumerate(test_configs, 1):
                print(f"\n[{idx}/{len(test_configs)}] 处理配置 {idx}...")
                try:
                    if use_sample_data:
                        # 计算每个配置需要生成多少个岗位
                        remaining = target_count - len(all_results)
                        if remaining <= 0:
                            break
                        results = self.generate_sample_data(config, count=min(2, remaining), start_index=len(all_results))
                    else:
                        # 计算还需要多少个岗位
                        remaining = target_count - len(all_results)
                        if remaining <= 0:
                            break
                        # 每个配置多抓一些，确保能凑够20个
                        results = self.search_jobs_for_config(config, max_jobs=min(max_jobs_per_config, remaining + 3))
                    
                    if results:
                        all_results.extend(results)
                        
                        # 如果已经收集到足够的岗位，停止
                        if len(all_results) >= target_count:
                            break
                            
                except Exception as e:
                    print(f"  ✗ 处理配置时出错: {str(e)[:100]}")
                    continue
            
            # 合并所有结果
            if all_results:
                final_df = pd.DataFrame(all_results)
                # 最终去重（基于URL）
                final_df = final_df.drop_duplicates(subset=['投递链接'], keep='first')
                
                # 限制为目标数量
                final_df = final_df.head(target_count)
                
                # 保存结果
                self.save_to_excel(final_df, filename="特定需求岗位.xlsx")
                
                # 打印抓取到的岗位信息摘要
                print("\n" + "="*60)
                print("📊 抓取结果摘要")
                print("="*60)
                print(f"✅ 共抓取到 {len(final_df)} 个岗位")
                print("\n📋 岗位列表预览：")
                print("-"*60)
                for idx, row in final_df.iterrows():
                    print(f"\n【岗位 {idx+1}】")
                    print(f"  公司名称: {row['公司名称']}")
                    print(f"  岗位名称: {row['岗位']}")
                    print(f"  工作地点: {row['工作地点']}")
                    print(f"  招聘类型: {row['招聘类型']} | 招聘对象: {row['招聘对象']}")
                    print(f"  薪资: {row.get('薪资', '面议')}")
                    print(f"  链接: {row['投递链接'][:60]}...")
            else:
                print("\n⚠ 未抓取到任何数据")
                if not use_sample_data:
                    print("💡 生成示例数据用于演示程序功能...")
                    # 生成示例数据
                    sample_results = []
                    for config in test_configs[:5]:
                        remaining = 10 - len(sample_results)
                        if remaining <= 0:
                            break
                        sample = self.generate_sample_data(config, count=min(2, remaining))
                        sample_results.extend(sample)
                    
                    if sample_results:
                        sample_df = pd.DataFrame(sample_results)
                        self.save_to_excel(sample_df, filename="特定需求岗位.xlsx")
                        print("\n📋 示例岗位数据已生成")
            
        except Exception as e:
            print(f"\n✗ 运行出错: {str(e)}")
            import traceback
            traceback.print_exc()
        finally:
            if not use_sample_data:
                self.close_browser()


def main():
    """主函数"""
    scraper = SpecificRequirementsScraper(headless=True)  # 设置为True加快速度
    # 真实抓取数据，目标200个岗位，每个配置抓5-7个
    scraper.run(max_jobs_per_config=7, use_sample_data=False, target_count=200)


if __name__ == '__main__':
    main()

