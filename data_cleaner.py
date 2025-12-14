import openpyxl
import re
from datetime import datetime

# 读取原始文件
wb_source = openpyxl.load_workbook('/Users/changchun/Documents/trae_projects/111/enhanced_aceoffer_jobs_20251204_153445.xlsx')
ws_source = wb_source['Sheet1']

# 创建新的工作簿
wb_target = openpyxl.load_workbook('/Users/changchun/Desktop/规范职位投递表_20251204.xlsx')
ws_target = wb_target.active
ws_target.title = '职位信息表'

# 设置表头
headers = ['序号', '岗位ID', '岗位名称', '公司名称', '招聘类型', '行业', '工作地点', '薪资', '发布日期', '岗位要求', '岗位描述', '投递链接']
for col_idx, header in enumerate(headers, 1):
    ws_target.cell(row=1, column=col_idx, value=header)

# 处理数据
processed_jobs = {}  # 用于去重，key为岗位ID
row_num = 2

for row_idx in range(2, ws_source.max_row + 1):
    job_id = ws_source.cell(row=row_idx, column=2).value  # B列：岗位ID
    job_name = ws_source.cell(row=row_idx, column=4).value  # D列：岗位名称
    recruit_type_raw = ws_source.cell(row=row_idx, column=3).value  # C列
    recruit_type = ws_source.cell(row=row_idx, column=5).value  # E列：招聘类型
    location_raw = ws_source.cell(row=row_idx, column=7).value  # G列
    salary = ws_source.cell(row=row_idx, column=8).value  # H列：薪资
    publish_date = ws_source.cell(row=row_idx, column=9).value  # I列：发布日期
    requirements = ws_source.cell(row=row_idx, column=10).value  # J列：岗位要求
    description = ws_source.cell(row=row_idx, column=11).value  # K列：岗位描述
    apply_link = ws_source.cell(row=row_idx, column=14).value  # N列：真实投递链接
    
    # 跳过没有岗位ID或岗位名称的行
    if not job_id or not job_name:
        continue
    
    # 去重：如果已经处理过这个岗位ID，跳过
    if job_id in processed_jobs:
        continue
    
    # 从描述中提取公司名称、行业、地点等信息
    company_name = ''
    industry = ''
    location = ''
    company_type = ''
    
    if description:
        # 解析格式如："制造 | 央/国企 | 立即投递" 或 "银行 | 央/国企 | 临沂、宁波、日照、潍坊 | 立即投递"
        parts = description.split('|')
        if len(parts) >= 2:
            industry = parts[0].strip()
            company_type = parts[1].strip()
            if len(parts) >= 3 and '立即投递' not in parts[2]:
                location = parts[2].strip()
            elif len(parts) >= 4:
                location = parts[2].strip()
    
    # 如果location_raw有值且location为空，使用location_raw
    if location_raw and not location:
        location = location_raw
    
    # 从岗位名称中提取公司名称（通常在岗位名称开头）
    if job_name:
        # 尝试提取公司名称（通常是第一个中文词组）
        match = re.match(r'^([^0-9]+?(?:公司|集团|银行|中心|医院|大学|投资|科技|游戏|基金))', job_name)
        if match:
            company_name = match.group(1)
        else:
            # 如果没有匹配到，尝试提取前面的部分
            match = re.match(r'^(.+?)(?:20\d{2}|招聘|校招)', job_name)
            if match:
                company_name = match.group(1).strip()
    
    # 确定招聘批次（秋招正式批、秋招补录等）
    batch = recruit_type_raw if recruit_type_raw else ''
    
    # 写入新表格
    ws_target.cell(row=row_num, column=1, value=row_num - 1)  # 序号
    ws_target.cell(row=row_num, column=2, value=job_id)  # 岗位ID
    ws_target.cell(row=row_num, column=3, value=job_name)  # 岗位名称
    ws_target.cell(row=row_num, column=4, value=company_name)  # 公司名称
    ws_target.cell(row=row_num, column=5, value=f"{batch} - {recruit_type}" if batch and recruit_type else (batch or recruit_type or ''))  # 招聘类型
    ws_target.cell(row=row_num, column=6, value=f"{industry} | {company_type}" if industry and company_type else (industry or company_type or ''))  # 行业
    ws_target.cell(row=row_num, column=7, value=location)  # 工作地点
    ws_target.cell(row=row_num, column=8, value=salary if salary else '面议')  # 薪资
    ws_target.cell(row=row_num, column=9, value=publish_date)  # 发布日期
    ws_target.cell(row=row_num, column=10, value=requirements)  # 岗位要求
    ws_target.cell(row=row_num, column=11, value=description)  # 岗位描述
    ws_target.cell(row=row_num, column=12, value=apply_link)  # 投递链接
    
    processed_jobs[job_id] = True
    row_num += 1

# 保存文件
wb_target.save('/Users/changchun/Desktop/规范职位投递表_20251204.xlsx')
print(f'成功处理 {row_num - 2} 条职位信息')
print('文件已保存至：/Users/changchun/Desktop/规范职位投递表_20251204.xlsx')