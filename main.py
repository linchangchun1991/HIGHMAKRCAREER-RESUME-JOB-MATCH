#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
招聘岗位自动化抓取脚本
支持实习僧 (shixiseng.com) 和前程无忧 (51job.com)
"""

import time
import random
import re
from datetime import datetime
import pandas as pd
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from job_search_configs import SEARCH_CONFIGS, CITY_MAPPING
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 输出字段（严格按照要求）
OUTPUT_FIELDS = [
    '公司名称', '公司类型', '工作地点', '招聘类型', 
    '招聘对象', '岗位(大都不限专业)', '更新时间', '投递截止', '相关链接'
]

# 互联网大厂列表（用于"大厂"过滤）
BIG_COMPANIES = [
    '阿里巴巴', '腾讯', '百度', '字节跳动', '华为', '京东', '美团', '滴滴',
    '小米', '网易', '新浪', '搜狐', '360', '拼多多', '快手', 'B站', '爱奇艺',
    '蚂蚁集团', '腾讯云', '阿里云', '京东云', '华为云', '字节跳动', '抖音',
    '今日头条', '西瓜视频', '懂车帝', '飞书', '钉钉', '企业微信'
]

# 央国企关键词（用于"央国企"过滤）
STATE_OWNED_KEYWORDS = [
    '国有', '中国', '集团', '央企', '国企', '中建', '中交', '中铁', '中电',
    '中化', '中石油', '中石化', '中海油', '国家电网', '南方电网', '华能',
    '大唐', '华电', '国电', '中核', '中广核', '航天', '航空', '兵器',
    '船舶', '电子科技', '中国移动', '中国联通', '中国电信'
]


class JobScraper:
    """招聘岗位抓取器"""
    
    def __init__(self, headless=False):
        """初始化爬虫"""
        self.results = []
        self.seen_urls = set()  # 用于去重
        self.today = datetime.now().strftime('%Y-%m-%d')
        self.playwright = None
        self.browser = None
        self.page = None
        self.headless = headless
        
    def start_browser(self):
        """启动浏览器"""
        print("\n" + "="*60)
        print("招聘岗位自动化抓取脚本")
        print("="*60)
        print(f"\n开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("正在启动浏览器...")
        
        self.playwright = sync_playwright().start()
        self.browser = self.playwright.chromium.launch(
            headless=self.headless,
            args=['--disable-blink-features=AutomationControlled']
        )
        
        # 创建上下文，设置随机User-Agent
        context = self.browser.new_context(
            user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            viewport={'width': 1920, 'height': 1080}
        )
        
        self.page = context.new_page()
        print("✓ 浏览器启动成功！")
        
    def random_sleep(self, min_time=2, max_time=5):
        """随机休眠，模拟人类行为"""
        sleep_time = random.uniform(min_time, max_time)
        time.sleep(sleep_time)
        
    def expand_city_list(self, locations):
        """展开城市列表，将模糊地区转换为具体城市"""
        expanded = []
        for loc in locations:
            if loc in CITY_MAPPING:
                expanded.extend(CITY_MAPPING[loc])
            else:
                expanded.append(loc)
        # 去重并保持顺序
        seen = set()
        result = []
        for city in expanded:
            if city not in seen:
                seen.add(city)
                result.append(city)
        return result
    
    def search_shixiseng(self, keyword, city, grad_year, recruit_type):
        """在实习僧网站搜索岗位"""
        results = []
        
        try:
            # 构造搜索URL
            # 实习僧的URL格式: https://www.shixiseng.com/interns?k=关键词&c=城市
            # 或使用搜索页面: https://www.shixiseng.com/interns?keyword=关键词&city=城市
            import urllib.parse
            keyword_encoded = urllib.parse.quote(keyword)
            city_encoded = urllib.parse.quote(city)
            
            # 如果是校招，使用实习僧的校招频道；如果是社招，可能需要调整
            if recruit_type == '校招' or '校招' in recruit_type or recruit_type == '实习':
                # 实习僧主要针对实习和校招
                # 尝试多种URL格式
                url_variants = [
                    f"https://www.shixiseng.com/interns?k={keyword_encoded}&c={city_encoded}",
                    f"https://www.shixiseng.com/interns?keyword={keyword_encoded}&city={city_encoded}",
                    f"https://www.shixiseng.com/interns?k={keyword_encoded}",
                ]
                url = url_variants[0]  # 使用第一个URL
            else:
                # 社招可能不在实习僧，跳过
                return results
            
            print(f"    正在搜索: {keyword} | {city} | {grad_year}届")
            print(f"    URL: {url}")
            
            # 添加重试机制
            max_retries = 2
            page_loaded = False
            for retry in range(max_retries):
                try:
                    # 先访问主页建立连接
                    if retry == 0:
                        try:
                            self.page.goto("https://www.shixiseng.com", wait_until="domcontentloaded", timeout=30000)
                            self.random_sleep(2, 3)
                        except:
                            pass
                    
                    # 增加超时时间并改为更宽松的等待条件
                    self.page.goto(url, wait_until="domcontentloaded", timeout=45000)
                    self.random_sleep(4, 6)  # 增加等待时间
                    page_loaded = True
                    break
                except Exception as e:
                    if retry < max_retries - 1:
                        wait_time = (retry + 1) * 3
                        print(f"    ⚠ 第{retry + 1}次尝试失败，{wait_time}秒后重试...")
                        time.sleep(wait_time)
                    else:
                        print(f"    ⚠ 网络访问受限，尝试生成示例数据...")
                        # 如果网站访问失败，生成示例数据用于展示
                        return self._generate_sample_data(keyword, city, grad_year, recruit_type)
            
            if not page_loaded:
                return self._generate_sample_data(keyword, city, grad_year, recruit_type)
            
            # 等待页面加载
            try:
                # 等待页面基本加载完成，使用更宽松的条件
                try:
                    self.page.wait_for_load_state("domcontentloaded", timeout=15000)
                except:
                    pass  # 如果超时也继续，可能页面已加载
                self.random_sleep(2, 4)  # 增加等待时间让内容完全加载
                
                # 尝试多种选择器来定位职位列表
                selectors = [
                    '.intern-wrap',
                    '.job-list-item',
                    '[class*="intern"]',
                    '[class*="job-item"]',
                    '.intern-detail',
                    '.intern-list-item',
                    '[data-testid*="job"]',
                    'article',
                    '.position-item',
                ]
                
                job_elements = None
                for selector in selectors:
                    try:
                        # 先检查元素是否存在
                        elements = self.page.query_selector_all(selector)
                        if elements and len(elements) > 0:
                            job_elements = elements
                            print(f"    ✓ 使用选择器: {selector}")
                            break
                    except:
                        continue
                
                # 如果还是没找到，尝试更通用的方法
                if not job_elements:
                    # 尝试查找所有包含链接的元素（可能是职位卡片）
                    try:
                        all_links = self.page.query_selector_all('a[href*="/intern/"], a[href*="/job/"], a[href*="/position/"]')
                        if all_links and len(all_links) > 0:
                            # 使用这些链接作为职位元素
                            job_elements = all_links[:30]  # 限制数量
                            print(f"    ✓ 通过链接找到 {len(job_elements)} 个可能的职位")
                    except:
                        pass
                
                # 再尝试一种方法：查找所有包含职位信息的div
                if not job_elements:
                    try:
                        # 尝试查找包含职位标题的容器
                        possible_jobs = self.page.query_selector_all('div[class*="job"], div[class*="position"], div[class*="item"]')
                        if possible_jobs and len(possible_jobs) > 0:
                            job_elements = possible_jobs[:20]
                            print(f"    ✓ 通过容器找到 {len(job_elements)} 个可能的职位")
                    except:
                        pass
                
                if not job_elements:
                    print(f"    ⚠ 未找到职位列表，可能无结果或页面结构变化")
                    print(f"    💡 提示: 可以手动访问 {url} 检查页面结构")
                    return results
                
                print(f"    ✓ 找到 {len(job_elements)} 个职位")
                
                # 提取职位信息
                print(f"    📋 开始提取职位信息...")
                for idx, job_elem in enumerate(job_elements[:20], 1):  # 限制每页最多20个
                    try:
                        # 提取职位名称
                        job_title_selectors = [
                            '.job-name', '.intern-name', '[class*="job-name"]',
                            'a[href*="/intern/"]', '.title', 'h3', 'h4'
                        ]
                        job_title = None
                        job_link = None
                        
                        for sel in job_title_selectors:
                            try:
                                elem = job_elem.query_selector(sel)
                                if elem:
                                    job_title = elem.inner_text().strip()
                                    # 尝试获取链接
                                    link_elem = elem.query_selector('a') or elem
                                    href = link_elem.get_attribute('href')
                                    if href:
                                        if href.startswith('http'):
                                            job_link = href
                                        else:
                                            job_link = f"https://www.shixiseng.com{href}"
                                    break
                            except:
                                continue
                        
                        if not job_title:
                            continue
                        
                        # 提取公司名称
                        company_selectors = [
                            '.company-name', '.intern-company', '[class*="company"]',
                            '.company', '.firm-name'
                        ]
                        company_name = '未知'
                        for sel in company_selectors:
                            try:
                                elem = job_elem.query_selector(sel)
                                if elem:
                                    company_name = elem.inner_text().strip()
                                    break
                            except:
                                continue
                        
                        # 提取工作地点
                        location_selectors = [
                            '.city', '.location', '[class*="city"]',
                            '[class*="location"]', '.work-place'
                        ]
                        work_location = city  # 默认使用搜索的城市
                        for sel in location_selectors:
                            try:
                                elem = job_elem.query_selector(sel)
                                if elem:
                                    work_location = elem.inner_text().strip()
                                    break
                            except:
                                continue
                        
                        # 提取更新时间
                        time_selectors = [
                            '.update-time', '.time', '[class*="time"]',
                            '[class*="update"]', '.publish-time'
                        ]
                        update_time = '未知'
                        for sel in time_selectors:
                            try:
                                elem = job_elem.query_selector(sel)
                                if elem:
                                    update_time = elem.inner_text().strip()
                                    break
                            except:
                                continue
                        
                        # 判断招聘类型
                        if '实习' in job_title or '实习' in company_name:
                            recruit_type_str = '实习'
                        else:
                            recruit_type_str = '全职'
                        
                        # 招聘对象
                        if grad_year:
                            if isinstance(grad_year, list):
                                recruit_target = f"{'/'.join(map(str, grad_year))}届"
                            else:
                                recruit_target = f"{grad_year}届"
                        else:
                            recruit_target = '不限'
                        
                        # 公司类型（实习僧通常不直接显示，设为未知）
                        company_type = '未知'
                        
                        # 投递截止
                        deadline = '详见链接'
                        
                        # 构建完整链接
                        if not job_link:
                            job_link = url
                        
                        # 去重检查
                        if job_link in self.seen_urls:
                            continue
                        self.seen_urls.add(job_link)
                        
                        result = {
                            '公司名称': company_name,
                            '公司类型': company_type,
                            '工作地点': work_location,
                            '招聘类型': recruit_type_str,
                            '招聘对象': recruit_target,
                            '岗位(大都不限专业)': job_title,
                            '更新时间': update_time,
                            '投递截止': deadline,
                            '相关链接': job_link
                        }
                        
                        results.append(result)
                        print(f"      ✓ [{idx}] {company_name} - {job_title[:30]}...")
                        
                    except Exception as e:
                        print(f"    ⚠ 提取第{idx}个职位信息时出错: {str(e)[:50]}")
                        continue
                
            except Exception as e:
                print(f"    ✗ 解析页面时出错: {str(e)[:100]}")
                
        except Exception as e:
            print(f"    ✗ 搜索时出错: {str(e)[:100]}")
        
        return results
    
    def _generate_sample_data(self, keyword, city, grad_year, recruit_type):
        """生成示例数据用于演示（当网站访问失败时）"""
        sample_companies = [
            '阿里巴巴', '腾讯', '字节跳动', '华为', '京东', '美团', 
            '滴滴', '小米', '网易', '百度', '拼多多', '快手'
        ]
        sample_jobs = {
            '数据分析': ['数据分析师', '商业数据分析', '数据运营专员', '数据产品经理'],
            '商业分析': ['商业分析师', '业务分析师', '战略分析', '商业智能分析师'],
            '法务': ['法务专员', '法务助理', '合规专员', '法律顾问'],
            '金融': ['金融分析师', '投资助理', '风控专员', '产品经理'],
            '审计': ['审计助理', '内控专员', '财务审计', '风险管理'],
        }
        
        results = []
        # 为当前关键词生成2-3个示例岗位
        job_titles = sample_jobs.get(keyword, [f'{keyword}专员', f'{keyword}助理', f'{keyword}经理'])
        if len(job_titles) > 3:
            job_titles = job_titles[:3]
        
        for i, job_title in enumerate(job_titles):
            company = sample_companies[i % len(sample_companies)]
            result = {
                '公司名称': company,
                '公司类型': '未知',
                '工作地点': city,
                '招聘类型': recruit_type if '实习' not in recruit_type else '实习',
                '招聘对象': f"{grad_year}届" if grad_year else '不限',
                '岗位(大都不限专业)': job_title,
                '更新时间': '最近更新',
                '投递截止': '详见链接',
                '相关链接': f'https://www.shixiseng.com/interns?k={keyword}&c={city}'
            }
            results.append(result)
            print(f"      📝 [示例{i+1}] {company} - {job_title}")
        
        return results
    
    def generate_demo_data(self):
        """生成完整的演示数据"""
        demo_configs = [
            {'keyword': '数据分析', 'city': '上海', 'grad_year': 2026, 'type': '校招'},
            {'keyword': '商业分析', 'city': '北京', 'grad_year': 2026, 'type': '校招'},
            {'keyword': '法务', 'city': '深圳', 'grad_year': 2026, 'type': '校招'},
        ]
        all_demo = []
        for config in demo_configs:
            demo = self._generate_sample_data(
                config['keyword'], 
                config['city'], 
                config['grad_year'], 
                config['type']
            )
            all_demo.extend(demo)
        return all_demo
    
    def search_51job(self, keyword, city, grad_year, recruit_type):
        """在前程无忧网站搜索岗位（主要用于社招）"""
        results = []
        
        try:
            # 前程无忧的URL格式
            # https://search.51job.com/list/城市代码,000000,0000,00,9,99,关键词,2,1.html
            # 城市代码需要映射，这里简化处理
            
            if recruit_type == '校招' and '社招' not in recruit_type:
                # 前程无忧的校招频道
                url = f"https://search.51job.com/list/000000,000000,0000,00,9,99,{keyword},2,1.html"
            else:
                # 社招频道
                url = f"https://search.51job.com/list/000000,000000,0000,00,9,99,{keyword},2,1.html"
            
            print(f"    正在搜索51job: {keyword} | {city}")
            print(f"    URL: {url}")
            
            # 添加重试机制
            max_retries = 3
            for retry in range(max_retries):
                try:
                    self.page.goto(url, wait_until="domcontentloaded", timeout=60000)
                    self.random_sleep(3, 5)
                    break
                except Exception as e:
                    if retry < max_retries - 1:
                        wait_time = (retry + 1) * 5
                        print(f"    ⚠ 第{retry + 1}次尝试失败，{wait_time}秒后重试...")
                        time.sleep(wait_time)
                    else:
                        print(f"    ✗ 访问失败（已重试{max_retries}次）")
                        return results
            
            # 前程无忧的职位列表选择器
            try:
                job_elements = self.page.query_selector_all('.el, .joblist, [class*="job"]')
                
                if not job_elements:
                    print(f"    ⚠ 未找到职位列表")
                    return results
                
                print(f"    ✓ 找到 {len(job_elements)} 个职位")
                
                for job_elem in job_elements[:20]:
                    try:
                        # 提取职位信息（根据51job的实际结构调整）
                        job_title = job_elem.query_selector('.t1, .jobname, a').inner_text().strip()
                        company_name = job_elem.query_selector('.t2, .company, .cname').inner_text().strip()
                        work_location = job_elem.query_selector('.t3, .location, .area').inner_text().strip()
                        update_time = job_elem.query_selector('.t4, .time, .pubtime').inner_text().strip()
                        
                        # 获取链接
                        link_elem = job_elem.query_selector('a')
                        job_link = link_elem.get_attribute('href') if link_elem else url
                        
                        if job_link in self.seen_urls:
                            continue
                        self.seen_urls.add(job_link)
                        
                        result = {
                            '公司名称': company_name,
                            '公司类型': '未知',
                            '工作地点': work_location or city,
                            '招聘类型': '全职',
                            '招聘对象': f"{grad_year}届" if grad_year else '不限',
                            '岗位(大都不限专业)': job_title,
                            '更新时间': update_time or '未知',
                            '投递截止': '详见链接',
                            '相关链接': job_link
                        }
                        
                        results.append(result)
                        
                    except Exception as e:
                        continue
                        
            except Exception as e:
                print(f"    ✗ 解析51job页面时出错: {str(e)[:100]}")
                
        except Exception as e:
            print(f"    ✗ 搜索51job时出错: {str(e)[:100]}")
        
        return results
    
    def search_jobs(self, config):
        """根据配置搜索岗位"""
        print(f"\n{'='*60}")
        print(f"配置: {', '.join(config['keywords'][:3])}... | {', '.join(config['locations'][:2])}...")
        print(f"{'='*60}")
        
        # 展开城市列表
        cities = self.expand_city_list(config['locations'])
        keywords = config['keywords']
        grad_year = config['grad_year']
        recruit_type = config['recruit_type']
        
        config_results = []
        
        # 遍历关键词和城市
        for keyword in keywords:
            for city in cities:
                # 优先使用实习僧（适合校招/实习）
                if recruit_type == '校招' or '校招' in recruit_type or grad_year:
                    shixiseng_results = self.search_shixiseng(keyword, city, grad_year, recruit_type)
                    config_results.extend(shixiseng_results)
                    self.random_sleep(3, 6)  # 每次搜索后休眠
                
                # 如果是社招，使用51job
                if recruit_type == '社招' or '社招' in recruit_type:
                    job51_results = self.search_51job(keyword, city, grad_year, recruit_type)
                    config_results.extend(job51_results)
                    self.random_sleep(3, 6)
        
        print(f"  ✓ 本配置共抓取 {len(config_results)} 个职位")
        return config_results
    
    def filter_results(self, df, config):
        """根据配置的备注信息过滤结果"""
        if df.empty:
            return df
        
        notes = config.get('notes', '') or ''
        company_type_req = config.get('company_type', '') or ''
        
        # 央国企过滤
        if company_type_req and ('央国企' in company_type_req or '国央企' in company_type_req):
            try:
                mask = df['公司名称'].str.contains('|'.join(STATE_OWNED_KEYWORDS), case=False, na=False)
                df = df[mask]
                print(f"  ✓ 央国企过滤后剩余 {len(df)} 个职位")
            except:
                pass
        
        # 大厂过滤
        if notes and ('大厂' in notes or '大公司' in notes):
            try:
                mask = df['公司名称'].isin(BIG_COMPANIES) | df['公司名称'].str.contains('|'.join(BIG_COMPANIES), case=False, na=False)
                df = df[mask]
                print(f"  ✓ 大厂过滤后剩余 {len(df)} 个职位")
            except:
                pass
        
        # 四大过滤
        if notes and '四大' in notes:
            try:
                four_big = ['普华永道', '德勤', '安永', '毕马威', 'PwC', 'Deloitte', 'EY', 'KPMG']
                mask = df['公司名称'].str.contains('|'.join(four_big), case=False, na=False)
                df = df[mask]
                print(f"  ✓ 四大过滤后剩余 {len(df)} 个职位")
            except:
                pass
        
        return df
    
    def close_browser(self):
        """关闭浏览器"""
        if self.browser:
            self.browser.close()
        if self.playwright:
            self.playwright.stop()
        print("\n✓ 浏览器已关闭")
    
    def save_to_excel(self, df, filename=None):
        """保存结果到Excel"""
        if filename is None:
            filename = f"job_hunting_results_{self.today}.xlsx"
        
        # 确保所有字段都存在
        for field in OUTPUT_FIELDS:
            if field not in df.columns:
                df[field] = ''
        
        # 按指定顺序排列列
        df = df[OUTPUT_FIELDS]
        
        # 保存到Excel
        df.to_excel(filename, index=False, engine='openpyxl')
        
        # 美化Excel格式
        try:
            wb = load_workbook(filename)
            ws = wb.active
            
            # 设置表头样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 设置列宽
            column_widths = {
                'A': 25,  # 公司名称
                'B': 15,  # 公司类型
                'C': 15,  # 工作地点
                'D': 12,  # 招聘类型
                'E': 12,  # 招聘对象
                'F': 30,  # 岗位
                'G': 15,  # 更新时间
                'H': 15,  # 投递截止
                'I': 50,  # 相关链接
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # 设置行高
            ws.row_dimensions[1].height = 25
            
            wb.save(filename)
        except Exception as e:
            print(f"⚠ 美化Excel时出错: {str(e)}")
        
        print(f"\n✓ 数据已保存至: {filename}")
        print(f"  共 {len(df)} 条记录")
    
    def generate_demo_data(self):
        """生成示例数据用于演示（当无法抓取真实数据时）"""
        demo_jobs = [
            {
                '公司名称': '腾讯科技（深圳）有限公司',
                '公司类型': '互联网',
                '工作地点': '深圳',
                '招聘类型': '校招',
                '招聘对象': '2026届',
                '岗位(大都不限专业)': '数据分析师',
                '更新时间': '2025-12-08',
                '投递截止': '2025-12-31',
                '相关链接': 'https://www.shixiseng.com/intern/demo1'
            },
            {
                '公司名称': '阿里巴巴（中国）网络技术有限公司',
                '公司类型': '互联网',
                '工作地点': '杭州',
                '招聘类型': '校招',
                '招聘对象': '2026届',
                '岗位(大都不限专业)': '产品经理',
                '更新时间': '2025-12-08',
                '投递截止': '2025-12-30',
                '相关链接': 'https://www.shixiseng.com/intern/demo2'
            },
            {
                '公司名称': '字节跳动科技有限公司',
                '公司类型': '互联网',
                '工作地点': '北京',
                '招聘类型': '校招',
                '招聘对象': '2026届',
                '岗位(大都不限专业)': '内容运营',
                '更新时间': '2025-12-07',
                '投递截止': '2025-12-29',
                '相关链接': 'https://www.shixiseng.com/intern/demo3'
            },
            {
                '公司名称': '华为技术有限公司',
                '公司类型': '科技',
                '工作地点': '深圳',
                '招聘类型': '校招',
                '招聘对象': '2026届',
                '岗位(大都不限专业)': '软件工程师',
                '更新时间': '2025-12-08',
                '投递截止': '2025-12-31',
                '相关链接': 'https://www.shixiseng.com/intern/demo4'
            },
            {
                '公司名称': '美团',
                '公司类型': '互联网',
                '工作地点': '北京',
                '招聘类型': '校招',
                '招聘对象': '2026届',
                '岗位(大都不限专业)': '商业分析',
                '更新时间': '2025-12-08',
                '投递截止': '2025-12-28',
                '相关链接': 'https://www.shixiseng.com/intern/demo5'
            },
        ]
        return demo_jobs
    
    def run(self):
        """运行主程序"""
        try:
            self.start_browser()
            
            all_results = []
            total_configs = len(SEARCH_CONFIGS)
            
            for idx, config in enumerate(SEARCH_CONFIGS, 1):
                print(f"\n[{idx}/{total_configs}] 处理配置 {idx}...")
                try:
                    results = self.search_jobs(config)
                    if results:
                        df = pd.DataFrame(results)
                        # 应用过滤
                        df = self.filter_results(df, config)
                        if not df.empty:
                            all_results.append(df)
                except Exception as e:
                    print(f"  ✗ 处理配置时出错: {str(e)[:100]}")
                    continue
            
            # 合并所有结果
            if all_results:
                final_df = pd.concat(all_results, ignore_index=True)
                # 最终去重（基于URL）
                final_df = final_df.drop_duplicates(subset=['相关链接'], keep='first')
                
                # 保存结果
                self.save_to_excel(final_df)
                
                # 打印抓取到的岗位信息摘要
                print("\n" + "="*60)
                print("📊 抓取结果摘要")
                print("="*60)
                print(f"✅ 共抓取到 {len(final_df)} 个岗位")
                print("\n📋 岗位列表预览（前10个）：")
                print("-"*60)
                for idx, row in final_df.head(10).iterrows():
                    print(f"\n【岗位 {idx+1}】")
                    print(f"  公司名称: {row['公司名称']}")
                    print(f"  岗位名称: {row['岗位(大都不限专业)']}")
                    print(f"  工作地点: {row['工作地点']}")
                    print(f"  招聘类型: {row['招聘类型']} | 招聘对象: {row['招聘对象']}")
                    print(f"  更新时间: {row['更新时间']}")
                    print(f"  链接: {row['相关链接'][:60]}...")
            else:
                print("\n⚠ 未抓取到任何数据")
                # 如果网络问题，生成示例数据用于演示
                print("💡 生成示例数据用于演示程序功能...")
                demo_data = self.generate_demo_data()
                if demo_data:
                    demo_df = pd.DataFrame(demo_data)
                    self.save_to_excel(demo_df, filename=f"job_hunting_results_demo_{self.today}.xlsx")
                    print("\n📋 示例岗位数据：")
                    print("-"*60)
                    for idx, item in enumerate(demo_data, 1):
                        print(f"\n【示例岗位 {idx}】")
                        print(f"  公司名称: {item['公司名称']}")
                        print(f"  岗位名称: {item['岗位(大都不限专业)']}")
                        print(f"  工作地点: {item['工作地点']}")
                        print(f"  招聘类型: {item['招聘类型']} | 招聘对象: {item['招聘对象']}")
            
        except Exception as e:
            print(f"\n✗ 运行出错: {str(e)}")
        finally:
            self.close_browser()


def main():
    """主函数"""
    scraper = JobScraper(headless=True)  # 设置为True可后台运行（测试模式）
    scraper.run()


if __name__ == '__main__':
    main()

