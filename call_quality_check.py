#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
先发系统销售录音质检全自动化脚本
功能：连接Chrome浏览器，随机抽取顾问录音，下载后转文字，调用AI质检，生成Excel报告
"""

import os
import time
import random
import json
import re
from datetime import datetime, timedelta
from pathlib import Path
import pandas as pd
import requests
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

# ==================== 配置区域（请填写API密钥）====================
DASHSCOPE_API_KEY = "sk-668c28bae516493d9ea8a3662118ec98"  # 阿里云DashScope API Key

# 浏览器连接配置
CDP_URL = "http://localhost:9222"

# ==================== 顾问名单 ====================
ADVISOR_LIST = [
    "赵涵", "郭传宝", "谭彤", "梁瑶欣", "汪雯雯", "迟浩", "辛玉亮", "冯映深", 
    "金鸿浩", "李盼盼", "暴佳鑫", "石玉峰", "关云程", "秦本杰", "秦正", "许承业", 
    "李广胜", "于磊", "刘东里", "张洪玮", "李丽", "张淼", "邱权", "王则爽", 
    "付昀舟", "宋存芳", "邵吉雨", "郑雪雯", "赵红", "刘长林", "徐程", "沙修宇", 
    "李文龙", "袁洪霞", "毛英辉", "吴秀娟", "景天雨", "曲佳明", "孙佳晶", "李晓", 
    "赵娜", "毛栖萍", "周晓琳", "胡嘉祥", "王珏玢", "刘赛赛", "闫佳楠", "于艳艳", 
    "高彬", "张越", "冯美琪", "武术强", "徐贝贝", "刘芝延", "王馨頔", "赖志伟", "纪旋"
]

# 随机抽取数量
RANDOM_SAMPLE_COUNT = 10

# ==================== 目录设置 ====================
RECORDINGS_DIR = Path("recordings")
RECORDINGS_DIR.mkdir(exist_ok=True)


class CallQualityChecker:
    """录音质检自动化工具"""
    
    def __init__(self):
        self.results = []
        self.yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        self.page = None
        
    def connect_browser(self):
        """连接到已打开的Chrome浏览器"""
        print(f"[1] 正在连接到浏览器 {CDP_URL}...")
        
        playwright = sync_playwright().start()
        
        # 连接到已有浏览器
        browser = playwright.chromium.connect_over_cdp(CDP_URL)
        
        # 获取所有上下文
        contexts = browser.contexts
        if not contexts:
            raise Exception("未找到浏览器上下文，请确保Chrome已用 --remote-debugging-port=9222 启动")
        
        # 获取第一个上下文的主页面
        context = contexts[0]
        pages = context.pages
        
        # 查找包含 xianfasj 的页面
        target_page = None
        xianfasj_pages = []
        for page in pages:
            try:
                url = page.url
                if 'xianfasj' in url:
                    xianfasj_pages.append((page, url))
            except:
                continue
        
        if xianfasj_pages:
            # 优先选择查询页面，而不是登录页面
            for page, url in xianfasj_pages:
                if '/login' not in url.lower():
                    target_page = page
                    print(f"   ✓ 找到目标页面: {url}")
                    break
            
            # 如果都是登录页面，使用第一个
            if not target_page:
                target_page, url = xianfasj_pages[0]
                print(f"   找到登录页面: {url}")
        else:
            # 如果没有找到xianfasj页面，尝试查找其他可能的页面
            print("   ⚠ 未找到xianfasj网站页面")
            print("   正在查找所有打开的页面...")
            for page in pages:
                try:
                    url = page.url
                    # 跳过扩展页面和空白页
                    if 'chrome-extension' not in url and url not in ['about:blank', 'chrome://newtab/']:
                        print(f"   发现页面: {url}")
                except:
                    continue
            
            # 尝试打开新页面并导航到先发系统
            print("   尝试打开新页面...")
            target_page = context.new_page()
            try:
                target_page.goto("https://www.xianfasj.com", wait_until='domcontentloaded', timeout=10000)
                print(f"   ✓ 已打开新页面: {target_page.url}")
            except Exception as e:
                print(f"   打开页面失败: {e}")
                print("   请手动在浏览器中打开 https://www.xianfasj.com 并登录")
        
        self.page = target_page
        print("   浏览器连接成功！")
        return browser, playwright
    
    def check_login_status(self):
        """检查是否已登录，如果未登录则等待"""
        print(f"[1.5] 检查登录状态...")
        
        try:
            current_url = self.page.url
            print(f"   当前页面: {current_url}")
            
            # 检查是否在登录页面
            if '/login' in current_url.lower() or '登录' in self.page.title():
                print("\n" + "="*60)
                print("⚠️  检测到当前在登录页面！")
                print("请完成以下操作：")
                print("1. 在浏览器中完成登录")
                print("2. 导航到通话记录查询页面")
                print("3. 确保页面已加载完成")
                print("="*60)
                print("\n等待登录完成（最多等待5分钟，每10秒检查一次）...")
                
                max_wait = 300  # 最多等待5分钟
                for i in range(max_wait):
                    time.sleep(1)
                    try:
                        new_url = self.page.url
                        # 检查是否已离开登录页面
                        if '/login' not in new_url.lower():
                            # 检查是否找到跟进人选择框或日期筛选框（说明在查询页面）
                            try:
                                # 尝试查找跟进人选择框
                                all_inputs = self.page.query_selector_all('input')
                                for inp in all_inputs:
                                    placeholder = inp.get_attribute('placeholder') or ''
                                    if '跟进人' in placeholder or '日期' in placeholder:
                                        print(f"\n✓ 检测到已登录并进入查询页面！")
                                        print(f"   新页面: {new_url}")
                                        self.page.wait_for_timeout(2000)  # 等待页面稳定
                                        return True
                            except:
                                pass
                            
                            # 或者检查URL是否包含查询相关关键词
                            if any(keyword in new_url.lower() for keyword in ['record', 'call', '通话', '查询', 'search']):
                                print(f"\n✓ 检测到已离开登录页面！")
                                print(f"   新页面: {new_url}")
                                self.page.wait_for_timeout(2000)  # 等待页面稳定
                                return True
                    except:
                        pass
                    
                    if (i + 1) % 10 == 0:
                        print(f"   等待中... ({i+1}秒)")
                
                print("\n⚠️  等待超时，但继续尝试执行...")
                return False
            else:
                print(f"   ✓ 已不在登录页面，继续执行...")
                # 检查是否在正确的查询页面（尝试查找跟进人选择框）
                try:
                    all_inputs = self.page.query_selector_all('input')
                    for inp in all_inputs:
                        placeholder = inp.get_attribute('placeholder') or ''
                        if '跟进人' in placeholder:
                            print(f"   ✓ 检测到查询页面元素")
                            return True
                except:
                    pass
                return True
                
        except Exception as e:
            print(f"   检查登录状态时出错: {e}，继续执行...")
            return True
    
    def set_date_filter(self):
        """设置日期筛选为昨天"""
        print(f"[2] 设置日期筛选为: {self.yesterday}")
        
        try:
            # 等待页面加载
            self.page.wait_for_timeout(2000)
            
            # 查找日期输入框（可能有多种选择器，尝试常见的选择器）
            date_selectors = [
                'input[placeholder*="日期"]',
                'input[placeholder*="开始"]',
                'input[type="date"]',
                '.ant-picker-input input',
                'input.el-input__inner[placeholder*="日期"]',
                'input[class*="date"]'
            ]
            
            start_date_input = None
            for selector in date_selectors:
                try:
                    elements = self.page.query_selector_all(selector)
                    if len(elements) >= 2:
                        start_date_input = elements[0]
                        end_date_input = elements[1]
                        break
                    elif len(elements) == 1:
                        # 可能是单个日期选择器
                        start_date_input = elements[0]
                        break
                except:
                    continue
            
            if not start_date_input:
                print("   警告：未找到日期输入框，尝试手动输入...")
                # 尝试通过文本定位
                self.page.wait_for_timeout(1000)
                return
            
            # 清空并输入日期
            start_date_input.click()
            self.page.wait_for_timeout(500)
            start_date_input.fill(self.yesterday)
            self.page.wait_for_timeout(300)
            
            # 如果有结束日期，也设置为昨天
            if 'end_date_input' in locals() and end_date_input:
                end_date_input.click()
                self.page.wait_for_timeout(500)
                end_date_input.fill(self.yesterday)
            
            # 按回车确认
            self.page.keyboard.press("Enter")
            self.page.wait_for_timeout(1000)
            
            print(f"   日期筛选设置完成")
            
        except Exception as e:
            print(f"   日期筛选设置时出错: {e}，继续执行...")
    
    def query_advisor(self, advisor_name):
        """查询指定顾问的通话记录"""
        print(f"   正在查询顾问: {advisor_name}...")
        
        try:
            # 确保在查询页面，如果不在则导航过去
            current_url = self.page.url
            if '/conversationList' not in current_url:
                print(f"   当前不在查询页面，正在导航...")
                self.page.goto("https://www.xianfasj.com/callControl/conversationList", wait_until='domcontentloaded', timeout=10000)
                self.page.wait_for_timeout(2000)
            
            # 等待页面加载
            self.page.wait_for_timeout(2000)
            
            # 打印当前URL以便调试
            current_url = self.page.url
            print(f"   当前页面URL: {current_url}")
            
            select_element = None
            
            # 方法1: 直接查找所有input，检查placeholder
            try:
                all_inputs = self.page.query_selector_all('input')
                print(f"   页面上共有 {len(all_inputs)} 个input元素")
                
                # 打印所有input的placeholder用于调试
                if all_inputs:
                    print(f"   前5个input元素的placeholder:")
                    for i, inp in enumerate(all_inputs[:5]):
                        try:
                            placeholder = inp.get_attribute('placeholder') or ''
                            input_type = inp.get_attribute('type') or ''
                            input_class = inp.get_attribute('class') or ''
                            print(f"     [{i+1}] type={input_type}, placeholder='{placeholder}', class='{input_class[:50]}'")
                        except:
                            pass
                for inp in all_inputs:
                    try:
                        placeholder = inp.get_attribute('placeholder') or ''
                        # 打印前几个placeholder用于调试
                        if placeholder and ('跟进' in placeholder or '请选择' in placeholder):
                            print(f"   发现输入框，placeholder: {placeholder}")
                        if placeholder and '跟进人' in placeholder:
                            select_element = inp
                            print(f"   ✓ 找到跟进人选择框: {placeholder}")
                            break
                    except:
                        continue
            except Exception as e:
                print(f"   方法1失败: {e}")
            
            # 方法2: 通过JavaScript查找（更灵活）
            if not select_element:
                try:
                    js_code = '''
                    () => {
                        // 方法2.1: 查找包含"跟进人："标签的元素，然后找相邻的input
                        const labels = Array.from(document.querySelectorAll('*')).filter(el => {
                            const text = el.textContent || '';
                            return (text.includes('跟进人：') || text.includes('跟进人:') || text.trim() === '跟进人') 
                                   && !text.includes('客户');
                        });
                        
                        for (const label of labels) {
                            // 查找父容器或兄弟元素中的input
                            let container = label.parentElement;
                            while (container && container !== document.body) {
                                // 查找同一容器中的input
                                const inputs = container.querySelectorAll('input');
                                for (const input of inputs) {
                                    const placeholder = input.getAttribute('placeholder') || '';
                                    if (placeholder.includes('跟进人') || placeholder.includes('请选择')) {
                                        return input;
                                    }
                                }
                                // 查找 .ant-select 或类似的选择组件
                                const selects = container.querySelectorAll('.ant-select, .el-select, [class*="select"]');
                                for (const select of selects) {
                                    const input = select.querySelector('input');
                                    if (input) return input;
                                }
                                container = container.parentElement;
                            }
                            
                            // 查找下一个兄弟元素
                            let nextSibling = label.nextElementSibling;
                            while (nextSibling) {
                                const input = nextSibling.querySelector('input');
                                if (input) {
                                    const placeholder = input.getAttribute('placeholder') || '';
                                    if (placeholder.includes('跟进人') || placeholder.includes('请选择')) {
                                        return input;
                                    }
                                }
                                nextSibling = nextSibling.nextElementSibling;
                            }
                        }
                        
                        // 方法2.2: 直接查找所有包含"跟进人"的input
                        const allInputs = document.querySelectorAll('input');
                        for (const input of allInputs) {
                            const placeholder = input.getAttribute('placeholder') || '';
                            if (placeholder.includes('跟进人') && !placeholder.includes('客户')) {
                                return input;
                            }
                        }
                        
                        return null;
                    }
                    '''
                    result_handle = self.page.evaluate_handle(js_code)
                    if result_handle:
                        element = result_handle.as_element()
                        if element:
                            select_element = element
                            placeholder = select_element.get_attribute('placeholder') or ''
                            print(f"   ✓ 通过JavaScript找到跟进人选择框: {placeholder}")
                except Exception as e:
                    print(f"   方法2失败: {e}")
            
            # 方法3: 通过更通用的选择器查找
            if not select_element:
                selectors = [
                    'input[placeholder*="请选择跟进人"]',
                    'input[placeholder*="跟进人"]',
                    '.ant-select input',
                    '.el-select input',
                    'input[placeholder*="请选择"]'
                ]
                for selector in selectors:
                    try:
                        elements = self.page.query_selector_all(selector)
                        for elem in elements:
                            placeholder = elem.get_attribute('placeholder') or ''
                            if placeholder and '跟进人' in placeholder and '客户' not in placeholder:
                                select_element = elem
                                print(f"   ✓ 通过选择器找到跟进人选择框: {placeholder}")
                                break
                        if select_element:
                            break
                    except:
                        continue
            
            if not select_element:
                # 如果还没找到，打印页面标题和部分HTML结构用于调试
                try:
                    page_title = self.page.title()
                    print(f"   页面标题: {page_title}")
                    print(f"   提示：如果无法找到跟进人选择框，请确保：")
                    print(f"     1. 已登录系统")
                    print(f"     2. 已导航到通话记录查询页面")
                    print(f"     3. 页面已完全加载")
                except:
                    pass
            
            if select_element:
                # 使用JavaScript直接操作选择框，避免被已选择项遮挡
                try:
                    # 方法1: 先通过JavaScript找到并点击清空按钮
                    clear_result = self.page.evaluate('''() => {
                        const clears = document.querySelectorAll('.ant-select-clear, .el-select__clear, [class*="clear"]');
                        for (const clear of clears) {
                            if (clear.offsetParent !== null) { // 可见
                                clear.click();
                                return true;
                            }
                        }
                        return false;
                    }''')
                    if clear_result:
                        print(f"   已通过JavaScript清空之前的选择")
                        self.page.wait_for_timeout(500)
                except:
                    pass
                
                # 方法2: 通过JavaScript操作选择框
                try:
                    select_success = self.page.evaluate('''(input) => {
                        // 找到选择框容器
                        const selectContainer = input.closest('.ant-select, .el-select, [class*="select"]');
                        if (selectContainer) {
                            // 先点击清空按钮（如果存在）
                            const clearBtn = selectContainer.querySelector('.ant-select-clear, .el-select__clear');
                            if (clearBtn) {
                                clearBtn.click();
                                return 'cleared';
                            }
                            
                            // 点击选择框容器打开下拉菜单
                            selectContainer.click();
                            return 'opened';
                        }
                        return false;
                    }''', select_element)
                    
                    if select_success:
                        print(f"   已打开选择框")
                        self.page.wait_for_timeout(500)
                    else:
                        # 如果JavaScript方法失败，尝试直接点击input（强制点击）
                        try:
                            select_element.click(force=True)
                            self.page.wait_for_timeout(500)
                        except:
                            pass
                except Exception as e:
                    print(f"   JavaScript操作失败: {e}，尝试直接点击...")
                    try:
                        select_element.click(force=True)
                        self.page.wait_for_timeout(500)
                    except:
                        pass
                
                # 在查询前，记录当前列表的第一条记录（用于验证查询是否生效）
                try:
                    pre_query_rows = self.page.query_selector_all('.ant-table-tbody tr, .el-table__body tr')
                    if pre_query_rows and len(pre_query_rows) > 0:
                        # 获取第一条记录的某些特征（比如第一列的内容）
                        pre_first_row_text = pre_query_rows[0].evaluate('''el => {
                            const firstCell = el.querySelector('td:first-child, th:first-child');
                            return firstCell ? (firstCell.innerText || firstCell.textContent || "").trim() : "";
                        }''')
                        print(f"   查询前第一条记录特征: {pre_first_row_text[:50] if pre_first_row_text else '空'}")
                except:
                    pre_first_row_text = None
                
                # 清空选择框内容（如果有的话）
                try:
                    select_element.click(force=True)
                    self.page.wait_for_timeout(300)
                    # 全选并删除
                    self.page.keyboard.press("Control+a") if os.name != 'nt' else self.page.keyboard.press("Meta+a")
                    self.page.wait_for_timeout(100)
                    self.page.keyboard.press("Backspace")
                    self.page.wait_for_timeout(300)
                except:
                    pass
                
                # 在选择框中输入顾问姓名
                print(f"   输入顾问姓名: {advisor_name}...")
                
                # 使用多种方法输入，确保成功
                input_success = False
                
                # 方法1: 使用JavaScript直接设置值
                try:
                    set_value_result = self.page.evaluate(f'''
                    (input) => {{
                        input.value = "{advisor_name}";
                        input.dispatchEvent(new Event("input", {{ bubbles: true }}));
                        input.dispatchEvent(new Event("change", {{ bubbles: true }}));
                        return input.value === "{advisor_name}";
                    }}
                    ''', select_element)
                    
                    if set_value_result:
                        print(f"   ✓ 已通过JavaScript设置顾问姓名")
                        input_success = True
                        self.page.wait_for_timeout(1000)
                except Exception as e:
                    print(f"   JavaScript设置失败: {e}")
                
                # 方法2: 如果JavaScript失败，使用fill方法
                if not input_success:
                    try:
                        select_element.fill(advisor_name)
                        input_success = True
                        self.page.wait_for_timeout(1000)
                    except Exception as e:
                        print(f"   fill方法失败: {e}")
                
                # 方法3: 如果fill也失败，使用type方法
                if not input_success:
                    try:
                        select_element.click()
                        self.page.wait_for_timeout(300)
                        self.page.keyboard.type(advisor_name, delay=100)
                        input_success = True
                        self.page.wait_for_timeout(1000)
                    except Exception as e:
                        print(f"   type方法失败: {e}")
                
                # 等待下拉选项加载
                self.page.wait_for_timeout(1500)
                
                # 尝试从下拉选项中选择
                try:
                    # 查找包含顾问姓名的下拉选项
                    option_selectors = [
                        f'.ant-select-item-option:has-text("{advisor_name}")',
                        f'.el-select-dropdown__item:has-text("{advisor_name}")',
                        f'div[role="option"]:has-text("{advisor_name}")',
                        f'li:has-text("{advisor_name}")',
                        f'text={advisor_name}',
                        f'[title="{advisor_name}"]',
                        f'[title*="{advisor_name}"]'
                    ]
                    
                    option_found = False
                    for selector in option_selectors:
                        try:
                            option = self.page.locator(selector).first
                            if option.is_visible(timeout=2000):
                                option.click()
                                print(f"   ✓ 已选择顾问: {advisor_name}")
                                option_found = True
                                break
                        except:
                            continue
                    
                    if not option_found:
                        # 如果没有找到选项，尝试按回车确认输入
                        print(f"   未找到下拉选项，按回车确认输入...")
                        self.page.keyboard.press("Enter")
                        self.page.wait_for_timeout(500)
                        self.page.keyboard.press("Tab")
                        self.page.wait_for_timeout(500)
                        print(f"   ✓ 已通过回车确认输入")
                        
                except Exception as e:
                    print(f"   选择选项时出错: {e}，尝试按回车...")
                    self.page.keyboard.press("Enter")
                    self.page.wait_for_timeout(500)
                    self.page.keyboard.press("Tab")
                
                self.page.wait_for_timeout(1000)  # 等待选择完成
            else:
                print(f"   无法找到跟进人选择框，跳过此顾问")
                return False
            
            # 查找并点击"查询"按钮
            print(f"   正在查找查询按钮...")
            
            # 方法1: 通过JavaScript查找并点击查询按钮（更可靠）
            try:
                js_code = '''
                () => {
                    // 先打印所有按钮用于调试
                    const allButtons = Array.from(document.querySelectorAll('button, .ant-btn, .el-button, [role="button"], span[role="button"]'));
                    const visibleButtons = allButtons.filter(btn => btn.offsetParent !== null);
                    const buttonTexts = visibleButtons.map(btn => (btn.textContent || '').trim()).filter(t => t);
                    
                    // 查找查询按钮（注意：按钮文本可能是"查 询"中间有空格）
                    for (const btn of visibleButtons) {
                        const text = (btn.textContent || '').trim();
                        // 匹配"查询"、"查 询"、"搜索"等
                        if (text.includes('查询') || text.replace(/\s+/g, '') === '查询' || text.includes('搜索') || btn.type === 'submit') {
                            btn.click();
                            return {success: true, text: text, allButtons: buttonTexts.slice(0, 5)};
                        }
                    }
                    return {success: false, allButtons: buttonTexts.slice(0, 10)};
                }
                '''
                result = self.page.evaluate(js_code)
                if result.get('allButtons'):
                    print(f"   页面上可见按钮: {result.get('allButtons')}")
                
                if result.get('success'):
                    print(f"   ✓ 通过JavaScript点击了查询按钮: '{result.get('text', '')}'")
                    # 等待查询完成，并等待列表刷新
                    self.page.wait_for_timeout(2000)
                    
                    # 等待表格数据加载完成（通过检查表格行数是否变化）
                    max_wait = 10
                    for wait_count in range(max_wait):
                        try:
                            rows = self.page.query_selector_all('.ant-table-tbody tr, .el-table__body tr')
                            if rows and len(rows) > 0:
                                # 检查行数是否稳定（连续两次相同）
                                self.page.wait_for_timeout(500)
                                rows2 = self.page.query_selector_all('.ant-table-tbody tr, .el-table__body tr')
                                if len(rows) == len(rows2):
                                    break
                        except:
                            pass
                        self.page.wait_for_timeout(500)
                    
                    # 如果页面跳转了，导航回查询页面
                    current_url_after = self.page.url
                    if '/conversationList' not in current_url_after:
                        print(f"   检测到页面跳转，正在返回查询页面...")
                        self.page.goto("https://www.xianfasj.com/callControl/conversationList", wait_until='domcontentloaded', timeout=10000)
                        self.page.wait_for_timeout(2000)
                    
                    # 验证查询是否成功：检查列表是否有数据，并且是否变化
                    list_selectors = [
                        '.ant-table-tbody tr',
                        '.el-table__body tr',
                        'tr[class*="row"]'
                    ]
                    for selector in list_selectors:
                        try:
                            rows = self.page.query_selector_all(selector)
                            if rows and len(rows) > 0:
                                # 检查查询后的第一条记录是否与查询前不同
                                if pre_first_row_text:
                                    post_first_row_text = rows[0].evaluate('''el => {
                                        const firstCell = el.querySelector('td:first-child, th:first-child');
                                        return firstCell ? (firstCell.innerText || firstCell.textContent || "").trim() : "";
                                    }''')
                                    if post_first_row_text == pre_first_row_text:
                                        print(f"   ⚠️ 警告：查询后第一条记录未变化，可能查询未生效")
                                        print(f"   查询前: {pre_first_row_text[:50]}, 查询后: {post_first_row_text[:50]}")
                                    else:
                                        print(f"   ✓ 查询成功，列表已更新（找到 {len(rows)} 条记录）")
                                else:
                                    print(f"   ✓ 查询成功，找到 {len(rows)} 条记录")
                                return True
                        except:
                            continue
                    
                    # 如果没有找到记录，可能真的没有数据
                    print(f"   ⚠ 查询后未找到记录（可能该顾问没有通话记录）")
                    return True
            except Exception as e:
                print(f"   JavaScript点击失败: {e}")
            
            # 方法2: 通过选择器查找（注意匹配"查 询"中间有空格的情况）
            query_selectors = [
                'button:has-text("查询")',
                'button:has-text("查 询")',  # 匹配中间有空格的情况
                'button:has-text("搜索")',
                '.ant-btn-primary:has-text("查询")',
                '.ant-btn-primary:has-text("查 询")',
                'button.ant-btn-primary',
                'button[type="submit"]',
                'button.el-button--primary'
            ]
            
            for selector in query_selectors:
                try:
                    query_button = self.page.locator(selector).first
                    if query_button.is_visible(timeout=1000):
                        text = query_button.text_content() or ''
                        query_button.click()
                        print(f"   ✓ 点击了查询按钮: '{text.strip()}'")
                        
                        # 等待查询完成，并等待列表刷新
                        self.page.wait_for_timeout(2000)
                        
                        # 等待表格数据加载完成
                        max_wait = 10
                        for wait_count in range(max_wait):
                            try:
                                rows = self.page.query_selector_all('.ant-table-tbody tr, .el-table__body tr')
                                if rows and len(rows) > 0:
                                    self.page.wait_for_timeout(500)
                                    rows2 = self.page.query_selector_all('.ant-table-tbody tr, .el-table__body tr')
                                    if len(rows) == len(rows2):
                                        break
                            except:
                                pass
                            self.page.wait_for_timeout(500)
                        
                        # 如果页面跳转了，导航回查询页面
                        current_url_after = self.page.url
                        if '/conversationList' not in current_url_after:
                            print(f"   检测到页面跳转，正在返回查询页面...")
                            self.page.goto("https://www.xianfasj.com/callControl/conversationList", wait_until='domcontentloaded', timeout=10000)
                            self.page.wait_for_timeout(2000)
                        
                        return True
                except:
                    continue
            
            # 方法3: 如果还是找不到，尝试通过表单提交
            print(f"   ⚠ 未找到查询按钮，尝试通过表单提交...")
            try:
                # 尝试找到表单并提交
                form = self.page.locator('form').first
                if form.is_visible(timeout=500):
                    form.evaluate('form => form.submit()')
                    print(f"   ✓ 已提交表单")
                    self.page.wait_for_timeout(3000)
                    return True
            except:
                pass
            
            # 方法4: 尝试点击选择框外的区域来触发查询（某些UI框架会自动查询）
            print(f"   ⚠ 尝试点击页面其他区域触发自动查询...")
            try:
                # 点击页面空白处
                self.page.click('body', position={'x': 100, 'y': 100})
                self.page.wait_for_timeout(2000)
            except:
                pass
            
            # 方法5: 尝试按回车键
            print(f"   ⚠ 尝试按回车键提交...")
            self.page.keyboard.press("Enter")
            self.page.wait_for_timeout(2000)
            
            # 方法6: 尝试通过JavaScript直接触发查询（如果选择框有onChange事件）
            try:
                # 触发input的change事件，可能某些框架会在change时自动查询
                js_trigger_change = '''
                () => {
                    const inputs = document.querySelectorAll('input[placeholder*="跟进人"]');
                    for (const input of inputs) {
                        if (input.value) {
                            // 触发change事件
                            input.dispatchEvent(new Event('change', { bubbles: true }));
                            input.dispatchEvent(new Event('blur', { bubbles: true }));
                        }
                    }
                    return true;
                }
                '''
                self.page.evaluate(js_trigger_change)
                self.page.wait_for_timeout(2000)
            except:
                pass
            
            # 再等待一下，看是否有自动查询
            self.page.wait_for_timeout(2000)
            
            print(f"   查询操作已完成，等待列表刷新...")
            return True  # 假设查询操作可以触发
                
        except Exception as e:
            print(f"   查询顾问时出错: {e}")
            return False
    
    def download_first_recording(self, advisor_name):
        """下载第一条通话录音"""
        try:
            # 等待列表加载，确保数据已刷新
            self.page.wait_for_timeout(3000)  # 增加等待时间
            
            # 打印当前页面URL用于调试
            current_url = self.page.url
            print(f"   当前页面: {current_url}")
            
            # 查找列表项（表格行或列表项）
            list_selectors = [
                'tr[class*="row"]',
                '.ant-table-tbody tr',
                '.el-table__body tr',
                'div[class*="list-item"]',
                'li[class*="item"]'
            ]
            
            rows = []
            for selector in list_selectors:
                try:
                    rows = self.page.query_selector_all(selector)
                    if rows:
                        print(f"   找到 {len(rows)} 条记录")
                        break
                except:
                    continue
            
            # 如果没有找到行，可能列表为空
            if not rows or len(rows) == 0:
                print(f"   列表为空，该顾问昨天没有通话记录")
                return None
            
            # 验证第一条记录是否属于当前顾问（重要！）
            first_row = rows[0]
            try:
                # 方法1: 尝试从第一行的所有单元格中提取文本
                cells = first_row.query_selector_all('td, th, div[class*="cell"], span[class*="cell"]')
                row_text_parts = []
                for cell in cells:
                    try:
                        cell_text = cell.evaluate('''el => {
                            const text = el.innerText || el.textContent || "";
                            return text.trim();
                        }''')
                        if cell_text:
                            row_text_parts.append(cell_text)
                    except:
                        continue
                
                # 方法2: 如果方法1失败，尝试从整行提取并清理
                if not row_text_parts:
                    row_text = first_row.evaluate('''el => {
                        const text = el.innerText || el.textContent || "";
                        return text.trim();
                    }''')
                    if row_text:
                        # 清理文本（移除多余的空白字符，但保留基本结构）
                        row_text = ' '.join(row_text.split())
                        row_text_parts = [row_text]
                
                # 合并所有文本
                full_row_text = ' '.join(row_text_parts)
                
                if full_row_text and len(full_row_text.strip()) > 5:  # 至少要有一些内容
                    # 检查是否包含顾问姓名
                    if advisor_name not in full_row_text:
                        print(f"   ⚠️ 警告：列表第一条记录不包含顾问'{advisor_name}'，可能查询未生效")
                        print(f"   第一条记录内容: {full_row_text[:150]}...")
                        print(f"   ⚠️ 跳过下载，避免下载错误的录音")
                        return None
                    else:
                        print(f"   ✓ 验证通过：列表第一条记录包含顾问'{advisor_name}'")
                else:
                    # 如果无法提取有效文本，可能是表格结构特殊，暂时跳过验证
                    print(f"   ⚠️ 无法提取有效记录文本（提取到的内容: '{full_row_text[:50]}'），跳过验证继续下载")
            except Exception as e:
                print(f"   验证记录时出错: {e}，继续下载...")
            
            # 查找下载按钮（在第一条记录中）
            
            download_selectors = [
                'button:has-text("下载")',
                'a:has-text("下载")',
                'button[title*="下载"]',
                '.download-btn',
                'button[class*="download"]',
                'i[class*="download"]',
                'svg[class*="download"]'
            ]
            
            download_button = None
            for selector in download_selectors:
                try:
                    buttons = first_row.query_selector_all(selector)
                    if buttons:
                        download_button = buttons[0]
                        break
                except:
                    continue
            
            # 如果第一行没找到，在整个页面中查找
            if not download_button:
                for selector in download_selectors:
                    try:
                        download_button = self.page.locator(selector).first
                        if download_button.is_visible(timeout=500):
                            break
                    except:
                        continue
            
            if download_button:
                # 设置下载监听
                with self.page.expect_download() as download_info:
                    download_button.click()
                    self.page.wait_for_timeout(1000)
                
                download = download_info.value
                
                # 保存文件
                filename = f"{advisor_name}_{self.yesterday}.mp3"
                filepath = RECORDINGS_DIR / filename
                
                download.save_as(filepath)
                print(f"   下载成功: {filename}")
                
                return filepath
            else:
                print(f"   未找到下载按钮")
                return None
                
        except PlaywrightTimeoutError:
            print(f"   下载超时，可能没有录音文件")
            return None
        except Exception as e:
            print(f"   下载录音时出错: {e}")
            return None
    
    def audio_to_text(self, audio_path):
        """使用DashScope将音频转换为文字"""
        if not DASHSCOPE_API_KEY:
            raise ValueError("请先设置 DASHSCOPE_API_KEY")
        
        try:
            import dashscope
            from dashscope.audio.asr import Transcription
            from dashscope import Files
            
            dashscope.api_key = DASHSCOPE_API_KEY
            
            print(f"   正在转写音频: {audio_path.name}...")
            
            # 先上传文件到dashscope
            print(f"   正在上传音频文件...")
            upload_result = Files.upload(
                file_path=str(audio_path),
                purpose='inference'
            )
            
            if upload_result.status_code != 200:
                print(f"   文件上传失败: {upload_result.message}")
                return None
            
            # 获取文件ID
            file_info = upload_result.output
            file_id = None
            
            # 提取file_id，格式：{'uploaded_files': [{'name': '...', 'file_id': '...'}]}
            if isinstance(file_info, dict):
                uploaded_files = file_info.get('uploaded_files', [])
                if uploaded_files and len(uploaded_files) > 0:
                    file_id = uploaded_files[0].get('file_id')
            
            if not file_id:
                print(f"   无法获取文件ID，上传结果: {file_info}")
                return None
            
            print(f"   文件上传成功，文件ID: {file_id}")
            
            # 尝试获取文件的访问URL
            try:
                file_info_result = Files.get(file_id=file_id)
                if file_info_result.status_code == 200:
                    file_info_data = file_info_result.output
                    # 检查是否有url字段
                    if isinstance(file_info_data, dict) and 'url' in file_info_data:
                        file_access_url = file_info_data['url']
                        print(f"   获取到文件访问URL: {file_access_url[:80]}...")
                    else:
                        file_access_url = None
                        print(f"   文件信息中无URL字段，使用file_id")
                else:
                    file_access_url = None
            except Exception as e:
                print(f"   获取文件URL失败: {e}，使用file_id")
                file_access_url = None
            
            # 尝试多种模型
            models_to_try = [
                'paraformer-realtime-v2',
                'paraformer-v2',
                'paraformer-zh',
                'paraformer',
                'fun-asr'
            ]
            
            for model_name in models_to_try:
                try:
                    # 尝试不同的URL格式
                    url_formats_to_try = []
                    if file_access_url:
                        url_formats_to_try.append(file_access_url)  # 优先使用获取到的URL
                    url_formats_to_try.extend([
                        file_id,  # 直接使用file_id
                        f"fileid://{file_id}",  # fileid协议
                        f"dashscope://{file_id}",  # dashscope协议
                    ])
                    
                    task_response = None
                    file_url_used = None
                    
                    for file_url_format in url_formats_to_try:
                        try:
                            print(f"   使用模型 {model_name}，URL格式: {file_url_format[:50]}...")
                            task_response = Transcription.async_call(
                                model=model_name,
                                file_urls=[file_url_format]
                            )
                            
                            if task_response.status_code == 200:
                                file_url_used = file_url_format
                                break
                            else:
                                print(f"     URL格式失败: {task_response.message}")
                        except Exception as url_error:
                            print(f"     URL格式错误: {url_error}")
                            continue
                    
                    if task_response and task_response.status_code == 200:
                        task_id = task_response.output.get('task_id')
                        if task_id:
                            print(f"   转写任务已提交，任务ID: {task_id}")
                            print(f"   等待转写完成（最多等待120秒）...")
                            
                            # 等待任务完成（wait方法会自动轮询）
                            try:
                                transcribe_response = Transcription.wait(task=task_id)
                                
                                if transcribe_response.status_code == 200:
                                    output = transcribe_response.output
                                    full_text = None
                                    
                                    # 提取文本内容
                                    if isinstance(output, dict):
                                        # 检查任务状态
                                        task_status = output.get('task_status', '').upper()
                                        if task_status == 'FAILED':
                                            error_msg = output.get('message', '未知错误')
                                            print(f"   转写任务失败: {error_msg}")
                                            continue
                                        
                                        # 方法1: 从sentence_list提取
                                        if 'sentence_list' in output:
                                            sentence_list = output['sentence_list']
                                            if isinstance(sentence_list, list) and len(sentence_list) > 0:
                                                texts = []
                                                for s in sentence_list:
                                                    if isinstance(s, dict):
                                                        texts.append(s.get('text', ''))
                                                    else:
                                                        texts.append(str(s))
                                                full_text = ' '.join(texts)
                                        
                                        # 方法2: 从sentences提取
                                        if not full_text and 'sentences' in output:
                                            sentences = output['sentences']
                                            if isinstance(sentences, list) and len(sentences) > 0:
                                                texts = []
                                                for s in sentences:
                                                    if isinstance(s, dict):
                                                        texts.append(s.get('text', ''))
                                                    else:
                                                        texts.append(str(s))
                                                full_text = ' '.join(texts)
                                        
                                        # 方法3: 直接获取text
                                        if not full_text and 'text' in output:
                                            full_text = output['text']
                                        
                                        # 方法4: 从results提取（需要从transcription_url获取）
                                        if not full_text and 'results' in output:
                                            results = output['results']
                                            if isinstance(results, list) and len(results) > 0:
                                                # 检查是否有transcription_url，需要下载转写结果
                                                for r in results:
                                                    if isinstance(r, dict):
                                                        transcription_url = r.get('transcription_url')
                                                        if transcription_url:
                                                            print(f"   找到transcription_url，正在下载转写结果...")
                                                            # 从transcription_url下载转写结果
                                                            try:
                                                                response = requests.get(transcription_url, timeout=30)
                                                                if response.status_code == 200:
                                                                    transcription_data = response.json()
                                                                    print(f"   已下载转写数据")
                                                                    print(f"   转写数据结构类型: {type(transcription_data)}")
                                                                    if isinstance(transcription_data, dict):
                                                                        print(f"   转写数据keys: {list(transcription_data.keys())[:10]}")
                                                                    # 尝试提取文本
                                                                    if isinstance(transcription_data, dict):
                                                                        # 方法1: 从transcripts字段提取（fun-asr模型的格式）
                                                                        if 'transcripts' in transcription_data:
                                                                            transcripts = transcription_data['transcripts']
                                                                            if isinstance(transcripts, list) and len(transcripts) > 0:
                                                                                # transcripts可能是字符串列表或字典列表
                                                                                texts = []
                                                                                for t in transcripts:
                                                                                    if isinstance(t, dict):
                                                                                        # 尝试不同的字段名
                                                                                        text = t.get('text') or t.get('transcript') or t.get('content') or str(t)
                                                                                        if text:
                                                                                            texts.append(text)
                                                                                    elif isinstance(t, str):
                                                                                        texts.append(t)
                                                                                full_text = ' '.join(texts)
                                                                        # 方法2: 从sentences字段提取
                                                                        elif 'sentences' in transcription_data:
                                                                            sentences = transcription_data['sentences']
                                                                            if isinstance(sentences, list) and len(sentences) > 0:
                                                                                texts = []
                                                                                for s in sentences:
                                                                                    if isinstance(s, dict):
                                                                                        texts.append(s.get('text', ''))
                                                                                    else:
                                                                                        texts.append(str(s))
                                                                                full_text = ' '.join(texts)
                                                                        # 方法3: 从sentence_list字段提取
                                                                        elif 'sentence_list' in transcription_data:
                                                                            sentence_list = transcription_data['sentence_list']
                                                                            if isinstance(sentence_list, list) and len(sentence_list) > 0:
                                                                                texts = []
                                                                                for s in sentence_list:
                                                                                    if isinstance(s, dict):
                                                                                        texts.append(s.get('text', ''))
                                                                                    else:
                                                                                        texts.append(str(s))
                                                                                full_text = ' '.join(texts)
                                                                        # 方法4: 直接获取text字段
                                                                        elif 'text' in transcription_data:
                                                                            full_text = transcription_data['text']
                                                                        
                                                                        if full_text and full_text.strip():
                                                                            break
                                                                    elif isinstance(transcription_data, list):
                                                                        # 如果直接是列表
                                                                        texts = []
                                                                        for item in transcription_data:
                                                                            if isinstance(item, dict):
                                                                                texts.append(item.get('text', ''))
                                                                            else:
                                                                                texts.append(str(item))
                                                                        full_text = ' '.join(texts)
                                                                        if full_text:
                                                                            break
                                                            except Exception as download_error:
                                                                print(f"   下载转写结果失败: {download_error}")
                                                                import traceback
                                                                traceback.print_exc()
                                                                continue
                                        
                                        # 方法5: 直接从results中查找text字段
                                        if not full_text and 'results' in output:
                                            results = output['results']
                                            if isinstance(results, list) and len(results) > 0:
                                                for r in results:
                                                    if isinstance(r, dict):
                                                        if 'text' in r:
                                                            full_text = r['text']
                                                            break
                                                        # 检查其他可能的文本字段
                                                        for key in ['transcription', 'content', 'transcript']:
                                                            if key in r and r[key]:
                                                                full_text = r[key]
                                                                break
                                                        if full_text:
                                                            break
                                    
                                    if full_text and full_text.strip():
                                        print(f"   转写完成，文本长度: {len(full_text)} 字符")
                                        return full_text
                                    else:
                                        # 打印完整的输出结构用于调试
                                        import json
                                        print(f"   转写完成但未提取到文本")
                                        print(f"   输出结构: {json.dumps(output, indent=2, ensure_ascii=False)[:500]}")
                                        
                                        # 尝试从output的所有可能字段中查找文本
                                        for key in output.keys():
                                            value = output[key]
                                            if isinstance(value, str) and len(value) > 10:
                                                print(f"   发现可能的文本字段 {key}: {value[:100]}")
                                            elif isinstance(value, list) and len(value) > 0:
                                                print(f"   发现列表字段 {key}: {type(value[0]) if value else 'empty'}")
                                                if isinstance(value[0], dict):
                                                    print(f"      第一个元素keys: {list(value[0].keys())[:5]}")
                                else:
                                    print(f"   转写任务失败: {transcribe_response.message}")
                            except Exception as wait_error:
                                print(f"   等待转写结果时出错: {wait_error}")
                                continue
                        else:
                            print(f"   无法获取任务ID")
                    else:
                        print(f"   提交任务失败: {task_response.message}")
                            
                except Exception as model_error:
                    print(f"   尝试模型 {model_name} 失败: {model_error}")
                    continue
            
            print(f"   所有模型尝试均失败")
            return None
                
        except ImportError:
            print("   错误：未安装 dashscope，请运行: pip install dashscope")
            return None
        except Exception as e:
            print(f"   转写音频时出错: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def ai_quality_check(self, advisor_name, transcript):
        """使用阿里云通义千问进行质检"""
        if not DASHSCOPE_API_KEY:
            raise ValueError("请先设置 DASHSCOPE_API_KEY")
        
        try:
            import dashscope
            from dashscope import Generation
            
            dashscope.api_key = DASHSCOPE_API_KEY
            
            system_prompt = "你是一个留学行业销冠教练。请分析这段销售通话录音文本。"
            
            user_prompt = f"""顾问姓名：{advisor_name}
通话内容：{transcript}

请从以下维度打分（0-10分）并简评：
1. 开场白（是否清晰、有身份锚点）
2. 需求挖掘（是否问出学校、专业、痛点）
3. 价值传递（是否提到国央企、实习等核心卖点）
4. 推进能力（是否尝试约面试/试听）

最后给出【未成交核心原因】和【改进建议】。"""
            
            # 构建完整的提示词（DashScope支持system+user格式）
            full_prompt = f"{system_prompt}\n\n{user_prompt}"
            
            print(f"   正在调用阿里云通义千问进行AI质检...")
            
            # 调用通义千问模型
            models_to_try = [
                'qwen-turbo',      # 快速模型
                'qwen-plus',       # 增强模型
                'qwen-max',        # 最强模型
                'qwen-max-longcontext'  # 长文本模型
            ]
            
            for model_name in models_to_try:
                try:
                    response = Generation.call(
                        model=model_name,
                        prompt=full_prompt,
                        temperature=0.7,
                        max_tokens=2000
                    )
                    
                    if response.status_code == 200:
                        ai_result = response.output.text
                        print(f"   AI质检完成（使用模型: {model_name}）")
                        return ai_result
                    else:
                        print(f"   模型 {model_name} 调用失败: {response.message}")
                        continue
                except Exception as model_error:
                    print(f"   模型 {model_name} 调用出错: {model_error}")
                    continue
            
            print(f"   所有模型尝试均失败")
            return None
            
        except ImportError:
            print("   错误：未安装 dashscope，请运行: pip install dashscope")
            return None
        except Exception as e:
            print(f"   AI质检时出错: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def generate_excel_report(self):
        """生成Excel质检报告"""
        if not self.results:
            print("   没有数据可生成报告")
            return
        
        print(f"[6] 正在生成Excel报告...")
        
        # 准备数据
        data = []
        for result in self.results:
            data.append({
                '顾问姓名': result['advisor'],
                '录音日期': result['date'],
                '音频文件名': result['filename'],
                '通话文字摘要': result['transcript'][:200] + '...' if result['transcript'] and len(result['transcript']) > 200 else (result['transcript'] or '无'),
                'AI评分与点评': result['ai_result'] or '质检失败',
                '状态': result['status']
            })
        
        df = pd.DataFrame(data)
        
        # 保存Excel
        output_file = f"质检日报_{self.yesterday}.xlsx"
        df.to_excel(output_file, index=False, engine='openpyxl')
        
        # 美化Excel格式，确保内容完整显示
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
            
            wb = load_workbook(output_file)
            ws = wb.active
            
            # 设置表头样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # 设置列宽
            column_widths = {
                'A': 12,  # 顾问姓名
                'B': 12,  # 录音日期
                'C': 30,  # 音频文件名
                'D': 50,  # 通话文字摘要
                'E': 80,  # AI评分与点评（最重要的列，设置较宽）
                'F': 10,  # 状态
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # 设置所有数据行的格式
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    # 启用自动换行
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    # 设置垂直对齐为顶部，这样多行文本会从顶部开始显示
            
            # 设置行高（自动调整以适应内容）
            ws.row_dimensions[1].height = 25  # 表头行高
            
            # 保存格式化的Excel
            wb.save(output_file)
            print(f"   Excel格式已优化，AI质检结果将完整显示")
        except Exception as e:
            print(f"   优化Excel格式时出错: {e}，但文件已保存")
        
        print(f"   报告已生成: {output_file}")
        return output_file
    
    def run(self):
        """执行完整流程"""
        print("=" * 60)
        print("先发系统销售录音质检自动化工具")
        print("=" * 60)
        
        # 检查API密钥
        if not DASHSCOPE_API_KEY:
            print("警告：未设置 DASHSCOPE_API_KEY，语音转文字和AI质检功能将无法使用")
        else:
            print(f"API密钥已配置，使用阿里云DashScope服务")
        
        browser = None
        playwright = None
        
        try:
            # 1. 连接浏览器
            browser, playwright = self.connect_browser()
            
            # 1.5. 检查登录状态
            self.check_login_status()
            
            # 2. 设置日期筛选
            self.set_date_filter()
            
            # 3. 随机抽取顾问
            selected_advisors = random.sample(ADVISOR_LIST, min(RANDOM_SAMPLE_COUNT, len(ADVISOR_LIST)))
            print(f"[3] 随机抽取 {len(selected_advisors)} 位顾问: {', '.join(selected_advisors)}")
            
            # 4. 循环处理每个顾问
            print(f"[4] 开始下载录音...")
            for i, advisor in enumerate(selected_advisors, 1):
                print(f"\n--- 处理第 {i}/{len(selected_advisors)} 位顾问: {advisor} ---")
                
                # 查询顾问
                if not self.query_advisor(advisor):
                    self.results.append({
                        'advisor': advisor,
                        'date': self.yesterday,
                        'filename': None,
                        'transcript': None,
                        'ai_result': None,
                        'status': '查询失败'
                    })
                    continue
                
                # 额外等待，确保查询结果已加载
                self.page.wait_for_timeout(3000)
                
                # 确保在查询页面
                current_url = self.page.url
                if '/conversationList' not in current_url:
                    print(f"   确保在查询页面...")
                    self.page.goto("https://www.xianfasj.com/callControl/conversationList", wait_until='domcontentloaded', timeout=10000)
                    self.page.wait_for_timeout(2000)
                
                # 下载录音（内部会验证是否属于该顾问）
                audio_path = self.download_first_recording(advisor)
                
                if not audio_path:
                    self.results.append({
                        'advisor': advisor,
                        'date': self.yesterday,
                        'filename': None,
                        'transcript': None,
                        'ai_result': None,
                        'status': '无通话记录'
                    })
                    continue
                
                # 5. 语音转文字
                print(f"[5-{i}] 语音转文字...")
                transcript = None
                ai_result = None
                
                if DASHSCOPE_API_KEY:
                    transcript = self.audio_to_text(audio_path)
                    
                    # 6. AI质检
                    if transcript:
                        print(f"[6-{i}] AI智能质检...")
                        ai_result = self.ai_quality_check(advisor, transcript)
                else:
                    transcript = "（未配置API密钥，跳过转写）"
                    ai_result = "（未配置API密钥，跳过质检）"
                
                # 记录结果
                self.results.append({
                    'advisor': advisor,
                    'date': self.yesterday,
                    'filename': audio_path.name if audio_path else None,
                    'transcript': transcript,
                    'ai_result': ai_result,
                    'status': '完成'
                })
                
                # 稍作延迟，避免请求过快
                time.sleep(1)
            
            # 7. 生成报告
            self.generate_excel_report()
            
            print("\n" + "=" * 60)
            print("流程执行完成！")
            print(f"成功处理: {sum(1 for r in self.results if r['status'] == '完成')} 位顾问")
            print(f"无记录: {sum(1 for r in self.results if r['status'] == '无通话记录')} 位顾问")
            print("=" * 60)
            
        except Exception as e:
            print(f"\n错误: {e}")
            import traceback
            traceback.print_exc()
        finally:
            # 不关闭浏览器，因为它是已打开的
            if playwright:
                playwright.stop()


if __name__ == "__main__":
    checker = CallQualityChecker()
    checker.run()

