#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
校招内推码爬虫 - 完整版
功能：
1. 爬取牛客网、V2EX、GitHub、知乎、掘金、微信公众号的校招内推信息
2. 使用正则表达式提取：公司名、岗位、内推码、内推链接、内推人、截止日期
3. 去重后保存到Excel文件，按公司名称排序
4. 支持每天定时自动运行更新

运行方式：
python referral_crawler.py          # 立即爬取一次
python referral_crawler.py --auto   # 启动每日自动更新模式
"""

import re
import time
import random
import logging
import argparse
from datetime import datetime, timedelta
from typing import List, Dict, Optional
from urllib.parse import urljoin, urlparse, parse_qs
import hashlib

import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import schedule

# ==================== 配置区 ====================
# 通义千问API配置（可选，如果使用AI提取）
TONGYI_API_KEY = ""  # 如果为空，则使用正则表达式提取
TONGYI_API_URL = "https://dashscope.aliyuncs.com/api/v1/services/aigc/text-generation/generation"

# 爬取配置
CRAWL_DELAY_MIN = 2  # 最小延迟（秒）
CRAWL_DELAY_MAX = 3  # 最大延迟（秒）
MAX_PAGES_PER_SOURCE = 5  # 每个来源最多爬取页数
TIMEOUT = 30  # 请求超时时间（秒）

# 输出文件
OUTPUT_FILE = "referral_data.xlsx"

# 日志配置
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('referral_crawler.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# User-Agent列表
USER_AGENTS = [
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
]

# ==================== 工具函数 ====================

def get_random_headers():
    """获取随机请求头"""
    return {
        'User-Agent': random.choice(USER_AGENTS),
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'none',
    }

def random_delay():
    """随机延迟"""
    time.sleep(random.uniform(CRAWL_DELAY_MIN, CRAWL_DELAY_MAX))

def safe_request(url: str, headers: dict = None, max_retries: int = 3) -> Optional[requests.Response]:
    """安全请求，带重试机制"""
    if headers is None:
        headers = get_random_headers()
    
    for attempt in range(max_retries):
        try:
            response = requests.get(url, headers=headers, timeout=TIMEOUT)
            if response.status_code == 200:
                return response
            elif response.status_code == 403:
                logger.warning(f"访问被拒绝: {url}, 尝试更换User-Agent")
                headers['User-Agent'] = random.choice(USER_AGENTS)
                time.sleep(5)
            else:
                logger.warning(f"请求失败: {url}, 状态码: {response.status_code}")
        except Exception as e:
            logger.error(f"请求出错 (尝试 {attempt+1}/{max_retries}): {url}, 错误: {str(e)}")
            if attempt < max_retries - 1:
                time.sleep(3)
    
    return None

def generate_unique_id(data: Dict) -> str:
    """生成唯一ID用于去重"""
    key_str = f"{data.get('公司', '')}_{data.get('内推码', '')}_{data.get('岗位', '')}"
    return hashlib.md5(key_str.encode('utf-8')).hexdigest()

# ==================== 数据提取器 ====================

class DataExtractor:
    """数据提取器 - 使用正则表达式提取信息"""
    
    @staticmethod
    def extract_company(text: str) -> Optional[str]:
        """提取公司名称"""
        if not text:
            return None
        
        # 公司名模式
        patterns = [
            r'([A-Za-z0-9\u4e00-\u9fa5]{2,20}(?:公司|集团|科技|有限|股份|企业|有限公司|股份有限公司))\s*[内推|校招|招聘]',
            r'【([^】]{2,20}?)(?:公司|集团|科技|有限|股份)?】',
            r'([A-Za-z0-9\u4e00-\u9fa5]{2,20})\s*内推',
            r'([A-Za-z0-9\u4e00-\u9fa5]{2,20})\s*校招',
            r'([A-Za-z0-9\u4e00-\u9fa5]{2,20})\s*招聘',
            r'\(([^)]{2,20}?)(?:公司|集团|科技)?\)',
            r'（([^）]{2,20}?)(?:公司|集团|科技)?）',
            r'@([A-Za-z0-9\u4e00-\u9fa5]{2,20})',
            r'#([A-Za-z0-9\u4e00-\u9fa5]{2,20})#',
        ]
        
        search_text = text[:500]  # 只搜索前500字符
        for pattern in patterns:
            match = re.search(pattern, search_text)
            if match:
                company = match.group(1).strip()
                company = re.sub(r'[：:：]', '', company)
                if 2 <= len(company) <= 20:
                    exclude_words = ['内推码', '校招', '招聘', '推荐', '岗位', '职位', '公司', '内推']
                    if not any(word in company for word in exclude_words):
                        return company
        
        # 从文本开头提取
        lines = text.split('\n')
        for line in lines[:3]:
            line = line.strip()[:40]
            if not line:
                continue
            line_clean = re.sub(r'[【】（）()\[\]@#：:：]', '', line)
            if 2 <= len(line_clean) <= 20:
                exclude_starts = ['内推', '推荐', '招聘', '需要', '急招', '岗位', '职位']
                if not any(line_clean.startswith(word) for word in exclude_starts):
                    return line_clean
        
        return None
    
    @staticmethod
    def extract_position(text: str) -> Optional[str]:
        """提取岗位名称"""
        if not text:
            return None
        
        patterns = [
            r'招聘[：:\s]*([^内推码\n]{2,30}?)(?:内推|校招|岗位|职位)',
            r'岗位[：:\s]*([^\n]{2,30})',
            r'职位[：:\s]*([^\n]{2,30})',
            r'([A-Za-z0-9\u4e00-\u9fa5]{2,30}(?:工程师|开发|算法|产品|运营|设计|测试|前端|后端|Java|Python|C\+\+|Go|PHP))',
            r'【([^】]{2,30}?)(?:工程师|开发|算法|产品|运营|设计|测试)】',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text[:500])
            if match:
                position = match.group(1).strip()
                if 2 <= len(position) <= 30:
                    return position
        
        return None
    
    @staticmethod
    def extract_referral_code(text: str) -> Optional[str]:
        """提取内推码"""
        if not text:
            return None
        
        patterns = [
            r'内推码[：:\s]+([A-Za-z0-9]{4,20})',
            r'推荐码[：:\s]+([A-Za-z0-9]{4,20})',
            r'内推[：:\s]+([A-Za-z0-9]{4,20})',
            r'推荐[：:\s]+([A-Za-z0-9]{4,20})',
            r'code[：:\s]+([A-Za-z0-9]{4,20})',
            r'CODE[：:\s]+([A-Za-z0-9]{4,20})',
            r'Referral[：:\s]+([A-Za-z0-9]{4,20})',
            r'[（(]内推码[：:\s]*([A-Za-z0-9]{4,20})[）)]',
            r'[（(]推荐码[：:\s]*([A-Za-z0-9]{4,20})[）)]',
            r'\b([A-Z]{2,4}[0-9]{4,8})\b',
            r'\b([A-Z]{4,8}[0-9]{2,4})\b',
            r'\b([0-9]{6,10}[A-Z]{2,4})\b',
            r'\b([A-Z]{5,12})\b',
        ]
        
        found_codes = []
        for pattern in patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            for match in matches:
                code = match.strip().upper()
                if 4 <= len(code) <= 20 and not re.match(r'^[0-9]+$', code):
                    if code not in found_codes:
                        found_codes.append(code)
        
        # 优先返回明确标签的
        labeled_patterns = [
            r'内推码[：:\s]+([A-Za-z0-9]{4,20})',
            r'推荐码[：:\s]+([A-Za-z0-9]{4,20})',
            r'内推[：:\s]+([A-Za-z0-9]{4,20})',
        ]
        for pattern in labeled_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                code = match.group(1).strip().upper()
                if 4 <= len(code) <= 20:
                    return code
        
        return found_codes[0] if found_codes else None
    
    @staticmethod
    def extract_referral_link(text: str) -> Optional[str]:
        """提取内推链接"""
        if not text:
            return None
        
        # URL模式
        url_pattern = r'https?://[^\s\n<>"\'\)]+'
        urls = re.findall(url_pattern, text)
        
        # 优先查找包含内推、校招、招聘等关键词的链接
        for url in urls:
            if any(keyword in url.lower() for keyword in ['referral', 'campus', '校招', '内推', '招聘', 'job', 'career']):
                return url
        
        # 如果没有找到相关链接，返回第一个URL
        return urls[0] if urls else None
    
    @staticmethod
    def extract_referrer(text: str) -> Optional[str]:
        """提取内推人"""
        if not text:
            return None
        
        patterns = [
            r'内推人[：:\s]*([^\n]{2,20})',
            r'推荐人[：:\s]*([^\n]{2,20})',
            r'联系[：:\s]*([^\n]{2,20})',
            r'微信[：:\s]*([a-zA-Z0-9_-]{3,20})',
            r'QQ[：:\s]*([0-9]{5,12})',
            r'邮箱[：:\s]*([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text[:300])
            if match:
                referrer = match.group(1).strip()
                if 2 <= len(referrer) <= 50:
                    return referrer
        
        return None
    
    @staticmethod
    def extract_contact(text: str) -> Optional[str]:
        """提取联系方式"""
        if not text:
            return None
        
        patterns = [
            r'微信[：:\s]*([a-zA-Z0-9_-]{3,20})',
            r'QQ[：:\s]*([0-9]{5,12})',
            r'邮箱[：:\s]*([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})',
            r'电话[：:\s]*([0-9-]{7,15})',
            r'联系[：:\s]*([^\n]{2,30})',
        ]
        
        contacts = []
        for pattern in patterns:
            matches = re.findall(pattern, text[:500])
            for match in matches:
                contact = match.strip()
                if contact and contact not in contacts:
                    contacts.append(contact)
        
        return ' / '.join(contacts) if contacts else None
    
    @staticmethod
    def extract_deadline(text: str) -> Optional[str]:
        """提取截止日期"""
        if not text:
            return None
        
        patterns = [
            r'截止[：:\s]*(\d{4}[-/]\d{1,2}[-/]\d{1,2})',
            r'截止[：:\s]*(\d{1,2}[-/]\d{1,2}[-/]\d{4})',
            r'截止[：:\s]*(\d{4}年\d{1,2}月\d{1,2}日)',
            r'截止[：:\s]*(\d{1,2}月\d{1,2}日)',
            r'有效期[：:\s]*(\d{4}[-/]\d{1,2}[-/]\d{1,2})',
            r'(\d{4}[-/]\d{1,2}[-/]\d{1,2})\s*截止',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text[:500])
            if match:
                date_str = match.group(1)
                # 标准化日期格式
                try:
                    if '年' in date_str:
                        date_str = date_str.replace('年', '-').replace('月', '-').replace('日', '')
                    date_str = re.sub(r'[/-]', '-', date_str)
                    parts = date_str.split('-')
                    if len(parts) == 3:
                        if len(parts[0]) == 4:  # YYYY-MM-DD
                            return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
                        else:  # DD-MM-YYYY
                            return f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
                except:
                    pass
                return date_str
        
        return None
    
    @staticmethod
    def extract_all(text: str, source: str = "") -> Dict:
        """提取所有信息"""
        if not text or len(text.strip()) < 10:
            return {}
        
        # 必须包含内推相关关键词
        if not any(keyword in text for keyword in ['内推', '推荐', '校招', '招聘', 'referral']):
            return {}
        
        data = {
            '公司': DataExtractor.extract_company(text),
            '岗位': DataExtractor.extract_position(text),
            '内推码': DataExtractor.extract_referral_code(text),
            '内推链接': DataExtractor.extract_referral_link(text),
            '内推人': DataExtractor.extract_referrer(text),
            '联系方式': DataExtractor.extract_contact(text),
            '截止日期': DataExtractor.extract_deadline(text),
            '来源': source,
            '更新时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }
        
        # 至少要有公司或内推码才认为是有效数据
        if not data['公司'] and not data['内推码']:
            return {}
        
        return data

# ==================== 爬虫类 ====================

class ReferralCrawler:
    """校招内推码爬虫主类"""
    
    def __init__(self):
        self.results: List[Dict] = []
        self.seen_ids: set = set()
        self.extractor = DataExtractor()
    
    def crawl_nowcoder(self) -> List[Dict]:
        """爬取牛客网"""
        logger.info("开始爬取牛客网...")
        results = []
        
        try:
            base_url = "https://www.nowcoder.com/search"
            params = {
                'type': 'post',
                'query': '内推码',
                'order': 'time',  # 按时间排序
            }
            
            for page in range(1, MAX_PAGES_PER_SOURCE + 1):
                params['page'] = page
                url = f"{base_url}?type={params['type']}&query={params['query']}&order={params['order']}&page={page}"
                
                logger.info(f"  爬取第 {page} 页: {url}")
                response = safe_request(url)
                
                if not response:
                    logger.warning(f"  无法访问第 {page} 页")
                    break
                
                soup = BeautifulSoup(response.text, 'html.parser')
                
                # 查找帖子列表（需要根据实际页面结构调整选择器）
                posts = soup.find_all(['div', 'article'], class_=re.compile(r'post|item|card|feed', re.I))
                
                if not posts:
                    # 尝试其他选择器
                    posts = soup.find_all('a', href=re.compile(r'/discuss'))
                
                if not posts:
                    logger.warning(f"  第 {page} 页未找到帖子内容")
                    break
                
                logger.info(f"  找到 {len(posts)} 个帖子")
                
                for post in posts:
                    try:
                        # 获取帖子文本
                        text = post.get_text(strip=True)
                        if not text or len(text) < 20:
                            continue
                        
                        # 提取数据
                        data = self.extractor.extract_all(text, "牛客网")
                        if data:
                            unique_id = generate_unique_id(data)
                            if unique_id not in self.seen_ids:
                                self.seen_ids.add(unique_id)
                                results.append(data)
                                logger.debug(f"  提取到: {data.get('公司')} - {data.get('内推码')}")
                    except Exception as e:
                        logger.error(f"  解析帖子时出错: {str(e)}")
                        continue
                
                random_delay()
                
                # 如果没有更多内容，停止
                if len(posts) < 10:
                    break
            
            logger.info(f"牛客网爬取完成，获得 {len(results)} 条数据")
        
        except Exception as e:
            logger.error(f"爬取牛客网时出错: {str(e)}")
        
        return results
    
    def crawl_v2ex(self) -> List[Dict]:
        """爬取V2EX职场节点"""
        logger.info("开始爬取V2EX...")
        results = []
        
        try:
            base_url = "https://www.v2ex.com/go/career"
            
            for page in range(1, MAX_PAGES_PER_SOURCE + 1):
                if page == 1:
                    url = base_url
                else:
                    url = f"{base_url}?p={page}"
                
                logger.info(f"  爬取第 {page} 页: {url}")
                response = safe_request(url)
                
                if not response:
                    logger.warning(f"  无法访问第 {page} 页")
                    break
                
                soup = BeautifulSoup(response.text, 'html.parser')
                
                # V2EX帖子选择器
                posts = soup.find_all('span', class_='item_title')
                
                if not posts:
                    logger.warning(f"  第 {page} 页未找到帖子")
                    break
                
                logger.info(f"  找到 {len(posts)} 个帖子")
                
                for post_title in posts:
                    try:
                        # 获取帖子链接
                        link_elem = post_title.find('a')
                        if not link_elem:
                            continue
                        
                        post_url = urljoin(base_url, link_elem.get('href', ''))
                        
                        # 访问帖子详情页
                        post_response = safe_request(post_url)
                        if not post_response:
                            continue
                        
                        post_soup = BeautifulSoup(post_response.text, 'html.parser')
                        content = post_soup.find('div', class_='topic_content')
                        
                        if not content:
                            continue
                        
                        text = content.get_text(strip=True)
                        if not text or len(text) < 20:
                            continue
                        
                        # 提取数据
                        data = self.extractor.extract_all(text, "V2EX")
                        if data:
                            unique_id = generate_unique_id(data)
                            if unique_id not in self.seen_ids:
                                self.seen_ids.add(unique_id)
                                results.append(data)
                                logger.debug(f"  提取到: {data.get('公司')} - {data.get('内推码')}")
                        
                        random_delay()
                    except Exception as e:
                        logger.error(f"  解析帖子时出错: {str(e)}")
                        continue
                
                random_delay()
                
                # 检查是否有下一页
                next_page = soup.find('a', string=re.compile(r'下一页|next', re.I))
                if not next_page:
                    break
            
            logger.info(f"V2EX爬取完成，获得 {len(results)} 条数据")
        
        except Exception as e:
            logger.error(f"爬取V2EX时出错: {str(e)}")
        
        return results
    
    def crawl_github(self) -> List[Dict]:
        """爬取GitHub内推仓库"""
        logger.info("开始爬取GitHub...")
        results = []
        
        try:
            # 常见的GitHub内推仓库（需要根据实际情况调整）
            github_repos = [
                "https://raw.githubusercontent.com/0voice/interview_internal_reference/main/README.md",
                # 可以添加更多GitHub仓库URL
            ]
            
            for repo_url in github_repos:
                logger.info(f"  爬取仓库: {repo_url}")
                response = safe_request(repo_url)
                
                if not response:
                    logger.warning(f"  无法访问: {repo_url}")
                    continue
                
                # GitHub的README通常是Markdown格式
                text = response.text
                
                # 按行分割，每行可能包含一条内推信息
                lines = text.split('\n')
                current_section = ""
                
                for line in lines:
                    # 跳过Markdown标题和空行
                    if line.startswith('#') or not line.strip():
                        continue
                    
                    # 累积文本直到遇到明显的分隔
                    if len(line.strip()) > 10:
                        current_section += line + "\n"
                    
                    # 如果累积的文本足够长，尝试提取
                    if len(current_section) > 100:
                        data = self.extractor.extract_all(current_section, "GitHub")
                        if data:
                            unique_id = generate_unique_id(data)
                            if unique_id not in self.seen_ids:
                                self.seen_ids.add(unique_id)
                                results.append(data)
                                logger.debug(f"  提取到: {data.get('公司')} - {data.get('内推码')}")
                        current_section = ""
                
                # 处理最后一段
                if current_section:
                    data = self.extractor.extract_all(current_section, "GitHub")
                    if data:
                        unique_id = generate_unique_id(data)
                        if unique_id not in self.seen_ids:
                            self.seen_ids.add(unique_id)
                            results.append(data)
                
                random_delay()
            
            logger.info(f"GitHub爬取完成，获得 {len(results)} 条数据")
        
        except Exception as e:
            logger.error(f"爬取GitHub时出错: {str(e)}")
        
        return results
    
    def crawl_zhihu(self) -> List[Dict]:
        """爬取知乎搜索'校招内推码'的回答"""
        logger.info("开始爬取知乎...")
        results = []
        
        try:
            # 知乎搜索API（需要根据实际情况调整）
            base_url = "https://www.zhihu.com/search"
            params = {
                'type': 'content',
                'q': '校招内推码',
            }
            
            for page in range(1, MAX_PAGES_PER_SOURCE + 1):
                url = f"{base_url}?type={params['type']}&q={params['q']}&page={page}"
                
                logger.info(f"  爬取第 {page} 页: {url}")
                response = safe_request(url)
                
                if not response:
                    logger.warning(f"  无法访问第 {page} 页")
                    break
                
                soup = BeautifulSoup(response.text, 'html.parser')
                
                # 知乎搜索结果选择器（需要根据实际页面结构调整）
                items = soup.find_all(['div', 'article'], class_=re.compile(r'item|result|answer|content', re.I))
                
                if not items:
                    # 尝试其他选择器
                    items = soup.find_all('a', href=re.compile(r'/question|/answer'))
                
                if not items:
                    logger.warning(f"  第 {page} 页未找到内容")
                    break
                
                logger.info(f"  找到 {len(items)} 个结果")
                
                for item in items:
                    try:
                        text = item.get_text(strip=True)
                        if not text or len(text) < 20:
                            continue
                        
                        # 提取数据
                        data = self.extractor.extract_all(text, "知乎")
                        if data:
                            unique_id = generate_unique_id(data)
                            if unique_id not in self.seen_ids:
                                self.seen_ids.add(unique_id)
                                results.append(data)
                                logger.debug(f"  提取到: {data.get('公司')} - {data.get('内推码')}")
                    except Exception as e:
                        logger.error(f"  解析内容时出错: {str(e)}")
                        continue
                
                random_delay()
                
                # 如果没有更多内容，停止
                if len(items) < 10:
                    break
            
            logger.info(f"知乎爬取完成，获得 {len(results)} 条数据")
        
        except Exception as e:
            logger.error(f"爬取知乎时出错: {str(e)}")
        
        return results
    
    def crawl_juejin(self) -> List[Dict]:
        """爬取掘金沸点的求职内推话题"""
        logger.info("开始爬取掘金...")
        results = []
        
        try:
            # 掘金沸点搜索
            base_url = "https://juejin.cn/search"
            params = {
                'query': '校招内推码',
                'type': 'all',
            }
            
            for page in range(1, MAX_PAGES_PER_SOURCE + 1):
                url = f"{base_url}?query={params['query']}&type={params['type']}&page={page}"
                
                logger.info(f"  爬取第 {page} 页: {url}")
                response = safe_request(url)
                
                if not response:
                    logger.warning(f"  无法访问第 {page} 页")
                    break
                
                soup = BeautifulSoup(response.text, 'html.parser')
                
                # 掘金内容选择器（需要根据实际页面结构调整）
                items = soup.find_all(['div', 'article'], class_=re.compile(r'item|pin|content|result', re.I))
                
                if not items:
                    # 尝试其他选择器
                    items = soup.find_all('a', href=re.compile(r'/pin|/post'))
                
                if not items:
                    logger.warning(f"  第 {page} 页未找到内容")
                    break
                
                logger.info(f"  找到 {len(items)} 个结果")
                
                for item in items:
                    try:
                        text = item.get_text(strip=True)
                        if not text or len(text) < 20:
                            continue
                        
                        # 提取数据
                        data = self.extractor.extract_all(text, "掘金")
                        if data:
                            unique_id = generate_unique_id(data)
                            if unique_id not in self.seen_ids:
                                self.seen_ids.add(unique_id)
                                results.append(data)
                                logger.debug(f"  提取到: {data.get('公司')} - {data.get('内推码')}")
                    except Exception as e:
                        logger.error(f"  解析内容时出错: {str(e)}")
                        continue
                
                random_delay()
                
                # 如果没有更多内容，停止
                if len(items) < 10:
                    break
            
            logger.info(f"掘金爬取完成，获得 {len(results)} 条数据")
        
        except Exception as e:
            logger.error(f"爬取掘金时出错: {str(e)}")
        
        return results
    
    def crawl_wechat(self) -> List[Dict]:
        """爬取微信公众号文章（通过搜狗微信搜索）"""
        logger.info("开始爬取微信公众号...")
        results = []
        
        try:
            # 搜狗微信搜索
            base_url = "https://weixin.sogou.com/weixin"
            params = {
                'type': '2',
                'query': '校招内推码',
            }
            
            for page in range(1, min(MAX_PAGES_PER_SOURCE, 3) + 1):  # 搜狗限制较多，只爬3页
                if page == 1:
                    url = f"{base_url}?type={params['type']}&query={params['query']}"
                else:
                    url = f"{base_url}?type={params['type']}&query={params['query']}&page={page}"
                
                logger.info(f"  爬取第 {page} 页: {url}")
                response = safe_request(url)
                
                if not response:
                    logger.warning(f"  无法访问第 {page} 页")
                    break
                
                soup = BeautifulSoup(response.text, 'html.parser')
                
                # 搜狗微信搜索结果选择器（需要根据实际页面结构调整）
                items = soup.find_all(['div', 'li'], class_=re.compile(r'item|result|news', re.I))
                
                if not items:
                    # 尝试其他选择器
                    items = soup.find_all('a', href=re.compile(r'/link'))
                
                if not items:
                    logger.warning(f"  第 {page} 页未找到内容")
                    break
                
                logger.info(f"  找到 {len(items)} 个结果")
                
                for item in items:
                    try:
                        # 获取文章摘要
                        text = item.get_text(strip=True)
                        if not text or len(text) < 20:
                            continue
                        
                        # 提取数据（从摘要中提取，可能信息不完整）
                        data = self.extractor.extract_all(text, "微信公众号")
                        if data:
                            unique_id = generate_unique_id(data)
                            if unique_id not in self.seen_ids:
                                self.seen_ids.add(unique_id)
                                results.append(data)
                                logger.debug(f"  提取到: {data.get('公司')} - {data.get('内推码')}")
                    except Exception as e:
                        logger.error(f"  解析内容时出错: {str(e)}")
                        continue
                
                random_delay()
                
                # 搜狗反爬较严格，只爬少量页面
                if page >= 3:
                    break
            
            logger.info(f"微信公众号爬取完成，获得 {len(results)} 条数据")
        
        except Exception as e:
            logger.error(f"爬取微信公众号时出错: {str(e)}")
        
        return results
    
    def crawl_all(self) -> List[Dict]:
        """爬取所有来源"""
        logger.info("="*60)
        logger.info("开始爬取所有来源的内推信息")
        logger.info("="*60)
        
        all_results = []
        
        # 爬取各个来源
        sources = [
            ("牛客网", self.crawl_nowcoder),
            ("V2EX", self.crawl_v2ex),
            ("GitHub", self.crawl_github),
            ("知乎", self.crawl_zhihu),
            ("掘金", self.crawl_juejin),
            ("微信公众号", self.crawl_wechat),
        ]
        
        for source_name, crawl_func in sources:
            try:
                results = crawl_func()
                all_results.extend(results)
                logger.info(f"{source_name} 完成，当前总计 {len(all_results)} 条数据")
            except Exception as e:
                logger.error(f"爬取 {source_name} 时出错: {str(e)}")
                continue
        
        logger.info("="*60)
        logger.info(f"所有来源爬取完成，共获得 {len(all_results)} 条数据")
        logger.info("="*60)
        
        return all_results

# ==================== Excel导出器 ====================

class ExcelExporter:
    """Excel导出器"""
    
    @staticmethod
    def save_to_excel(data: List[Dict], filename: str = OUTPUT_FILE):
        """保存数据到Excel"""
        if not data:
            logger.warning("没有数据可保存")
            return
        
        try:
            # 创建DataFrame
            df = pd.DataFrame(data)
            
            # 确保所有必需的列都存在
            required_columns = ['公司', '岗位', '内推码', '内推链接', '内推人', '联系方式', '截止日期', '来源', '更新时间']
            for col in required_columns:
                if col not in df.columns:
                    df[col] = ''
            
            # 按列顺序排列
            df = df[required_columns]
            
            # 填充空值
            df = df.fillna('')
            
            # 按公司名称排序
            df = df.sort_values('公司', ascending=True, na_position='last')
            
            # 保存到Excel
            df.to_excel(filename, index=False, engine='openpyxl')
            
            # 美化Excel格式
            wb = load_workbook(filename)
            ws = wb.active
            
            # 设置表头样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 设置列宽
            column_widths = {
                'A': 25,  # 公司
                'B': 30,  # 岗位
                'C': 20,  # 内推码
                'D': 40,  # 内推链接
                'E': 20,  # 内推人
                'F': 25,  # 联系方式
                'G': 15,  # 截止日期
                'H': 15,  # 来源
                'I': 20,  # 更新时间
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # 设置行高
            ws.row_dimensions[1].height = 25
            
            # 设置数据行样式
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            wb.save(filename)
            
            logger.info(f"数据已保存至: {filename}")
            logger.info(f"共 {len(df)} 条记录")
        
        except Exception as e:
            logger.error(f"保存Excel时出错: {str(e)}")
            import traceback
            traceback.print_exc()

# ==================== 定时任务 ====================

def run_crawler():
    """运行爬虫任务"""
    logger.info("\n" + "="*60)
    logger.info("开始执行爬虫任务")
    logger.info("="*60)
    
    try:
        crawler = ReferralCrawler()
        results = crawler.crawl_all()
        
        if results:
            # 合并已有数据（如果文件存在）
            try:
                existing_df = pd.read_excel(OUTPUT_FILE, engine='openpyxl')
                existing_data = existing_df.to_dict('records')
                
                # 去重合并
                existing_ids = {generate_unique_id(item) for item in existing_data}
                new_results = [r for r in results if generate_unique_id(r) not in existing_ids]
                
                all_results = existing_data + new_results
                logger.info(f"合并数据: 原有 {len(existing_data)} 条，新增 {len(new_results)} 条，总计 {len(all_results)} 条")
            except:
                all_results = results
                logger.info("未找到已有数据文件，使用新数据")
            
            # 保存到Excel
            ExcelExporter.save_to_excel(all_results)
            
            logger.info("="*60)
            logger.info("爬虫任务完成")
            logger.info("="*60)
        else:
            logger.warning("未获取到任何数据")
    
    except Exception as e:
        logger.error(f"执行爬虫任务时出错: {str(e)}")
        import traceback
        traceback.print_exc()

# ==================== 主函数 ====================

def main():
    """主函数"""
    parser = argparse.ArgumentParser(description='校招内推码爬虫')
    parser.add_argument('--auto', action='store_true', help='启动每日自动更新模式')
    args = parser.parse_args()
    
    if args.auto:
        logger.info("="*60)
        logger.info("启动定时任务模式")
        logger.info("="*60)
        logger.info("将在每天 09:00 自动运行爬虫")
        logger.info("按 Ctrl+C 停止")
        logger.info("="*60)
        
        # 立即运行一次
        run_crawler()
        
        # 设置定时任务：每天09:00运行
        schedule.every().day.at("09:00").do(run_crawler)
        
        # 也可以设置其他时间，例如每小时运行一次（用于测试）
        # schedule.every().hour.do(run_crawler)
        
        try:
            while True:
                schedule.run_pending()
                time.sleep(60)  # 每分钟检查一次
        except KeyboardInterrupt:
            logger.info("\n定时任务已停止")
    else:
        # 立即运行一次
        run_crawler()

if __name__ == '__main__':
    main()

