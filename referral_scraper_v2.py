#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
每日内推信息自动抓取与报告生成系统 v2.0
功能：
1. 自动抓取牛客网、知乎等平台的最新内推信息
2. 生成Markdown格式的日报
3. 更新Excel汇总表
4. 提取Top 3捡漏机会
"""

import time
import re
from datetime import datetime, timedelta
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

class EnhancedReferralScraper:
    def __init__(self):
        """初始化爬虫"""
        self.chrome_options = Options()
        user_data_dir = os.path.expanduser('~/Library/Application Support/Google/Chrome')
        self.chrome_options.add_argument(f'--user-data-dir={user_data_dir}')
        self.chrome_options.add_argument('--profile-directory=Default')
        self.chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        self.chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])
        self.chrome_options.add_experimental_option('useAutomationExtension', False)
        
        self.driver = None
        self.results = []
        self.today = datetime.now().strftime('%Y-%m-%d')
        
        # 重点公司列表
        self.key_companies = [
            '字节跳动', '腾讯', '阿里巴巴', '阿里', '百度', '美团', '京东',
            '拼多多', '小米', '华为', '网易', '滴滴', '快手', 'B站', 'bilibili',
            '小红书', '蔚来', '理想', '比亚迪', '大疆', '海康威视', '科大讯飞',
            '米哈游', '基恩士', '施耐德', '博世', 'OPPO', 'vivo', '荣耀'
        ]
        
    def start_driver(self):
        """启动浏览器"""
        print("\n" + "="*60)
        print("每日内推信息自动抓取系统 v2.0")
        print("="*60)
        print(f"\n开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("正在启动Chrome浏览器...")
        
        try:
            self.driver = webdriver.Chrome(options=self.chrome_options)
            self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            print("✓ 浏览器启动成功！")
        except Exception as e:
            print(f"✗ 浏览器启动失败: {str(e)}")
            print("\n请确保：")
            print("1. 已安装Chrome浏览器")
            print("2. 已安装ChromeDriver (brew install chromedriver)")
            raise
    
    def scrape_nowcoder(self, keywords=['内推码', '2026校招', '内推', '急招']):
        """抓取牛客网内推信息"""
        print("\n" + "-"*60)
        print("[1/3] 正在抓取牛客网...")
        print("-"*60)
        
        base_urls = [
            "https://www.nowcoder.com/discuss/experience",
            "https://www.nowcoder.com/discuss/tag/639",  # 内推标签
        ]
        
        try:
            for keyword in keywords:
                try:
                    search_url = f"https://www.nowcoder.com/search?query={keyword}&type=discuss"
                    print(f"\n搜索关键词: {keyword}")
                    self.driver.get(search_url)
                    time.sleep(3)
                    
                    # 获取帖子列表
                    posts = self.driver.find_elements(By.CSS_SELECTOR, '.discuss-item, .feed-item, .post-item')
                    print(f"找到 {len(posts)} 个帖子")
                    
                    count = 0
                    for post in posts[:15]:  # 取前15条
                        try:
                            # 提取标题和链接
                            title_elem = post.find_element(By.CSS_SELECTOR, 'a.discuss-title, a.feed-title, .title a')
                            title = title_elem.text.strip()
                            link = title_elem.get_attribute('href')
                            
                            # 检查是否包含内推相关关键词
                            if not any(kw in title for kw in ['内推', '校招', '招聘']):
                                continue
                            
                            # 提取发布时间
                            try:
                                time_elem = post.find_element(By.CSS_SELECTOR, '.time, .post-time, .feed-time')
                                post_time = time_elem.text
                            except:
                                post_time = "未知"
                            
                            # 检查是否为近期发布（简单判断）
                            if self.is_recent(post_time):
                                company = self.extract_company_name(title)
                                referral_code = self.extract_referral_code(title)
                                
                                self.results.append({
                                    '公司名称': company,
                                    '岗位/方向': '校招',
                                    '招聘类型': '校招',
                                    '内推码/直推邮箱': referral_code,
                                    '投递链接/来源': link,
                                    '备注': f"发布时间: {post_time}",
                                    '来源平台': '牛客网',
                                    '抓取时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                                    '标题': title
                                })
                                count += 1
                                print(f"  ✓ [{count}] {company} - {referral_code}")
                        except Exception as e:
                            continue
                    
                    print(f"本次搜索共抓取 {count} 条有效信息")
                    time.sleep(2)
                    
                except Exception as e:
                    print(f"  ✗ 搜索关键词 '{keyword}' 时出错: {str(e)}")
                    continue
            
            print(f"\n牛客网抓取完成，共获取 {len(self.results)} 条信息")
                    
        except Exception as e:
            print(f"\n✗ 抓取牛客网时出错: {str(e)}")
    
    def is_recent(self, time_str):
        """判断是否为近期发布（简单判断）"""
        recent_keywords = ['今天', '小时', '分钟', '刚刚', '昨天', '1天前', '2天前']
        return any(kw in time_str for kw in recent_keywords)
    
    def extract_referral_code(self, text):
        """从文本中提取内推码"""
        patterns = [
            r'内推码[：:](\w+)',
            r'推荐码[：:](\w+)',
            r'内推[：:]\s*(\w+)',
            r'码[：:](\w+)',
            r'\b[A-Z0-9]{6,10}\b',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                code = match.group(1) if match.lastindex else match.group(0)
                # 过滤掉一些常见的非内推码
                if code not in ['JAVA', 'PYTHON', 'HTTP', 'HTTPS']:
                    return code
        
        return "见原文"
    
    def extract_company_name(self, text):
        """从文本中提取公司名称"""
        for company in self.key_companies:
            if company in text:
                return company
        return "其他公司"
    
    def generate_markdown_report(self, filename=None):
        """生成Markdown格式的日报"""
        if not self.results:
            print("\n没有数据可生成报告！")
            return
        
        if filename is None:
            filename = f"今日内推汇总_{self.today}.md"
        
        desktop_path = os.path.expanduser('~/Desktop')
        file_path = os.path.join(desktop_path, filename)
        
        # 数据处理
        df = pd.DataFrame(self.results)
        df = df.drop_duplicates(subset=['公司名称', '内推码/直推邮箱'])
        
        # 按公司重要性排序
        def company_priority(company):
            if company in self.key_companies[:10]:  # 前10家重点公司
                return 0
            elif company in self.key_companies:
                return 1
            else:
                return 2
        
        df['priority'] = df['公司名称'].apply(company_priority)
        df = df.sort_values('priority')
        
        # 生成Markdown内容
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(f"# 今日最新内推/直推岗位汇总表\n\n")
            f.write(f"**更新日期**: {self.today}  \n")
            f.write(f"**数据来源**: 牛客网等平台  \n")
            f.write(f"**抓取时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  \n")
            f.write(f"**有效信息**: {len(df)} 条\n\n")
            f.write("---\n\n")
            
            # 表格
            f.write("## 📊 内推信息汇总\n\n")
            f.write("| 公司名称 | 岗位/方向 | 招聘类型 | 内推码/直推邮箱 | 投递链接/来源 | 备注 |\n")
            f.write("|---------|----------|---------|----------------|--------------|------|\n")
            
            for _, row in df.iterrows():
                f.write(f"| {row['公司名称']} | {row['岗位/方向']} | {row['招聘类型']} | ")
                f.write(f"{row['内推码/直推邮箱']} | {row['投递链接/来源']} | {row['备注']} |\n")
            
            # Top 3捡漏机会
            f.write("\n---\n\n")
            f.write("## 🔥 今日Top 3捡漏机会\n\n")
            
            top_opportunities = self.identify_top_opportunities(df)
            for i, opp in enumerate(top_opportunities, 1):
                f.write(f"### {i}. {opp['emoji']} {opp['company']}\n\n")
                f.write(f"**亮点**: {opp['highlight']}  \n")
                f.write(f"**内推码**: {opp['code']}  \n")
                f.write(f"**投递链接**: {opp['link']}  \n")
                f.write(f"**备注**: {opp['note']}\n\n")
                f.write("---\n\n")
            
            # 使用建议
            f.write("## 📝 使用建议\n\n")
            f.write("1. **尽早投递**: 大部分公司采用先到先得原则\n")
            f.write("2. **多平台尝试**: 不要只依赖一个内推码\n")
            f.write("3. **简历优化**: 使用内推码前先优化简历\n")
            f.write("4. **关注时效**: 部分内推码有使用期限\n")
            f.write("5. **跟进进度**: 投递后主动查询进度\n\n")
            
            f.write("---\n\n")
            f.write("**免责声明**: 本汇总表仅供参考，所有信息来源于公开渠道。\n\n")
            f.write("**祝各位求职顺利，早日拿到心仪的Offer！** 🎉\n")
        
        print(f"\n✓ Markdown报告已生成: {file_path}")
        return file_path
    
    def identify_top_opportunities(self, df):
        """识别Top 3捡漏机会"""
        opportunities = []
        
        # 优先选择重点公司
        for company in self.key_companies[:10]:
            company_data = df[df['公司名称'] == company]
            if not company_data.empty:
                row = company_data.iloc[0]
                opportunities.append({
                    'emoji': '🌟',
                    'company': company,
                    'highlight': '大厂机会，简历优先筛选',
                    'code': row['内推码/直推邮箱'],
                    'link': row['投递链接/来源'],
                    'note': row['备注']
                })
                if len(opportunities) >= 3:
                    break
        
        # 如果不足3个，补充其他公司
        while len(opportunities) < 3 and len(opportunities) < len(df):
            remaining = df[~df['公司名称'].isin([o['company'] for o in opportunities])]
            if not remaining.empty:
                row = remaining.iloc[0]
                opportunities.append({
                    'emoji': '💡',
                    'company': row['公司名称'],
                    'highlight': '新机会，竞争较小',
                    'code': row['内推码/直推邮箱'],
                    'link': row['投递链接/来源'],
                    'note': row['备注']
                })
            else:
                break
        
        return opportunities
    
    def update_excel(self, filename='每日内推.xlsx'):
        """更新Excel汇总表"""
        if not self.results:
            print("\n没有数据可更新Excel！")
            return
        
        desktop_path = os.path.expanduser('~/Desktop')
        file_path = os.path.join(desktop_path, filename)
        
        # 数据处理
        df = pd.DataFrame(self.results)
        df = df.drop_duplicates(subset=['公司名称', '内推码/直推邮箱'])
        
        # 如果文件已存在，追加数据
        if os.path.exists(file_path):
            try:
                existing_df = pd.read_excel(file_path)
                df = pd.concat([existing_df, df], ignore_index=True)
                df = df.drop_duplicates(subset=['公司名称', '内推码/直推邮箱'], keep='last')
                print(f"\n✓ 已合并现有数据")
            except Exception as e:
                print(f"\n⚠ 读取现有Excel失败，将创建新文件: {str(e)}")
        
        # 保存到Excel
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='内推汇总', index=False)
            
            # 格式化Excel
            self.format_excel(file_path)
            
            print(f"✓ Excel已更新: {file_path}")
            print(f"✓ 共 {len(df)} 条记录")
            
        except Exception as e:
            print(f"\n✗ 更新Excel失败: {str(e)}")
    
    def format_excel(self, file_path):
        """格式化Excel表格"""
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # 设置表头样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 调整列宽
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 40
            ws.column_dimensions['F'].width = 30
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 20
            
            wb.save(file_path)
            print("✓ Excel格式化完成")
            
        except Exception as e:
            print(f"⚠ Excel格式化失败: {str(e)}")
    
    def close(self):
        """关闭浏览器"""
        if self.driver:
            self.driver.quit()
            print("\n✓ 浏览器已关闭")

def main():
    """主函数"""
    scraper = EnhancedReferralScraper()
    
    try:
        # 启动浏览器
        scraper.start_driver()
        
        # 抓取牛客网
        scraper.scrape_nowcoder()
        
        # 生成Markdown报告
        scraper.generate_markdown_report()
        
        # 更新Excel
        scraper.update_excel()
        
        print("\n" + "="*60)
        print("任务完成！")
        print("="*60)
        print(f"\n结束时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"\n生成文件：")
        print(f"  1. 今日内推汇总_{scraper.today}.md (Markdown报告)")
        print(f"  2. 每日内推.xlsx (Excel汇总表)")
        print(f"\n文件位置: ~/Desktop/")
        
    except KeyboardInterrupt:
        print("\n\n⚠ 用户中断执行")
    except Exception as e:
        print(f"\n\n✗ 程序执行出错: {str(e)}")
        import traceback
        traceback.print_exc()
    finally:
        # 关闭浏览器
        scraper.close()

if __name__ == "__main__":
    main()
