#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AceOffer 招聘信息抓取脚本
功能：使用 Playwright 抓取 https://material.aceoffer.cn/recruit 上的招聘信息

安装依赖：
    pip install playwright pandas openpyxl

安装Playwright浏览器：
    playwright install chromium

使用方法：
    python aceoffer_recruit_scraper.py
"""

import asyncio
import random
import re
import time
from datetime import datetime
from typing import List, Dict, Optional
import pandas as pd
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError

# ==================== 配置区域 ====================

# Chrome 用户数据目录路径（请根据实际情况修改）
# 使用独立的用户数据目录，避免与正在运行的 Chrome 冲突
# ⚠️ 注意：使用独立目录需要重新登录网站
CHROME_USER_DATA_DIR = "/Users/changchun/Library/Application Support/Google/Chrome_Scraper"

# 目标URL
TARGET_URL = "https://material.aceoffer.cn/recruit"

# 随机等待时间范围（秒）- 模拟人类操作，防止被反爬（已优化为更快速度）
RANDOM_WAIT_MIN = 0.5
RANDOM_WAIT_MAX = 1.5

# 最大抓取页数（设置为 None 表示抓取所有页）
MAX_PAGES = 20  # 抓取前20页（确保能抓取到100个岗位）

# 每页最大抓取数量（设置为 None 表示抓取所有，测试时可以设置较小值）
MAX_ITEMS_PER_PAGE = None  # 抓取每页所有卡片

# 最大抓取岗位总数（设置为 None 表示不限制）
MAX_TOTAL_ITEMS = 100  # 只抓取前100个岗位

# 是否启用日期过滤（网申截止的岗位通常不需要日期过滤）
ONLY_TODAY_UPDATED = False  # 设置为 True 时启用日期过滤，False 表示抓取所有

# 日期过滤天数（只抓取最近N天更新的岗位）
DATE_FILTER_DAYS = 2  # 设置为 1 表示只抓取今天，设置为 2 表示抓取最近2天，以此类推

# 连续空页数阈值（当启用日期过滤时，如果连续N页没有目标日期范围内的岗位，则停止翻页）
CONSECUTIVE_EMPTY_PAGES_THRESHOLD = 2  # 连续2页没有目标日期范围内的岗位就停止

# Excel文件路径（覆盖更新）
EXCEL_FILE_PATH = "网申截止倒计时公司名单.xlsx"  # 固定文件名，用于覆盖更新

# ==================== CSS选择器配置 ====================
# 使用多种选择器组合，自动尝试匹配

# "网申截止倒计时"标签选择器（需要先点击这个标签）
NET_APPLY_TAB_SELECTOR = "text=网申截止倒计时, button:has-text('网申截止倒计时'), [class*='tab']:has-text('网申截止倒计时')"

# 列表容器选择器（"网申开启"列表的容器）
LIST_CONTAINER_SELECTOR = ".recruit-list, .job-list, [class*='list'], [class*='container'], .content, main"

# 单个招聘卡片选择器（尝试多种可能的选择器）
JOB_CARD_SELECTORS = [
    ".job-card",
    ".recruit-item", 
    "[class*='card']",
    "[class*='item']",
    "[class*='job']",
    "li[class*='item']",
    "div[class*='card']",
    ".el-card",
    ".ant-card"
]

# 公司名称选择器（在卡片内，尝试多种）
COMPANY_NAME_SELECTORS = [
    ".company-name",
    "[class*='company']",
    "[class*='name']",
    "h3, h4, h5",
    ".title",
    "[class*='title']"
]

# 公司类型关键词（从标签中提取）
COMPANY_TYPE_KEYWORDS = ["外资", "央/国企", "内资", "国企", "央企", "外企"]

# 工作地点选择器
LOCATION_SELECTORS = [
    ".location",
    "[class*='location']",
    "[class*='city']",
    "[class*='address']",
    ".address"
]

# 招聘类型关键词（从标签中提取）
RECRUIT_TYPE_KEYWORDS = ["暑期实习", "秋招正式批", "春招", "实习", "校招", "社招"]

# 岗位/标签关键词选择器（提取所有标签）
POSITION_TAG_SELECTORS = [
    ".tag",
    "[class*='tag']",
    "[class*='label']",
    ".label",
    "span[class*='tag']",
    ".el-tag",
    ".ant-tag"
]

# 更新日期选择器
UPDATE_DATE_SELECTORS = [
    ".update-date",
    "[class*='date']",
    "[class*='time']",
    "[class*='update']"
]

# "立即投递"按钮选择器（优先使用文本匹配）
APPLY_BUTTON_SELECTORS = [
    "button:has-text('立即投递')",
    "a:has-text('立即投递')",
    "*:has-text('立即投递')",
    "[class*='apply']",
    "[class*='投递']",
    ".apply-btn",
    ".btn-apply"
]

# 下一页按钮选择器
NEXT_PAGE_SELECTORS = [
    "button:has-text('下一页')",
    "a:has-text('下一页')",
    "button:has-text('Next')",
    ".next-page",
    "[class*='next']",
    ".pagination-next",
    ".el-pagination .btn-next"
]

# 招聘对象选择器（在新打开的页面中，如果存在）
RECRUIT_TARGET_SELECTORS = [
    ".recruit-target",
    "[class*='target']",
    "[class*='对象']",
    "*:has-text('招聘对象')",
    "*:has-text('面向')"
]

# 投递截止时间选择器（在新打开的页面中，如果存在）
DEADLINE_SELECTORS = [
    ".deadline",
    "[class*='deadline']",
    "[class*='截止']",
    "*:has-text('截止时间')",
    "*:has-text('投递截止')",
    "*:has-text('截止日期')"
]

# ==================== 主类 ====================

class AceOfferRecruitScraper:
    def __init__(self):
        """初始化爬虫"""
        self.results: List[Dict] = []
        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None
        
    async def random_wait(self, min_seconds: float = None, max_seconds: float = None):
        """随机等待，模拟人类操作"""
        min_sec = min_seconds or RANDOM_WAIT_MIN
        max_sec = max_seconds or RANDOM_WAIT_MAX
        wait_time = random.uniform(min_sec, max_sec)
        await asyncio.sleep(wait_time)
        
    async def start_browser(self):
        """启动浏览器，加载用户配置文件"""
        print("\n" + "="*60)
        print("AceOffer 招聘信息抓取脚本")
        print("="*60)
        print(f"\n开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("正在启动浏览器...")
        
        self.playwright = await async_playwright().start()
        
        # 启动 Chromium，使用用户数据目录保持登录状态
        print(f"使用用户数据目录: {CHROME_USER_DATA_DIR}")
        
        # 尝试启动浏览器，如果失败则提示
        try:
            self.browser = await self.playwright.chromium.launch_persistent_context(
                user_data_dir=CHROME_USER_DATA_DIR,
                headless=False,  # 设置为 True 可启用无头模式
                channel="chrome",  # 使用系统安装的 Chrome，如果只有 chromium 则改为 None
                args=[
                    '--disable-blink-features=AutomationControlled',
                    '--disable-dev-shm-usage',
                    # 移除代理禁用设置，使用系统代理（TUN模式会自动生效）
                ],
                viewport={'width': 1920, 'height': 1080},
                user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                ignore_https_errors=True  # 忽略HTTPS错误
            )
        except Exception as e:
            if "ProcessSingleton" in str(e) or "profile is already in use" in str(e).lower():
                print("\n" + "="*60)
                print("⚠ 错误：Chrome 用户数据目录已被占用")
                print("="*60)
                print("\n解决方案：")
                print("1. 关闭所有 Chrome 浏览器窗口，然后重新运行脚本")
                print("2. 或者修改脚本中的 CHROME_USER_DATA_DIR 为独立目录")
                print("   （例如：'/Users/changchun/Library/Application Support/Google/Chrome_Scraper'）")
                print("   注意：使用独立目录需要重新登录网站")
                print("="*60)
                raise
            else:
                raise
        
        # 获取第一个页面（persistent context 会自动创建一个页面）
        pages = self.browser.pages
        if pages:
            self.page = pages[0]
        else:
            self.page = await self.browser.new_page()
            
        print("✓ 浏览器启动成功！")
        await self.random_wait(2, 4)
        
    async def navigate_to_target(self):
        """访问目标URL"""
        print(f"\n正在访问: {TARGET_URL}")
        try:
            await self.page.goto(TARGET_URL, wait_until="networkidle", timeout=60000)
            await self.random_wait(3, 5)
            print("✓ 页面加载完成")
        except Exception as e:
            error_msg = str(e)
            if "PROXY" in error_msg or "proxy" in error_msg.lower():
                print(f"\n{'='*60}")
                print("⚠ 代理连接失败")
                print("="*60)
                print("\n解决方案：")
                print("1. 检查Chrome浏览器的代理设置（已打开的浏览器窗口中）")
                print("2. 或者在浏览器地址栏手动访问网站登录后，脚本会继续")
                print("3. 等待10秒后脚本将重试...")
                print("="*60)
                await asyncio.sleep(10)
                # 重试一次
                try:
                    await self.page.goto(TARGET_URL, wait_until="networkidle", timeout=60000)
                    await self.random_wait(3, 5)
                    print("✓ 页面加载完成（重试成功）")
                except Exception as e2:
                    print(f"⚠ 重试仍然失败: {str(e2)}")
                    print("请手动在浏览器中打开网站并登录，然后按回车继续...")
                    # 不抛出异常，让用户有机会手动操作
                    await asyncio.sleep(5)
            else:
                print(f"⚠ 访问页面时出错: {error_msg}")
                raise
            
    async def click_net_apply_tab(self):
        """点击"网申截止倒计时"标签"""
        print("\n检查并点击'网申截止倒计时'标签...")
        try:
            # 尝试多种方式找到并点击"网申截止倒计时"标签
            tab_selectors = [
                "text=网申截止倒计时",
                "button:has-text('网申截止倒计时')",
                "a:has-text('网申截止倒计时')",
                "[class*='tab']:has-text('网申截止倒计时')",
                ".tab:has-text('网申截止倒计时')"
            ]
            
            for selector in tab_selectors:
                try:
                    tab = await self.page.query_selector(selector)
                    if tab:
                        is_visible = await tab.is_visible()
                        if is_visible:
                            # 检查是否已选中（可能需要检查class或aria-selected）
                            classes = await tab.get_attribute("class") or ""
                            if "active" not in classes.lower() and "selected" not in classes.lower():
                                print(f"  找到'网申截止倒计时'标签，正在点击...")
                                await tab.click()
                                await self.random_wait(2, 3)
                                print("  ✓ 已点击'网申截止倒计时'标签")
                            else:
                                print("  ✓ '网申截止倒计时'标签已选中")
                            return True
                except Exception:
                    continue
                    
            print("  ⚠ 未找到'网申截止倒计时'标签，可能已在正确页面")
            return False
        except Exception as e:
            print(f"  ⚠ 点击'网申截止倒计时'标签时出错: {str(e)}")
            return False
            
    async def wait_for_list_loaded(self):
        """等待招聘列表加载完成"""
        print("\n等待招聘列表加载...")
        try:
            # 等待更长时间，确保页面完全加载
            await self.random_wait(3, 5)
            
            # 尝试多个选择器
            for selector in LIST_CONTAINER_SELECTOR.split(", "):
                try:
                    await self.page.wait_for_selector(
                        selector.strip(),
                        timeout=15000,
                        state="visible"
                    )
                    await self.random_wait(2, 3)
                    print("✓ 列表加载完成")
                    break
                except PlaywrightTimeoutError:
                    continue
            
            # 滚动页面，确保所有卡片都加载出来（增加滚动次数）
            print("正在滚动页面以加载所有卡片...")
            for i in range(5):  # 增加滚动次数
                await self.page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                await self.random_wait(1.5, 2.5)
            # 滚动回顶部
            await self.page.evaluate("window.scrollTo(0, 0)")
            await self.random_wait(2, 3)
            print("✓ 页面滚动完成")
        except Exception as e:
            print(f"⚠ 等待列表时出错: {str(e)}")
            
    async def extract_text_with_selectors(self, element, selectors: list, default: str = "") -> str:
        """尝试多个选择器提取文本"""
        for selector in selectors:
            try:
                sub_element = await element.query_selector(selector)
                if sub_element:
                    text = await sub_element.inner_text()
                    if text and text.strip():
                        return text.strip()
            except Exception:
                continue
        return default
        
    async def extract_all_text_with_selectors(self, element, selectors: list, separator: str = ", ") -> str:
        """尝试多个选择器提取所有匹配元素的文本"""
        for selector in selectors:
            try:
                elements = await element.query_selector_all(selector)
                if elements:
                    texts = []
                    for elem in elements:
                        text = await elem.inner_text()
                        if text and text.strip():
                            texts.append(text.strip())
                    if texts:
                        return separator.join(texts)
            except Exception:
                continue
        return ""
        
    def extract_keywords_from_text(self, text: str, keywords: list) -> str:
        """从文本中提取包含关键词的部分"""
        found = []
        for keyword in keywords:
            if keyword in text:
                found.append(keyword)
        return ", ".join(found) if found else ""
        
    async def extract_text_from_element(self, element, selector: str, default: str = "") -> str:
        """从元素中提取文本（兼容旧接口）"""
        return await self.extract_text_with_selectors(element, [selector], default)
        
    async def extract_all_text_from_elements(self, element, selector: str, separator: str = ", ") -> str:
        """从多个元素中提取文本并合并（兼容旧接口）"""
        return await self.extract_all_text_with_selectors(element, [selector], separator)
        
    async def extract_job_info_from_card(self, card_element) -> Dict:
        """从单个招聘卡片中提取基础信息"""
        job_info = {
            '公司名称': '',
            '公司类型': '',
            '工作地点': '',
            '招聘类型': '',
            '招聘对象': '',
            '岗位': '',
            '更新时间': '',
            '投递截止': '',
            '相关链接': ''
        }
        
        try:
            # 获取整个卡片的文本，用于智能提取
            card_text = await card_element.inner_text()
            
            # 提取公司名称（优先从标题元素提取）
            job_info['公司名称'] = await self.extract_text_with_selectors(
                card_element, COMPANY_NAME_SELECTORS
            )
            
            # 如果没找到，尝试从卡片文本中提取
            if not job_info['公司名称'] and card_text:
                lines = [line.strip() for line in card_text.split('\n') if line.strip()]
                # 尝试前几行，找到最可能是公司名称的行
                for line in lines[:10]:  # 只检查前10行
                    # 公司名称通常较长，且不包含特定关键词
                    if len(line) > 5 and len(line) < 150:
                        # 排除明显的标签和按钮文本
                        exclude_keywords = ['立即投递', 'NEW', '校招', '实习', '招聘', '投递', '点击', '查看', '更多', '详情', '>>', '<<']
                        if not any(kw in line for kw in exclude_keywords):
                            # 排除纯数字、纯符号，必须包含中文或英文
                            if any(c.isalpha() or '\u4e00' <= c <= '\u9fff' for c in line):
                                # 优先选择包含"公司"、"企业"、"集团"等关键词的行
                                if any(kw in line for kw in ['公司', '企业', '集团', '银行', '科技', '信息', '有限']):
                                    job_info['公司名称'] = line
                                    break
                                # 如果没有找到包含关键词的，选择第一个符合条件的
                                if not job_info['公司名称']:
                                    job_info['公司名称'] = line
            
            # 提取招聘类型（从标签中提取，或从文本中匹配关键词）
            # 先尝试从标签元素提取
            recruit_type_text = ""
            for selector in POSITION_TAG_SELECTORS:
                try:
                    tags = await card_element.query_selector_all(selector)
                    for tag in tags:
                        tag_text = await tag.inner_text()
                        if tag_text:
                            for keyword in RECRUIT_TYPE_KEYWORDS:
                                if keyword in tag_text:
                                    recruit_type_text = tag_text.strip()
                                    break
                        if recruit_type_text:
                            break
                    if recruit_type_text:
                        break
                except Exception:
                    continue
            
            if not recruit_type_text and card_text:
                # 从整个文本中匹配关键词
                recruit_type_text = self.extract_keywords_from_text(card_text, RECRUIT_TYPE_KEYWORDS)
            
            job_info['招聘类型'] = recruit_type_text
            
            # 提取公司类型（从标签中提取关键词）
            company_type_text = ""
            if card_text:
                company_type_text = self.extract_keywords_from_text(card_text, COMPANY_TYPE_KEYWORDS)
            job_info['公司类型'] = company_type_text
            
            # 提取工作地点（尝试多个选择器）
            job_info['工作地点'] = await self.extract_all_text_with_selectors(
                card_element, LOCATION_SELECTORS, " "
            )
            # 如果没找到，尝试从文本中提取（常见城市名）
            if not job_info['工作地点'] and card_text:
                # 常见城市列表
                cities = ["北京", "上海", "广州", "深圳", "杭州", "南京", "苏州", "成都", "重庆", 
                         "武汉", "西安", "天津", "青岛", "大连", "宁波", "无锡", "长沙", "郑州",
                         "济南", "合肥", "福州", "厦门", "昆明", "南宁", "香港", "台北", "嘉兴"]
                found_cities = [city for city in cities if city in card_text]
                if found_cities:
                    job_info['工作地点'] = " ".join(found_cities)
            
            # 提取岗位/标签关键词（排除已提取的招聘类型和公司类型）
            all_tags = await self.extract_all_text_with_selectors(
                card_element, POSITION_TAG_SELECTORS, ", "
            )
            if all_tags:
                # 过滤掉招聘类型和公司类型
                tags_list = [tag.strip() for tag in all_tags.split(",")]
                filtered_tags = []
                for tag in tags_list:
                    is_recruit_type = any(kw in tag for kw in RECRUIT_TYPE_KEYWORDS)
                    is_company_type = any(kw in tag for kw in COMPANY_TYPE_KEYWORDS)
                    if not is_recruit_type and not is_company_type:
                        filtered_tags.append(tag)
                job_info['岗位'] = ", ".join(filtered_tags)
            
            # 提取更新日期
            job_info['更新时间'] = await self.extract_text_with_selectors(
                card_element, UPDATE_DATE_SELECTORS
            )
            
        except Exception as e:
            print(f"    提取基础信息时出错: {str(e)}")
            import traceback
            traceback.print_exc()
            
        return job_info
        
    async def get_apply_link(self, card_element, job_info: Dict, seen_links: set) -> tuple:
        """点击"立即投递"按钮，获取真实投递链接并提取完整信息
        
        Returns:
            tuple: (apply_link, extracted_info_dict)
        """
        apply_link = ""
        extracted_info = {
            '招聘对象': '',
            '投递截止': '',
            '岗位': '',
            '公司类型': '',
            '工作地点': '',
            '更新时间': ''
        }
        initial_page_count = len(self.browser.pages)
        
        try:
            # 尝试多个选择器查找"立即投递"按钮
            apply_button = None
            for selector in APPLY_BUTTON_SELECTORS:
                try:
                    apply_button = await card_element.query_selector(selector)
                    if apply_button:
                        is_visible = await apply_button.is_visible()
                        if is_visible:
                            break
                        else:
                            apply_button = None
                except Exception:
                    continue
                    
            if not apply_button:
                print(f"    ⚠ 未找到'立即投递'按钮")
                return apply_link, extracted_info
            
            # 点击按钮，等待新标签页打开
            print(f"    点击'立即投递'按钮...")
            async with self.browser.expect_page(timeout=8000) as new_page_info:  # 减少超时时间
                await apply_button.click()
            
            new_page = await new_page_info.value
            
            # 快速等待页面加载（不等待networkidle，只等待DOM加载）
            try:
                await new_page.wait_for_load_state("domcontentloaded", timeout=4000)
            except Exception:
                pass
            
            # 获取新页面的URL
            apply_link = new_page.url
            
            # 检查是否已处理过（避免重复打开）
            if apply_link in seen_links:
                print(f"    ⚠ 链接已处理过，跳过")
                await new_page.close()
                return apply_link, extracted_info
            
            print(f"    ✓ 获取到链接: {apply_link[:80]}...")
            
            # 从页面提取完整信息
            extracted_info = await self.extract_info_from_link_page(new_page)
            
            # 打印提取到的信息
            if extracted_info['招聘对象']:
                print(f"    ✓ 招聘对象: {extracted_info['招聘对象'][:50]}...")
            if extracted_info['投递截止']:
                print(f"    ✓ 投递截止: {extracted_info['投递截止']}")
            if extracted_info['岗位']:
                print(f"    ✓ 岗位: {extracted_info['岗位'][:50]}...")
            
            # 快速关闭新标签页（减少等待时间）
            await new_page.close()
            await asyncio.sleep(0.3)  # 减少等待时间
            
            # 切换回原页面
            await self.page.bring_to_front()
            
        except PlaywrightTimeoutError:
            print(f"    ⚠ 等待新标签页超时")
        except Exception as e:
            print(f"    ⚠ 获取投递链接时出错: {str(e)}")
            # 确保关闭可能打开的新页面
            try:
                pages = self.browser.pages
                if len(pages) > initial_page_count:
                    for p in pages[initial_page_count:]:
                        await p.close()
            except Exception:
                pass
                
        return apply_link, extracted_info
    
    async def extract_info_from_link_page(self, page) -> Dict:
        """从链接页面提取完整信息"""
        extracted_info = {
            '招聘对象': '',
            '投递截止': '',
            '岗位': '',
            '公司类型': '',
            '工作地点': '',
            '更新时间': ''
        }
        
        try:
            # 获取页面文本内容
            page_text = await page.inner_text("body")
            if not page_text:
                return extracted_info
            
            # 1. 提取招聘对象（使用正则表达式更准确）
            recruit_target_patterns = [
                r'招聘对象[：:]\s*([^。\n]{10,200})',
                r'面向[：:]\s*([^。\n]{10,200})',
                r'应聘对象[：:]\s*([^。\n]{10,200})',
            ]
            for pattern in recruit_target_patterns:
                matches = re.findall(pattern, page_text[:5000])
                if matches:
                    match = matches[0].strip()
                    # 清理文本，移除多余空格和换行
                    match = re.sub(r'\s+', ' ', match)
                    if len(match) > 10 and len(match) < 300:
                        extracted_info['招聘对象'] = match[:300]
                        break
            
            # 2. 提取投递截止时间
            deadline_patterns = [
                r'截止[至到]?\s*[:：]?\s*(\d{4}[-/]\d{1,2}[-/]\d{1,2})',
                r'投递截止[：:]?\s*(\d{4}[-/]\d{1,2}[-/]\d{1,2})',
                r'报名截止[：:]?\s*(\d{4}[-/]\d{1,2}[-/]\d{1,2})',
                r'(\d{4}[-/]\d{1,2}[-/]\d{1,2}).*?截止',
            ]
            for pattern in deadline_patterns:
                matches = re.findall(pattern, page_text[:5000])
                if matches:
                    extracted_info['投递截止'] = matches[0].strip()
                    break
            
            # 3. 提取岗位信息（更精确）
            position_patterns = [
                r'岗位[类别名称]?[：:]\s*([^。\n]{5,50})',
                r'职位[名称]?[：:]\s*([^。\n]{5,50})',
                r'招聘岗位[：:]\s*([^。\n]{5,50})',
            ]
            found_positions = []
            for pattern in position_patterns:
                matches = re.findall(pattern, page_text[:3000])
                for match in matches[:3]:  # 最多取3个
                    match = match.strip()
                    if match and len(match) > 3 and len(match) < 50:
                        # 排除明显的无关内容
                        exclude_words = ['要求', '条件', '不得', '报考', '对象']
                        if not any(word in match for word in exclude_words):
                            found_positions.append(match)
            if found_positions:
                extracted_info['岗位'] = ", ".join(found_positions[:5])[:200]
            
            # 4. 提取工作地点
            cities = ["北京", "上海", "广州", "深圳", "杭州", "南京", "苏州", "成都", "重庆", 
                     "武汉", "西安", "天津", "青岛", "大连", "宁波", "无锡", "长沙", "郑州",
                     "济南", "合肥", "福州", "厦门", "昆明", "南宁", "香港", "台北", "嘉兴"]
            found_cities = [city for city in cities if city in page_text[:3000]]
            if found_cities:
                extracted_info['工作地点'] = " ".join(found_cities[:10])
                    
        except Exception as e:
            pass
        
        return extracted_info
        
    async def scrape_current_page(self) -> int:
        """抓取当前页的所有招聘信息"""
        print("\n" + "-"*60)
        print("开始抓取当前页...")
        
        # 用于跟踪已处理的链接，避免重复
        seen_links = set()
        
        try:
            # 优先通过"立即投递"按钮定位真实的招聘卡片
            cards = []
            print("正在通过'立即投递'按钮定位招聘卡片...")
            try:
                # 先统计有多少个"立即投递"按钮
                apply_buttons = await self.page.query_selector_all("text=立即投递")
                print(f"  页面上共有 {len(apply_buttons)} 个'立即投递'按钮")
                
                if apply_buttons:
                    # 直接使用XPath找到按钮的父容器
                    xpath = "//button[contains(text(), '立即投递')]/ancestor::*[self::li or (self::div and position()>2)][1] | //a[contains(text(), '立即投递')]/ancestor::*[self::li or (self::div and position()>2)][1]"
                    all_cards = await self.page.query_selector_all(f"xpath={xpath}")
                    print(f"  通过XPath找到 {len(all_cards)} 个可能的卡片容器")
                    seen_elements = set()
                    for card in all_cards:
                        try:
                            # 验证不是按钮本身：检查文本长度
                            card_text = await card.inner_text()
                            if card_text and len(card_text) > 10:  # 确保是容器，不是按钮
                                element_id = await card.evaluate("el => el.outerHTML.substring(0, 200)")
                                if element_id and element_id not in seen_elements:
                                    seen_elements.add(element_id)
                                    cards.append(card)
                        except Exception:
                            continue
                    
                    print(f"✓ 找到 {len(cards)} 个唯一的招聘卡片")
            except Exception as e:
                print(f"    定位卡片时出错: {str(e)}")
                import traceback
                traceback.print_exc()
            
            # 如果XPath方法失败，尝试其他选择器
            if not cards:
                print("尝试使用其他选择器...")
                for selector in JOB_CARD_SELECTORS:
                    try:
                        cards = await self.page.query_selector_all(selector)
                        if cards and len(cards) > 0:
                            # 过滤：只保留包含"立即投递"按钮的卡片
                            filtered_cards = []
                            for card in cards:
                                try:
                                    has_apply = await card.query_selector("*:has-text('立即投递')")
                                    if has_apply:
                                        filtered_cards.append(card)
                                except Exception:
                                    pass
                            if filtered_cards:
                                cards = filtered_cards
                                print(f"✓ 使用选择器 '{selector}' 找到 {len(cards)} 个有效的招聘卡片")
                                break
                    except Exception:
                        continue
                
                # 如果还是没找到，尝试通过"立即投递"按钮的父元素来定位卡片
                if not cards and apply_buttons:
                    print("尝试通过'立即投递'按钮的父元素定位卡片...")
                    try:
                        temp_cards = []
                        for idx, btn in enumerate(apply_buttons[:100], 1):  # 限制处理前100个按钮
                            try:
                                # 使用 evaluate 找到最近的父容器
                                parent = await btn.evaluate_handle("""
                                    (btn) => {
                                        let el = btn;
                                        // 向上查找最多15层，找到包含足够文本的容器
                                        for (let i = 0; i < 15; i++) {
                                            el = el.parentElement;
                                            if (!el) break;
                                            let text = el.innerText || '';
                                            let classes = el.className || '';
                                            // 查找包含足够文本的容器（可能是卡片）
                                            if (text.length > 30) {
                                                // 检查是否是合适的容器
                                                if (classes.includes('item') || classes.includes('card') || 
                                                    classes.includes('job') || classes.includes('recruit') ||
                                                    el.tagName === 'LI' || el.tagName === 'DIV' || el.tagName === 'ARTICLE') {
                                                    return el;
                                                }
                                            }
                                        }
                                        // 如果没找到，返回按钮的父元素的父元素
                                        return btn.parentElement?.parentElement || btn.parentElement;
                                    }
                                """)
                                if parent:
                                    parent_elem = await parent.as_element()
                                    card_text = await parent_elem.inner_text()
                                    if card_text and len(card_text) > 20:  # 确保是卡片而不是按钮本身
                                        temp_cards.append(parent_elem)
                            except Exception as e:
                                continue
                        
                        if temp_cards:
                            # 去重：使用更精确的去重方法
                            seen_cards = set()
                            unique_cards = []
                            for card in temp_cards:
                                try:
                                    # 使用卡片的文本内容前100个字符作为唯一标识
                                    card_text = await card.inner_text()
                                    card_id = card_text[:100] if card_text else ""
                                    if card_id and card_id not in seen_cards:
                                        seen_cards.add(card_id)
                                        unique_cards.append(card)
                                except:
                                    continue
                            cards = unique_cards
                            print(f"✓ 通过按钮父元素找到 {len(cards)} 个招聘卡片")
                    except Exception as e:
                        print(f"  通过按钮父元素定位时出错: {str(e)}")
                        import traceback
                        traceback.print_exc()
                    
            if not cards:
                print("⚠ 未找到招聘卡片，尝试最后的方法：直接使用所有包含'立即投递'的元素...")
                # 最后尝试：直接查找所有包含按钮的容器
                try:
                    # 查找所有可能包含按钮的容器
                    all_containers = await self.page.query_selector_all("div, li, article")
                    potential_cards = []
                    for container in all_containers:
                        try:
                            container_text = await container.inner_text()
                            # 如果容器包含"立即投递"且有足够的内容，可能是卡片
                            if container_text and "立即投递" in container_text and len(container_text) > 30:
                                # 检查是否包含公司名称或岗位信息的关键词
                                if any(keyword in container_text for keyword in ["公司", "招聘", "岗位", "职位", "实习", "校招"]):
                                    potential_cards.append(container)
                        except:
                            continue
                    
                    if potential_cards:
                        # 去重
                        seen_texts = set()
                        unique_cards = []
                        for card in potential_cards:
                            try:
                                card_text = await card.inner_text()
                                text_id = card_text[:100] if card_text else ""
                                if text_id and text_id not in seen_texts:
                                    seen_texts.add(text_id)
                                    unique_cards.append(card)
                            except:
                                continue
                        
                        if unique_cards:
                            cards = unique_cards[:100]  # 限制为100个
                            print(f"✓ 通过通用方法找到 {len(cards)} 个可能的招聘卡片")
                except Exception as e:
                    print(f"  通用方法也失败: {str(e)}")
                
                if not cards:
                    print("⚠ 未找到招聘卡片，请检查页面是否正常加载")
                    # 尝试打印页面HTML的一部分来调试
                    try:
                        page_content = await self.page.content()
                        if "立即投递" in page_content:
                            print("  ✓ 页面中包含'立即投递'文本，但未找到卡片元素")
                            print("  💡 提示：可能需要调整选择器或等待更长时间")
                        else:
                            print("  ⚠ 页面中未找到'立即投递'文本")
                    except:
                        pass
                    return 0
                
            # 如果设置了每页最大数量，只处理前N个
            if MAX_ITEMS_PER_PAGE and len(cards) > MAX_ITEMS_PER_PAGE:
                print(f"⚠ 当前页有 {len(cards)} 个卡片，但限制每页最多抓取 {MAX_ITEMS_PER_PAGE} 个")
                cards = cards[:MAX_ITEMS_PER_PAGE]
                
            # 遍历每个卡片
            for idx, card in enumerate(cards, 1):
                print(f"\n处理第 {idx}/{len(cards)} 个卡片...")
                
                # 验证卡片是否有"立即投递"按钮（确保是有效的招聘卡片）
                try:
                    has_apply_btn = await card.query_selector("*:has-text('立即投递')")
                    if not has_apply_btn:
                        print(f"  ⚠ 跳过：未找到'立即投递'按钮，可能不是有效的招聘卡片")
                        continue
                except Exception:
                    print(f"  ⚠ 跳过：无法验证卡片有效性")
                    continue
                
                # 提取基础信息（先提取，用于日期过滤）
                job_info = await self.extract_job_info_from_card(card)
                
                # 日期过滤：如果启用了日期过滤，先检查日期再决定是否点击按钮
                if ONLY_TODAY_UPDATED:
                    update_time = job_info.get('更新时间', '')
                    if not self.is_recent_days_updated(update_time, days=DATE_FILTER_DAYS):
                        print(f"  ⚠ 跳过：不在最近{DATE_FILTER_DAYS}天内的岗位（更新时间: {update_time or '无'}）")
                        continue
                
                # 获取投递链接并提取完整信息
                apply_link, extracted_info = await self.get_apply_link(card, job_info, seen_links)
                
                # 检查链接是否已处理过（去重）
                if apply_link and apply_link in seen_links:
                    print(f"  ⚠ 跳过：链接已处理过（重复）")
                    continue
                
                # 更新信息：优先使用从链接页面提取的信息，如果为空则使用卡片信息
                job_info['相关链接'] = apply_link
                
                # 补充空字段：如果job_info中的字段为空，使用extracted_info中的值
                if not job_info['招聘对象'] and extracted_info.get('招聘对象'):
                    job_info['招聘对象'] = extracted_info['招聘对象']
                if not job_info['投递截止'] and extracted_info.get('投递截止'):
                    job_info['投递截止'] = extracted_info['投递截止']
                if not job_info['岗位'] and extracted_info.get('岗位'):
                    job_info['岗位'] = extracted_info['岗位']
                if not job_info['公司类型'] and extracted_info.get('公司类型'):
                    job_info['公司类型'] = extracted_info['公司类型']
                if not job_info['工作地点'] and extracted_info.get('工作地点'):
                    job_info['工作地点'] = extracted_info['工作地点']
                if not job_info['更新时间'] and extracted_info.get('更新时间'):
                    job_info['更新时间'] = extracted_info['更新时间']
                
                # 验证数据有效性：至少要有公司名称或相关链接
                if not job_info['公司名称'] and not job_info['相关链接']:
                    print(f"  ⚠ 跳过：数据无效（无公司名称且无链接）")
                    continue
                
                # 日期过滤已经在点击按钮之前完成，这里不需要再次检查
                
                # 记录已处理的链接
                if apply_link:
                    seen_links.add(apply_link)
                
                # 保存结果
                self.results.append(job_info)
                print(f"  ✓ 公司: {job_info['公司名称'] or '(未提取)'}, 类型: {job_info['招聘类型'] or '(未提取)'}, 链接: {'已获取' if job_info['相关链接'] else '未获取'}")
                
                # 检查是否达到最大抓取数量
                if MAX_TOTAL_ITEMS and len(self.results) >= MAX_TOTAL_ITEMS:
                    print(f"\n⚠ 已达到最大抓取数量限制 ({MAX_TOTAL_ITEMS})，停止抓取")
                    return len(self.results)
                
                # 短暂等待，避免请求过快
                await asyncio.sleep(0.3)
                
            print(f"\n✓ 当前页抓取完成，共 {len(cards)} 条记录")
            return len(cards)
            
        except Exception as e:
            print(f"⚠ 抓取当前页时出错: {str(e)}")
            import traceback
            traceback.print_exc()
            return 0
            
    async def has_next_page(self) -> bool:
        """检查是否有下一页"""
        try:
            # 尝试多个选择器
            for selector in NEXT_PAGE_SELECTORS:
                try:
                    next_button = await self.page.query_selector(selector)
                    if next_button:
                        is_visible = await next_button.is_visible()
                        if is_visible:
                            # 检查按钮是否可点击（未被禁用）
                            is_disabled = await next_button.get_attribute("disabled")
                            if not is_disabled:
                                return True
                except Exception:
                    continue
            return False
        except Exception:
            return False
            
    async def go_to_next_page(self) -> bool:
        """点击下一页按钮"""
        try:
            # 尝试多个选择器
            next_button = None
            for selector in NEXT_PAGE_SELECTORS:
                try:
                    next_button = await self.page.query_selector(selector)
                    if next_button:
                        is_visible = await next_button.is_visible()
                        if is_visible:
                            is_disabled = await next_button.get_attribute("disabled")
                            if not is_disabled:
                                break
                        else:
                            next_button = None
                except Exception:
                    continue
                    
            if not next_button:
                return False
                
            print("\n" + "-"*60)
            print("点击'下一页'按钮...")
            await next_button.click()
            await self.random_wait(3, 5)
            
            # 等待列表重新加载
            await self.wait_for_list_loaded()
            
            return True
            
        except Exception as e:
            print(f"⚠ 翻页时出错: {str(e)}")
            return False
            
    def is_recent_days_updated(self, update_date_str: str, days: int = DATE_FILTER_DAYS) -> bool:
        """检查更新日期是否在最近N天内"""
        if not update_date_str or update_date_str == '' or update_date_str == '无':
            # 如果没有更新日期，假设可能是最新开启的，包含进来
            return True
        
        try:
            from datetime import timedelta
            today = datetime.now()
            cutoff_date = today - timedelta(days=days-1)  # days-1 因为包含今天
            
            # 尝试多种日期格式
            date_formats = ['%Y-%m-%d', '%Y/%m/%d', '%Y年%m月%d日', '%m-%d', '%m/%d', '%m月%d日']
            for fmt in date_formats:
                try:
                    parsed_date = datetime.strptime(update_date_str.strip(), fmt)
                    # 如果只有月日，需要补充年份（假设是今年）
                    if fmt in ['%m-%d', '%m/%d', '%m月%d日']:
                        parsed_date = parsed_date.replace(year=today.year)
                        # 如果日期已经过了今年，可能是去年的
                        if parsed_date > today:
                            parsed_date = parsed_date.replace(year=today.year - 1)
                    
                    # 检查日期是否在范围内
                    if parsed_date.date() >= cutoff_date.date():
                        return True
                except:
                    continue
            return False
        except:
            return False
    
    def is_today_updated(self, update_date_str: str) -> bool:
        """检查更新日期是否为今天（兼容旧方法）"""
        return self.is_recent_days_updated(update_date_str, days=1)
    
    async def scrape_all_pages(self):
        """抓取所有页面的招聘信息"""
        page_num = 1
        today_updated_count = 0
        skipped_count = 0
        consecutive_empty_pages = 0  # 连续没有当天更新岗位的页数
        
        if MAX_TOTAL_ITEMS:
            print(f"\n{'='*60}")
            print(f"⚠ 抓取限制：最多抓取 {MAX_TOTAL_ITEMS} 个岗位")
            print(f"{'='*60}\n")
        
        if ONLY_TODAY_UPDATED:
            print(f"\n{'='*60}")
            print(f"⚠ 日期过滤模式：只抓取最近 {DATE_FILTER_DAYS} 天更新的岗位")
            print(f"⚠ 智能翻页：连续 {CONSECUTIVE_EMPTY_PAGES_THRESHOLD} 页没有目标日期范围内的岗位将自动停止")
            print(f"{'='*60}\n")
        
        while True:
            print(f"\n{'='*60}")
            print(f"第 {page_num} 页")
            print(f"{'='*60}")
            
            # 记录抓取前的数量
            results_before = len(self.results)
            today_count_before = sum(1 for r in self.results if self.is_recent_days_updated(r.get('更新时间', ''), days=DATE_FILTER_DAYS)) if ONLY_TODAY_UPDATED else 0
            
            # 抓取当前页
            count = await self.scrape_current_page()
            
            if count == 0:
                print("⚠ 当前页没有数据，停止抓取")
                break
            
            # 如果启用了日期过滤，统计最近N天更新的数量
            if ONLY_TODAY_UPDATED:
                today_count_after = sum(1 for r in self.results if self.is_recent_days_updated(r.get('更新时间', ''), days=DATE_FILTER_DAYS))
                today_added_this_page = today_count_after - today_count_before
                
                if today_added_this_page > 0:
                    print(f"  当前页最近{DATE_FILTER_DAYS}天更新的岗位: {today_added_this_page} 条")
                    today_updated_count = today_count_after
                    skipped_count = len(self.results) - today_updated_count
                    consecutive_empty_pages = 0  # 重置连续空页计数
                else:
                    consecutive_empty_pages += 1
                    print(f"  当前页没有最近{DATE_FILTER_DAYS}天更新的岗位（连续 {consecutive_empty_pages} 页）")
                    
                    # 如果连续N页都没有目标日期范围内的岗位，停止翻页
                    if consecutive_empty_pages >= CONSECUTIVE_EMPTY_PAGES_THRESHOLD:
                        print(f"\n⚠ 连续 {consecutive_empty_pages} 页没有最近{DATE_FILTER_DAYS}天更新的岗位，停止翻页")
                        break
            else:
                # 未启用日期过滤时，正常统计
                today_updated_count = len(self.results)
            
            # 检查是否达到最大抓取数量
            if MAX_TOTAL_ITEMS and len(self.results) >= MAX_TOTAL_ITEMS:
                print(f"\n已达到最大抓取数量限制 ({MAX_TOTAL_ITEMS})，停止抓取")
                break
            
            # 检查是否达到最大页数
            if MAX_PAGES and page_num >= MAX_PAGES:
                print(f"\n已达到最大页数限制 ({MAX_PAGES})，停止抓取")
                break
                
            # 检查是否有下一页
            if not await self.has_next_page():
                print("\n✓ 已到达最后一页")
                break
                
            # 翻页
            if not await self.go_to_next_page():
                print("\n⚠ 无法翻页，停止抓取")
                break
                
            page_num += 1
        
        if ONLY_TODAY_UPDATED:
            print(f"\n{'='*60}")
            print(f"日期过滤统计:")
            print(f"  最近{DATE_FILTER_DAYS}天更新的岗位: {today_updated_count} 条")
            print(f"  已跳过非最近{DATE_FILTER_DAYS}天更新的岗位: {skipped_count} 条")
            print(f"{'='*60}")
            
    async def save_to_excel(self, overwrite: bool = False):
        """保存数据到Excel文件
        
        Args:
            overwrite: 是否为覆盖更新模式（覆盖现有文件，只保留今天的数据）
        """
        import os
        
        if not self.results:
            print("\n⚠ 没有数据可保存")
            return
            
        print(f"\n{'='*60}")
        if overwrite:
            print(f"覆盖更新Excel文件（只保留最近{DATE_FILTER_DAYS}天更新的岗位）...")
        else:
            print("保存数据到Excel...")
        print(f"{'='*60}")
        
        # 创建DataFrame
        new_df = pd.DataFrame(self.results)
        
        # 确保所有必需的列都存在（如果不存在则创建空列）
        columns_order = [
            '公司名称', '公司类型', '工作地点', '招聘类型', 
            '招聘对象', '岗位', '更新时间', '投递截止', '相关链接'
        ]
        
        # 确保所有列都存在
        for col in columns_order:
            if col not in new_df.columns:
                new_df[col] = ''
        
        # 按指定顺序排列列
        new_df = new_df[columns_order]
        
        # 数据清理：将空字符串替换为NaN，然后过滤掉完全空的行
        new_df = new_df.replace('', pd.NA)
        
        # 过滤掉没有公司名称且没有相关链接的无效行
        new_df = new_df[new_df['公司名称'].notna() | new_df['相关链接'].notna()]
        
        # 将NaN替换回空字符串，保持Excel格式整洁
        new_df = new_df.fillna('')
        
        # 数据验证和清理
        print(f"\n新抓取数据: {len(new_df)} 条记录")
        
        # 移除公司名称为空且相关链接也为空的记录
        valid_new_df = new_df[(new_df['公司名称'] != '') | (new_df['相关链接'] != '')]
        print(f"过滤无效记录后: {len(valid_new_df)} 条有效记录")
        
        # 清理空值：将空字符串和NaN统一处理
        for col in valid_new_df.columns:
            valid_new_df[col] = valid_new_df[col].replace('', pd.NA)
            valid_new_df[col] = valid_new_df[col].fillna('')
        
        # 只保留有公司名称的记录（确保数据质量）
        valid_new_df = valid_new_df[valid_new_df['公司名称'] != '']
        print(f"最终新记录: {len(valid_new_df)} 条（仅保留有公司名称的记录）")
        
        # 如果是覆盖更新模式，直接保存今天的数据（覆盖原文件）
        if overwrite:
            # 覆盖模式：只保存今天的数据，覆盖原文件
            # 去重：优先基于相关链接
            before_dedup = len(valid_new_df)
            valid_new_df = valid_new_df.drop_duplicates(subset=['相关链接'], keep='first')
            no_link_df = valid_new_df[valid_new_df['相关链接'] == '']
            has_link_df = valid_new_df[valid_new_df['相关链接'] != '']
            if len(no_link_df) > 0:
                no_link_df = no_link_df.drop_duplicates(subset=['公司名称'], keep='first')
            valid_new_df = pd.concat([has_link_df, no_link_df], ignore_index=True)
            
            after_dedup = len(valid_new_df)
            if before_dedup > after_dedup:
                print(f"去重后: {after_dedup} 条记录（去除了 {before_dedup - after_dedup} 条重复记录）")
            
            print(f"\n覆盖更新模式：将保存 {len(valid_new_df)} 条最近{DATE_FILTER_DAYS}天更新的岗位到文件")
            df = valid_new_df.reset_index(drop=True)
            filename = EXCEL_FILE_PATH
        else:
            # 非增量模式：只处理新数据
            # 去重：优先基于相关链接
            before_dedup = len(valid_new_df)
            valid_new_df = valid_new_df.drop_duplicates(subset=['相关链接'], keep='first')
            no_link_df = valid_new_df[valid_new_df['相关链接'] == '']
            has_link_df = valid_new_df[valid_new_df['相关链接'] != '']
            if len(no_link_df) > 0:
                no_link_df = no_link_df.drop_duplicates(subset=['公司名称'], keep='first')
            valid_new_df = pd.concat([has_link_df, no_link_df], ignore_index=True)
            
            after_dedup = len(valid_new_df)
            if before_dedup > after_dedup:
                print(f"去重后: {after_dedup} 条记录（去除了 {before_dedup - after_dedup} 条重复记录）")
            
            df = valid_new_df.reset_index(drop=True)
            
            # 生成文件名（带时间戳）
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"recruit_data_{timestamp}.xlsx"
        
        # 保存到Excel
        try:
            df.to_excel(filename, index=False, engine='openpyxl')
            
            # 如果是覆盖更新模式，给所有行添加颜色标注（因为都是今天更新的）
            if overwrite:
                try:
                    from openpyxl import load_workbook
                    from openpyxl.styles import PatternFill, Font, Alignment
                    
                    # 打开已保存的Excel文件
                    wb = load_workbook(filename)
                    ws = wb.active
                    
                    # 定义颜色：浅绿色背景，深绿色文字
                    fill_color = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')  # 浅绿色
                    header_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')  # 稍深的绿色（表头）
                    font_color = Font(color='1B5E20', bold=True)  # 深绿色文字
                    
                    # 设置表头样式
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = font_color
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 设置数据行样式（除了表头）
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        for cell in row:
                            cell.fill = fill_color
                    
                    # 自动调整列宽
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 60)  # 最大宽度60
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    # 保存文件
                    wb.save(filename)
                    print(f"\n✓ 已添加颜色标注（浅绿色背景）")
                except Exception as e:
                    print(f"  ⚠ 添加颜色标注时出错: {str(e)}，但文件已保存")
            
            print(f"\n✓ 数据已保存到: {filename}")
            print(f"  共 {len(df)} 条记录")
            print(f"\n列名: {', '.join(df.columns.tolist())}")
        except Exception as e:
            print(f"\n⚠ 保存Excel时出错: {str(e)}")
            # 尝试保存为CSV作为备选
            csv_filename = filename.replace('.xlsx', '.csv')
            df.to_csv(csv_filename, index=False, encoding='utf-8-sig')
            print(f"  已保存为CSV格式: {csv_filename}")
            
    async def run(self, overwrite: bool = False):
        """主运行函数
        
        Args:
            overwrite: 是否为覆盖更新模式（覆盖现有文件）
        """
        try:
            # 启动浏览器
            await self.start_browser()
            
            # 访问目标URL
            await self.navigate_to_target()
            
            # 点击"网申截止倒计时"标签
            await self.click_net_apply_tab()
            
            # 等待列表加载
            await self.wait_for_list_loaded()
            
            # 抓取所有页面
            await self.scrape_all_pages()
            
            # 保存数据
            await self.save_to_excel(overwrite=overwrite)
            
            print(f"\n{'='*60}")
            print("抓取完成！")
            print(f"结束时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            print(f"共抓取 {len(self.results)} 条记录")
            print(f"{'='*60}\n")
            
        except Exception as e:
            print(f"\n⚠ 运行过程中出错: {str(e)}")
            import traceback
            traceback.print_exc()
        finally:
            # 关闭浏览器
            if self.browser:
                await self.browser.close()
            if self.playwright:
                await self.playwright.stop()
                
# ==================== 主程序入口 ====================

import argparse

async def main():
    """主函数"""
    # 解析命令行参数
    parser = argparse.ArgumentParser(description='AceOffer校招信息抓取脚本')
    parser.add_argument('--overwrite', action='store_true', help='覆盖现有文件，使用固定文件名')
    args = parser.parse_args()
    
    scraper = AceOfferRecruitScraper()
    await scraper.run(overwrite=args.overwrite)

if __name__ == "__main__":
    asyncio.run(main())

